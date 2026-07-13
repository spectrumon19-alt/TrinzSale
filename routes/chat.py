import os
import re
import io
import json
import logging
from datetime import datetime, date

from flask import Blueprint, jsonify, request, send_file
from psycopg2.extras import RealDictCursor

from db import get_db_connection, release_db_connection
from auth import token_required, permission_required

logger = logging.getLogger(__name__)
chat_bp = Blueprint('chat', __name__)

# ── Schema context fed to Claude ───────────────────────────────────────────────
SCHEMA_CONTEXT = """
You are an expert SQL analyst for a Point-of-Sale (POS) system using PostgreSQL.

DATABASE TABLES:

1. products (product_id PK, name, pack_size, sku UNIQUE, gst_rate DECIMAL, purchase_rate DECIMAL,
   selling_rate DECIMAL, status ['Active'|'Inactive'], barcode, hsn_code)

2. inventory (product_id PK→products, stock_quantity INTEGER)

3. sales_invoices (invoice_id PK, invoice_number UNIQUE, invoice_date TIMESTAMP, customer_name,
   customer_contact, customer_mobile, user_id→users, mode_of_payment, upi_transaction_id,
   total_amount DECIMAL, total_gst DECIMAL, discount_percentage DECIMAL, discount_amount DECIMAL,
   status ['Completed'|'Cancelled'], irn VARCHAR(64), irn_generated_at TIMESTAMP,
   qr_data TEXT, einvoice_status ['pending'|'generated'])

4. sales_invoice_items (item_id PK, invoice_id→sales_invoices, product_id→products,
   quantity INTEGER, rate_at_sale DECIMAL, gst_rate_at_sale DECIMAL,
   exclusive_gst_amount DECIMAL, sgst DECIMAL, cgst DECIMAL,
   total_line_amount DECIMAL, discount_percentage DECIMAL)

5. purchase_orders (purchase_order_id PK, supplier_id→suppliers, supplier_name,
   purchase_order_number UNIQUE, purchase_date TIMESTAMP, user_id→users,
   total_amount DECIMAL, status ['Completed'|'Cancelled'])

6. purchase_order_items (item_id PK, purchase_order_id→purchase_orders, product_id→products,
   quantity INTEGER, purchase_rate DECIMAL, gst_rate DECIMAL,
   taxable_value DECIMAL, sgst DECIMAL, cgst DECIMAL, total_amount DECIMAL)

7. suppliers (supplier_id PK, supplier_name, supplier_gst_number UNIQUE, contact_person,
   email, address, mobile, bank_name, bank_account_number, ifsc_code, current_balance DECIMAL)

8. supplier_transactions (transaction_id PK, supplier_id→suppliers,
   transaction_type ['credit'|'debit'], amount DECIMAL, previous_balance DECIMAL,
   new_balance DECIMAL, purchase_order_number, note, created_at TIMESTAMP)

9. credit_customers (customer_id PK, customer_code UNIQUE, name, mobile, email,
   address, current_balance DECIMAL, created_at TIMESTAMP)

10. credit_transactions (transaction_id PK, customer_id→credit_customers,
    transaction_type ['credit'|'debit'], amount DECIMAL, invoice_no,
    note, previous_balance DECIMAL, created_at TIMESTAMP)

   Sign convention (Tally): credit_customers.current_balance is NEGATIVE when the
   customer OWES the shop (receivable); a credit sale posts a 'debit'. So
   "customers who owe / receivables" = rows WHERE current_balance < 0, and the
   amount owed = -current_balance. A POSITIVE balance is a customer advance.
   For suppliers it is the opposite: current_balance > 0 means the shop OWES the
   supplier (payable); a credit purchase posts a 'credit'.

11. users (user_id PK, username UNIQUE, role ['Admin'|'Cashier'|'Manager'|'Super Admin'],
    full_name, email, mobile)

12. sales_returns (return_id PK, return_number UNIQUE, original_invoice_id→sales_invoices,
    return_date DATE, customer_name, customer_mobile, return_reason, refund_method,
    subtotal DECIMAL, total_gst DECIMAL, total_amount DECIMAL, status, user_id→users)

13. sales_return_items (return_item_id PK, return_id→sales_returns, product_id→products,
    product_name, sku, quantity INTEGER, rate_at_return DECIMAL, gst_rate DECIMAL,
    exclusive_gst_amount DECIMAL, sgst DECIMAL, cgst DECIMAL, total_line_amount DECIMAL)

KEY FACTS:
- All monetary values are in Indian Rupees (INR).
- sales_invoices.total_amount is the GRAND TOTAL (inclusive of GST).
- sales_invoice_items.exclusive_gst_amount is the TAXABLE value for that line item.
- GST = sgst + cgst on each line item.
- mode_of_payment typical values: 'Cash', 'UPI', 'Card', 'Credit'.
- TODAY in PostgreSQL: CURRENT_DATE. This month: DATE_TRUNC('month', CURRENT_DATE).
- Financial year starts April 1. Current FY: if EXTRACT(month FROM CURRENT_DATE) >= 4
  then DATE_TRUNC('year', CURRENT_DATE) + INTERVAL '3 months'
  else DATE_TRUNC('year', CURRENT_DATE) - INTERVAL '9 months'.
- Always filter sales_invoices by status = 'Completed' unless asked otherwise.
- Gross margin = (exclusive_gst_amount - purchase_rate * quantity) / exclusive_gst_amount * 100.

WORKED EXAMPLES (question → correct SQL):

Q: "What were total sales today?"
SELECT COALESCE(SUM(total_amount), 0) AS total_sales,
       COUNT(*) AS invoice_count
FROM sales_invoices
WHERE status = 'Completed' AND invoice_date::date = CURRENT_DATE
LIMIT 500;

Q: "Top 5 selling products this month"
SELECT p.name AS product_name,
       SUM(sii.quantity) AS units_sold,
       SUM(sii.total_line_amount) AS revenue
FROM sales_invoice_items sii
JOIN sales_invoices si ON si.invoice_id = sii.invoice_id
JOIN products p ON p.product_id = sii.product_id
WHERE si.status = 'Completed'
  AND si.invoice_date >= DATE_TRUNC('month', CURRENT_DATE)
GROUP BY p.name
ORDER BY units_sold DESC
LIMIT 5;

Q: "Which customers owe us the most money?" (receivables — note the sign convention)
SELECT COALESCE(name, '-') AS customer_name,
       COALESCE(mobile, '-') AS mobile,
       -current_balance AS amount_owed
FROM credit_customers
WHERE current_balance < 0
ORDER BY current_balance ASC
LIMIT 500;

Q: "How much do we owe our suppliers?" (payables — opposite sign convention from customers)
SELECT COALESCE(supplier_name, '-') AS supplier_name,
       current_balance AS amount_payable
FROM suppliers
WHERE current_balance > 0
ORDER BY current_balance DESC
LIMIT 500;

Q: "GST collected this financial year"
SELECT COALESCE(SUM(total_gst), 0) AS total_gst_collected
FROM sales_invoices
WHERE status = 'Completed'
  AND invoice_date >= (
    CASE WHEN EXTRACT(month FROM CURRENT_DATE) >= 4
      THEN DATE_TRUNC('year', CURRENT_DATE) + INTERVAL '3 months'
      ELSE DATE_TRUNC('year', CURRENT_DATE) - INTERVAL '9 months'
    END)
LIMIT 500;

Q: "Sales made on credit this week, by customer"
SELECT COALESCE(customer_name, '-') AS customer_name,
       COALESCE(customer_mobile, '-') AS mobile,
       SUM(total_amount) AS credit_sales_total
FROM sales_invoices
WHERE status = 'Completed'
  AND mode_of_payment = 'Credit'
  AND invoice_date >= DATE_TRUNC('week', CURRENT_DATE)
GROUP BY customer_name, customer_mobile
ORDER BY credit_sales_total DESC
LIMIT 500;

Q: "Products low on stock (below 10 units)"
SELECT p.name AS product_name,
       COALESCE(p.sku, '-') AS sku,
       i.stock_quantity
FROM inventory i
JOIN products p ON p.product_id = i.product_id
WHERE i.stock_quantity < 10 AND p.status = 'Active'
ORDER BY i.stock_quantity ASC
LIMIT 500;

Q: "Gross margin by product last month"
SELECT p.name AS product_name,
       ROUND(AVG((sii.exclusive_gst_amount - p.purchase_rate * sii.quantity)
            / NULLIF(sii.exclusive_gst_amount, 0) * 100), 2) AS gross_margin_pct
FROM sales_invoice_items sii
JOIN sales_invoices si ON si.invoice_id = sii.invoice_id
JOIN products p ON p.product_id = sii.product_id
WHERE si.status = 'Completed'
  AND si.invoice_date >= DATE_TRUNC('month', CURRENT_DATE) - INTERVAL '1 month'
  AND si.invoice_date <  DATE_TRUNC('month', CURRENT_DATE)
GROUP BY p.name
ORDER BY gross_margin_pct DESC
LIMIT 500;
"""

SYSTEM_SQL = (
    "You are a SQL expert. Given the database schema and a user question, write a single "
    "PostgreSQL SELECT query. Rules:\n"
    "1. Output ONLY the SQL query — no markdown fences, no explanation.\n"
    "2. Only SELECT statements are allowed. No INSERT, UPDATE, DELETE, DROP, TRUNCATE, ALTER, "
    "GRANT, REVOKE, EXEC, EXECUTE, COPY, CALL.\n"
    "3. Always alias columns with readable names (e.g. SUM(total_amount) AS total_sales).\n"
    "4. For text columns add COALESCE(col, '-') to avoid NULLs in output.\n"
    "5. Limit results to 500 rows using LIMIT 500.\n"
    "6. For 'top N' questions use ORDER BY ... DESC LIMIT N.\n"
    "7. For 'today' use CURRENT_DATE; for 'this month' use DATE_TRUNC('month', CURRENT_DATE); "
    "for 'this year' use current financial year bounds.\n"
    "8. Format amounts as-is (raw numbers); the UI will format them.\n"
    "\nSchema:\n" + SCHEMA_CONTEXT
)

SYSTEM_FORMAT = (
    "You are a business analyst for a retail POS. Given a user question, the SQL that was run, "
    "and the raw result data (JSON), write a concise, insightful business answer in 2-5 sentences. "
    "Use Indian number formatting (lakh, crore) for large amounts. "
    "Highlight the single most important finding. "
    "Do NOT show or repeat the SQL. Do NOT use markdown headers. "
    "If data is empty, say so clearly and suggest why."
)

_BLOCKED = re.compile(
    r'\b(INSERT|UPDATE|DELETE|DROP|TRUNCATE|ALTER|GRANT|REVOKE|EXEC|EXECUTE|COPY|CALL|CREATE)\b',
    re.IGNORECASE
)

def _validate_sql(sql: str):
    m = _BLOCKED.search(sql)
    if m:
        raise ValueError(f"Disallowed SQL keyword: {m.group()}")
    if not re.search(r'\bSELECT\b', sql, re.IGNORECASE):
        raise ValueError("Only SELECT queries are permitted.")


def _run_sql(sql: str):
    # Defense in depth: the AI-generated SQL runs inside a server-enforced
    # READ ONLY transaction, so even if the keyword blocklist in _validate_sql
    # is bypassed, Postgres rejects any write at execution time.
    from db import run_readonly_query
    return run_readonly_query(sql, timeout_ms=12000, max_rows=501)


def _get_active_ai_config() -> dict:
    """Load the active AI config from DB; fall back to ANTHROPIC_API_KEY env var."""
    try:
        from psycopg2.extras import RealDictCursor as _RDC
        conn = get_db_connection()
        try:
            with conn.cursor(cursor_factory=_RDC) as cur:
                cur.execute("""
                    SELECT * FROM ai_settings WHERE is_active=TRUE LIMIT 1
                """)
                row = cur.fetchone()
        finally:
            release_db_connection(conn)
        if row:
            return dict(row)
    except Exception:
        pass

    # Fallback: use ANTHROPIC_API_KEY from environment
    key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not key:
        raise RuntimeError(
            "No AI provider is configured. Go to Admin → AI Settings to set up an AI provider, "
            "or add ANTHROPIC_API_KEY to your .env file."
        )
    return {
        'provider': 'anthropic',
        'api_key': key,
        'model': 'claude-haiku-4-5-20251001',
        'api_base_url': '',
        'extra_config': {},
    }


def _call_llm(system: str, user_msg: str, max_tokens: int = 1024,
              feature: str = 'chat', user_id=None, username: str = '',
              cacheable: bool = False) -> str:
    from routes.ai_settings import call_provider
    config = _get_active_ai_config()
    return call_provider(config, system, user_msg, max_tokens,
                         feature=feature, user_id=user_id, username=username,
                         cacheable=cacheable)


def _strip_sql(raw: str) -> str:
    """Strip markdown, chain-of-thought tags, and prose preamble from LLM output."""
    sql = raw.strip()
    # Remove reasoning/thinking blocks (deepseek-r1, qwq, o1-style models)
    sql = re.sub(r'<think>.*?</think>', '', sql, flags=re.DOTALL | re.IGNORECASE).strip()
    # Remove markdown code fences
    sql = re.sub(r'^```(?:sql)?\s*\n?', '', sql, flags=re.IGNORECASE)
    sql = re.sub(r'\n?\s*```\s*$', '', sql).strip()
    # If SQL already starts with SELECT or WITH (CTE), return as-is — no stripping needed
    if re.match(r'\s*(SELECT|WITH)\b', sql, re.IGNORECASE):
        return sql
    # Otherwise the LLM added prose before the SQL; find first SQL keyword on its own line
    m = re.search(r'\n[ \t]*(SELECT|WITH)\b', sql, re.IGNORECASE)
    if m:
        return sql[m.start() + 1:].strip()   # skip the leading newline
    # Last resort: find SELECT anywhere inline (e.g. "Here: SELECT ...")
    m = re.search(r'\b(SELECT|WITH)\b', sql, re.IGNORECASE)
    if m:
        return sql[m.start():]
    return sql


def _nl_to_sql(question: str, history: list, user_id=None, username: str = '') -> str:
    history_str = ''
    if history:
        lines = []
        for h in history[-4:]:
            lines.append(f"User: {h.get('user','')}")
            if h.get('sql'):
                lines.append(f"SQL used: {h['sql']}")
        history_str = '\n'.join(lines) + '\n\n'
    base_prompt = f"{history_str}Question: {question}"

    for attempt in range(2):
        extra = ("\n\nOutput ONLY the raw SQL — no markdown, no explanation, no preamble. "
                 "Start your response with SELECT or WITH.") if attempt else ""
        # SYSTEM_SQL is a frozen constant (schema + few-shot examples) sent
        # unchanged on every call — mark it cacheable so Anthropic reuses the
        # cached prefix instead of reprocessing ~2-3K tokens each time.
        raw = _call_llm(SYSTEM_SQL, base_prompt + extra, max_tokens=1024,
                        feature='sql-query', user_id=user_id, username=username,
                        cacheable=True)
        logger.debug('SQL gen attempt %d raw: %.300s', attempt + 1, raw)
        sql = _strip_sql(raw)
        if re.search(r'\bSELECT\b', sql, re.IGNORECASE):
            return sql
        logger.warning('SQL gen attempt %d no SELECT found. Stripped: %.200s', attempt + 1, sql)

    raise ValueError(
        "The AI could not generate a valid SQL query for that question. "
        "Check your AI provider settings or try rephrasing the question."
    )


def _format_answer(question: str, sql: str, columns: list, rows: list,
                    user_id=None, username: str = '') -> str:
    sample = rows[:20]
    def _default(o):
        if isinstance(o, (date, datetime)):
            return o.isoformat()
        return str(o)
    data_json = json.dumps({'columns': columns, 'rows': sample}, default=_default)
    prompt = (
        f"Question: {question}\n"
        f"SQL: {sql}\n"
        f"Total rows returned: {len(rows)}\n"
        f"Sample data (up to 20 rows): {data_json}"
    )
    return _call_llm(SYSTEM_FORMAT, prompt, max_tokens=400,
                     feature='chat', user_id=user_id, username=username)


# ── POST /api/chat ─────────────────────────────────────────────────────────────
@chat_bp.route('/chat', methods=['POST'])
@permission_required('ai-chat')
def chat(current_user):
    body = request.get_json(force=True, silent=True) or {}
    question = (body.get('question') or '').strip()
    history  = body.get('history') or []

    if not question:
        return jsonify({'error': 'Question is required.'}), 400
    if len(question) > 1000:
        return jsonify({'error': 'Question too long (max 1000 characters).'}), 400

    uid  = current_user.get('user_id')
    uname = current_user.get('username', '')

    try:
        sql = _nl_to_sql(question, history, user_id=uid, username=uname)
    except RuntimeError as e:
        logger.warning('Chat AI config error: %s', e)
        return jsonify({'error': str(e)}), 503
    except ValueError as e:
        logger.warning('Chat SQL generation failed: %s', e)
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logger.exception('Chat SQL generation error: %s', e)
        return jsonify({'error': 'Failed to generate SQL. Please try again.'}), 500

    try:
        _validate_sql(sql)
    except ValueError as e:
        logger.warning('Chat SQL blocked: %s | sql=%s', e, sql[:200])
        return jsonify({'error': str(e)}), 400

    try:
        columns, rows = _run_sql(sql)
        total_rows   = len(rows)
        has_more     = total_rows > 500
        display_rows = rows[:500]
        answer       = _format_answer(question, sql, columns, display_rows,
                                      user_id=uid, username=uname)

        def _serial(o):
            if isinstance(o, (date, datetime)):
                return o.isoformat()
            return str(o)

        return jsonify({
            'answer':     answer,
            'sql':        sql,
            'columns':    columns,
            'rows':       json.loads(json.dumps(display_rows, default=_serial)),
            'total_rows': total_rows,
            'has_more':   has_more,
            'has_data':   bool(columns and display_rows),
        })
    except RuntimeError as e:
        logger.warning('Chat AI format error: %s', e)
        return jsonify({'error': str(e)}), 503
    except Exception as e:
        logger.exception('Chat query error: %s | sql=%s', e, sql[:200])
        return jsonify({'error': 'Query failed: ' + str(e)}), 500


# ── POST /api/chat/export ──────────────────────────────────────────────────────
@chat_bp.route('/chat/export', methods=['POST'])
@permission_required('ai-chat')
def chat_export(current_user):
    body = request.get_json(force=True, silent=True) or {}
    columns = body.get('columns') or []
    rows    = body.get('rows') or []
    question = body.get('question') or 'AI Report'
    sql      = body.get('sql') or ''

    if not columns:
        return jsonify({'error': 'No data to export.'}), 400

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'Report'

        PURPLE = '6366F1'
        ALT    = 'F5F3FF'
        thin   = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB'),
        )

        # Header row
        for c_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=c_idx, value=col.replace('_', ' ').title())
            cell.fill = PatternFill('solid', fgColor=PURPLE)
            cell.font = Font(color='FFFFFF', bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin

        ws.row_dimensions[1].height = 20

        # Data rows
        for r_idx, row in enumerate(rows, 2):
            fill = PatternFill('solid', fgColor=ALT) if r_idx % 2 == 0 else None
            for c_idx, col in enumerate(columns, 1):
                val = row.get(col)
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if fill:
                    cell.fill = fill
                cell.border = thin
                cell.alignment = Alignment(vertical='center')

        # Auto-width
        for c_idx, col in enumerate(columns, 1):
            max_len = max(
                len(str(col)),
                *(len(str(row.get(col) or '')) for row in rows[:200])
            )
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 4, 45)

        # Sheet 2: query info
        ws2 = wb.create_sheet('Query Info')
        ws2['A1'] = 'Question'
        ws2['B1'] = question
        ws2['A2'] = 'SQL Query'
        ws2['B2'] = sql
        ws2['A3'] = 'Exported At'
        ws2['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws2['A4'] = 'Total Rows'
        ws2['B4'] = len(rows)
        for r in range(1, 5):
            ws2.cell(row=r, column=1).font = Font(bold=True)
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 60

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_name = re.sub(r'[^\w\s-]', '', question)[:40].strip().replace(' ', '_') or 'report'
        filename  = f"ai_report_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    except ImportError:
        return jsonify({'error': 'openpyxl is not installed. Run: pip install openpyxl'}), 500
    except Exception as e:
        logger.exception('Chat export error: %s', e)
        return jsonify({'error': 'Export failed: ' + str(e)}), 500
