"""
Data Export API — Export business data in Excel format for backup or import into other systems.
"""
import io
from flask import Blueprint, jsonify, request, Response
from db import get_db_connection, release_db_connection
from auth import token_required
from psycopg2.extras import RealDictCursor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

data_export_bp = Blueprint('data_export', __name__)


def _format_excel(ws, data, headers):
    """Format Excel worksheet with headers and data."""
    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add data rows
    for row_num, row_data in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            value = row_data.get(header.lower().replace(' ', '_'))

            if isinstance(value, float):
                cell.number_format = '₹#,##0.00'
            elif isinstance(value, datetime):
                cell.value = value.strftime('%Y-%m-%d %H:%M:%S')
            else:
                cell.value = value

    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        ws.column_dimensions[chr(64 + col_num)].width = min(max_length + 2, 50)


@data_export_bp.route('/data/export/products', methods=['GET'])
@token_required
def export_products(payload):
    """Export products master data."""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        cur.execute("""
            SELECT
                name, sku, pack_size, gst_rate as "gst_rate",
                purchase_rate as "purchase_rate", selling_rate as "selling_rate", status
            FROM products
            ORDER BY name
        """)
        products = cur.fetchall()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        headers = ['Name', 'SKU', 'Pack Size', 'GST Rate (%)', 'Purchase Rate', 'Selling Rate', 'Status']
        ws.append(headers)

        # Add data
        for product in products:
            ws.append([
                product.get('name'),
                product.get('sku'),
                product.get('pack_size'),
                float(product.get('gst_rate', 0)),
                float(product.get('purchase_rate', 0)),
                float(product.get('selling_rate', 0)),
                product.get('status', 'Active')
            ])

        # Format headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        # Auto-adjust columns
        for col_num, header in enumerate(headers, 1):
            ws.column_dimensions[chr(64 + col_num)].width = 18

        # Generate file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=products_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)


@data_export_bp.route('/data/export/customers', methods=['GET'])
@token_required
def export_customers(payload):
    """Export customers master data."""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        cur.execute("""
            SELECT
                name, email, mobile, address,
                customer_code, current_balance
            FROM credit_customers
            ORDER BY name
        """)
        customers = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"

        headers = ['Name', 'Email', 'Mobile', 'Address', 'Customer Code', 'Current Balance']
        ws.append(headers)

        for customer in customers:
            ws.append([
                customer.get('name'),
                customer.get('email'),
                customer.get('mobile'),
                customer.get('address'),
                customer.get('customer_code'),
                float(customer.get('current_balance', 0))
            ])

        # Format headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_num)].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=customers_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)


@data_export_bp.route('/data/export/sales', methods=['GET'])
@token_required
def export_sales(payload):
    """Export sales invoices with line items."""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        cur.execute("""
            SELECT
                si.invoice_number, si.invoice_date, si.customer_name, si.customer_contact,
                si.total_amount, si.total_gst, si.mode_of_payment, si.status
            FROM sales_invoices si
            WHERE si.status IN ('Completed', 'Done')
            ORDER BY si.invoice_date DESC
        """)
        sales = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"

        headers = ['Invoice Number', 'Date', 'Customer Name', 'Contact', 'Total Amount', 'Total GST', 'Payment Mode', 'Status']
        ws.append(headers)

        for row in sales:
            ws.append([
                row.get('invoice_number'),
                row.get('invoice_date'),
                row.get('customer_name'),
                row.get('customer_contact'),
                float(row.get('total_amount', 0)),
                float(row.get('total_gst', 0)),
                row.get('mode_of_payment', ''),
                row.get('status', '')
            ])

        # Format headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_num)].width = 16

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=sales_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)


@data_export_bp.route('/data/export/purchases', methods=['GET'])
@token_required
def export_purchases(payload):
    """Export purchase orders with line items."""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        cur.execute("""
            SELECT
                purchase_order_number, purchase_date, supplier_name, supplier_gst_number,
                total_amount, status
            FROM purchase_orders
            ORDER BY purchase_date DESC
        """)
        purchases = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Purchases"

        headers = ['PO Number', 'Date', 'Supplier', 'Supplier GST', 'Total Amount', 'Status']
        ws.append(headers)

        for row in purchases:
            ws.append([
                row.get('purchase_order_number'),
                row.get('purchase_date'),
                row.get('supplier_name'),
                row.get('supplier_gst_number'),
                float(row.get('total_amount', 0)),
                row.get('status', '')
            ])

        # Format headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_num)].width = 16

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=purchases_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)


@data_export_bp.route('/data/export/inventory', methods=['GET'])
@token_required
def export_inventory(payload):
    """Export current inventory levels."""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        cur.execute("""
            SELECT
                p.name as "product_name", p.sku, p.pack_size,
                i.stock_quantity as "stock_quantity",
                p.purchase_rate, p.selling_rate,
                (i.stock_quantity * p.purchase_rate) as "stock_value"
            FROM inventory i
            JOIN products p ON i.product_id = p.product_id
            ORDER BY p.name
        """)
        inventory = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"

        headers = ['Product Name', 'SKU', 'Pack Size', 'Stock Quantity', 'Purchase Rate', 'Selling Rate', 'Stock Value']
        ws.append(headers)

        for row in inventory:
            ws.append([
                row.get('name'),
                row.get('sku'),
                row.get('pack_size'),
                int(row.get('stock_quantity', 0)),
                float(row.get('purchase_rate', 0)),
                float(row.get('selling_rate', 0)),
                float(row.get('stock_value', 0))
            ])

        # Format headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_num)].width = 16

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=inventory_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)


@data_export_bp.route('/data/export/complete', methods=['GET'])
@token_required
def export_complete(payload):
    """Export all data in a single workbook."""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Products sheet
        cur.execute("SELECT name, sku, pack_size, gst_rate, purchase_rate, selling_rate, status FROM products ORDER BY name")
        products = cur.fetchall()
        ws_products = wb.create_sheet("Products")
        headers = ['Name', 'SKU', 'Pack Size', 'GST Rate (%)', 'Purchase Rate', 'Selling Rate', 'Status']
        ws_products.append(headers)
        for p in products:
            ws_products.append([p.get('name'), p.get('sku'), p.get('pack_size'), float(p.get('gst_rate', 0)), float(p.get('purchase_rate', 0)), float(p.get('selling_rate', 0)), p.get('status', 'Active')])

        # Customers sheet
        cur.execute("SELECT name, email, mobile, address, customer_code, current_balance FROM credit_customers ORDER BY name")
        customers = cur.fetchall()
        ws_customers = wb.create_sheet("Customers")
        headers = ['Name', 'Email', 'Mobile', 'Address', 'Customer Code', 'Current Balance']
        ws_customers.append(headers)
        for c in customers:
            ws_customers.append([c.get('name'), c.get('email'), c.get('mobile'), c.get('address'), c.get('customer_code'), float(c.get('current_balance', 0))])

        # Inventory sheet
        cur.execute("SELECT p.name as name, p.sku, i.stock_quantity, p.purchase_rate, p.selling_rate FROM inventory i JOIN products p ON i.product_id = p.product_id ORDER BY p.name")
        inventory = cur.fetchall()
        ws_inventory = wb.create_sheet("Inventory")
        headers = ['Product Name', 'SKU', 'Stock Quantity', 'Purchase Rate', 'Selling Rate']
        ws_inventory.append(headers)
        for inv in inventory:
            ws_inventory.append([inv.get('name'), inv.get('sku'), int(inv.get('stock_quantity', 0)), float(inv.get('purchase_rate', 0)), float(inv.get('selling_rate', 0))])

        # Sales sheet
        cur.execute("""SELECT invoice_number, invoice_date, customer_name, total_amount, mode_of_payment
                      FROM sales_invoices WHERE status IN ('Completed', 'Done') ORDER BY invoice_date DESC LIMIT 1000""")
        sales = cur.fetchall()
        ws_sales = wb.create_sheet("Sales")
        headers = ['Invoice Number', 'Date', 'Customer', 'Amount', 'Payment Mode']
        ws_sales.append(headers)
        for s in sales:
            ws_sales.append([s.get('invoice_number'), s.get('invoice_date'), s.get('customer_name'), float(s.get('total_amount', 0)), s.get('mode_of_payment', '')])

        # Format all sheets
        for ws in wb.sheetnames:
            sheet = wb[ws]
            for col_num in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                sheet.column_dimensions[chr(64 + col_num)].width = 16

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=trintzpos_complete_backup_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)
