import logging
from flask import Blueprint, request, jsonify
from db import get_db_connection, release_db_connection
from auth import token_required, admin_required
from psycopg2.extras import RealDictCursor

logger = logging.getLogger(__name__)
products_bp = Blueprint('products', __name__)


@products_bp.route('/products', methods=['GET'])
@token_required
def get_products(payload):
    query = request.args.get('q', '')
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        if query:
            search_query = f"%{query}%"
            cur.execute("""
                SELECT p.*, i.stock_quantity
                FROM products p
                LEFT JOIN inventory i ON p.product_id = i.product_id
                WHERE p.status = 'Active'
                  AND (p.name ILIKE %s OR p.sku ILIKE %s)
                ORDER BY p.name
            """, (search_query, search_query))
        else:
            cur.execute("""
                SELECT p.*, i.stock_quantity
                FROM products p
                LEFT JOIN inventory i ON p.product_id = i.product_id
                WHERE p.status = 'Active'
                ORDER BY p.name
            """)
        return jsonify(cur.fetchall()), 200
    except Exception as e:
        logger.exception("Failed to fetch products")
        return jsonify({'message': 'Failed to fetch products'}), 500
    finally:
        cur.close()
        release_db_connection(conn)


@products_bp.route('/products', methods=['POST'])
@admin_required   # BUG-008 fixed: replaced inline role check with decorator
def create_product(payload):
    data = request.get_json()
    if not data:
        return jsonify({'message': 'No data provided'}), 400

    for field in ['name', 'sku', 'gst_rate', 'selling_rate']:
        if field not in data or data[field] is None:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    try:
        name         = str(data['name'])
        sku          = str(data['sku'])
        pack_size    = str(data['pack_size']) if data.get('pack_size') is not None else None
        gst_rate     = float(data['gst_rate'])
        purchase_rate = float(data['purchase_rate']) if data.get('purchase_rate') is not None else None
        selling_rate = float(data['selling_rate'])
        initial_stock = int(data.get('initial_stock', 0))
        barcode      = str(data['barcode']).strip() if data.get('barcode') else None
    except (ValueError, TypeError) as e:
        return jsonify({'message': f'Invalid data type: {e}'}), 400

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            INSERT INTO products (name, pack_size, sku, gst_rate, purchase_rate, selling_rate, barcode)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING product_id
        """, (name, pack_size, sku, gst_rate, purchase_rate, selling_rate, barcode))

        result = cur.fetchone()
        if result is None:
            raise Exception("INSERT returned no product_id")
        product_id = result['product_id']

        cur.execute(
            "INSERT INTO inventory (product_id, stock_quantity) VALUES (%s, %s)",
            (product_id, initial_stock)
        )
        conn.commit()

        cur.execute("""
            SELECT p.*, i.stock_quantity
            FROM products p
            LEFT JOIN inventory i ON p.product_id = i.product_id
            WHERE p.product_id = %s
        """, (product_id,))
        return jsonify(cur.fetchone()), 201
    except Exception as e:
        conn.rollback()
        logger.exception("Failed to create product")
        # BUG-010 fixed: no traceback or internal details sent to client
        return jsonify({'message': 'Failed to create product'}), 500
    finally:
        cur.close()
        release_db_connection(conn)


@products_bp.route('/products/barcode/<string:code>', methods=['GET'])
@token_required
def get_product_by_barcode(payload, code):
    """Exact lookup by barcode or SKU — used by hardware/camera scanner."""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT p.*, i.stock_quantity
            FROM products p
            LEFT JOIN inventory i ON p.product_id = i.product_id
            WHERE p.status = 'Active'
              AND (p.barcode = %s OR p.sku = %s)
            LIMIT 1
        """, (code, code))
        product = cur.fetchone()
        if not product:
            return jsonify({'message': 'Product not found'}), 404
        return jsonify(product), 200
    except Exception as e:
        logger.exception("Barcode lookup failed for code %s", code)
        return jsonify({'message': 'Lookup failed'}), 500
    finally:
        cur.close()
        release_db_connection(conn)


@products_bp.route('/products/export', methods=['GET'])
@admin_required
def export_products(payload):
    """Export all products to Excel (.xlsx)."""
    try:
        import openpyxl
        from io import BytesIO
        from flask import Response as FlaskResponse
    except ImportError as e:
        return jsonify({'message': f'Missing library: {e}'}), 500

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT p.name, p.sku, p.pack_size, p.gst_rate, p.purchase_rate,
                   p.selling_rate, p.barcode, p.status, i.stock_quantity
            FROM products p
            LEFT JOIN inventory i ON p.product_id = i.product_id
            ORDER BY p.name
        """)
        rows = cur.fetchall()
    finally:
        cur.close()
        release_db_connection(conn)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'
    cols = ['name', 'sku', 'pack_size', 'gst_rate', 'purchase_rate',
            'selling_rate', 'barcode', 'stock_quantity', 'status']
    ws.append(cols)
    for row in rows:
        ws.append([row.get(c) for c in cols])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FlaskResponse(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=products_export.xlsx'}
    )


@products_bp.route('/products/import', methods=['POST'])
@admin_required
def import_products(payload):
    """Upsert products from uploaded .xlsx or .csv file."""
    if 'file' not in request.files:
        return jsonify({'message': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'message': 'Empty filename'}), 400

    fname = f.filename.lower()
    try:
        rows = []
        if fname.endswith('.xlsx') or fname.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1))
            headers = [str(c.value).strip().lower() if c.value is not None else '' for c in header_row]
            for excel_row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in excel_row):
                    continue
                rows.append(dict(zip(headers, excel_row)))
        elif fname.endswith('.csv'):
            import csv, io
            content = f.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            for r in reader:
                rows.append({k.strip().lower(): v for k, v in r.items()})
        else:
            return jsonify({'message': 'Unsupported file type. Use .xlsx or .csv'}), 400
    except Exception as e:
        logger.exception("Failed to parse import file")
        return jsonify({'message': f'Failed to parse file: {e}'}), 400

    imported = updated = 0
    errors = []

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        for i, row in enumerate(rows, start=2):
            cur.execute("SAVEPOINT sp_row")
            try:
                name = str(row.get('name') or '').strip()
                sku  = str(row.get('sku')  or '').strip()
                if not name or not sku:
                    errors.append({'row': i, 'message': 'name and sku are required'})
                    cur.execute("ROLLBACK TO SAVEPOINT sp_row")
                    continue

                try:
                    gst_rate     = float(row.get('gst_rate')     or 0)
                    selling_rate = float(row.get('selling_rate')  or 0)
                except (ValueError, TypeError):
                    errors.append({'row': i, 'message': 'gst_rate and selling_rate must be numbers'})
                    cur.execute("ROLLBACK TO SAVEPOINT sp_row")
                    continue

                if selling_rate <= 0:
                    errors.append({'row': i, 'message': 'selling_rate must be > 0'})
                    cur.execute("ROLLBACK TO SAVEPOINT sp_row")
                    continue

                pack_size     = str(row.get('pack_size')     or '').strip() or None
                pr_val        = row.get('purchase_rate')
                purchase_rate = float(pr_val) if pr_val not in (None, '', 'None') else None
                bc_val        = row.get('barcode')
                barcode       = str(bc_val).strip() or None if bc_val not in (None, '', 'None') else None

                cur.execute("SELECT product_id FROM products WHERE sku = %s", (sku,))
                existing = cur.fetchone()

                if existing:
                    cur.execute("""
                        UPDATE products
                        SET name=%s, pack_size=%s, gst_rate=%s, purchase_rate=%s,
                            selling_rate=%s, barcode=%s
                        WHERE sku=%s
                    """, (name, pack_size, gst_rate, purchase_rate, selling_rate, barcode, sku))
                    updated += 1
                else:
                    cur.execute("""
                        INSERT INTO products (name, sku, pack_size, gst_rate, purchase_rate, selling_rate, barcode)
                        VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING product_id
                    """, (name, sku, pack_size, gst_rate, purchase_rate, selling_rate, barcode))
                    product_id = cur.fetchone()['product_id']
                    sq_raw = row.get('stock_quantity') or row.get('initial_stock') or 0
                    try:
                        stock = int(float(sq_raw))
                    except (ValueError, TypeError):
                        stock = 0
                    cur.execute(
                        "INSERT INTO inventory (product_id, stock_quantity) VALUES (%s, %s)",
                        (product_id, stock)
                    )
                    imported += 1

                cur.execute("RELEASE SAVEPOINT sp_row")
            except Exception as e:
                cur.execute("ROLLBACK TO SAVEPOINT sp_row")
                logger.exception("Error on import row %d", i)
                errors.append({'row': i, 'message': str(e)})

        conn.commit()
    except Exception as e:
        conn.rollback()
        logger.exception("Import transaction failed")
        return jsonify({'message': f'Import failed: {e}'}), 500
    finally:
        cur.close()
        release_db_connection(conn)

    return jsonify({
        'imported': imported,
        'updated':  updated,
        'errors':   errors,
        'message':  f'Done: {imported} added, {updated} updated, {len(errors)} errors'
    }), 200


@products_bp.route('/products/<int:product_id>', methods=['GET'])
@token_required
def get_product_by_id(payload, product_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT p.*, i.stock_quantity
            FROM products p
            LEFT JOIN inventory i ON p.product_id = i.product_id
            WHERE p.product_id = %s
        """, (product_id,))
        product = cur.fetchone()
        if not product:
            return jsonify({'message': 'Product not found'}), 404
        return jsonify(product), 200
    except Exception as e:
        logger.exception("Failed to fetch product %s", product_id)
        return jsonify({'message': 'Failed to fetch product'}), 500
    finally:
        cur.close()
        release_db_connection(conn)


@products_bp.route('/products/<int:product_id>', methods=['PUT'])
@admin_required   # BUG-008 fixed
def update_product(payload, product_id):
    data = request.get_json()
    if not data:
        return jsonify({'message': 'No data provided'}), 400

    for field in ['name', 'sku', 'gst_rate', 'selling_rate']:
        if field not in data or data[field] is None:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    try:
        name         = str(data['name'])
        sku          = str(data['sku'])
        pack_size    = str(data['pack_size']) if data.get('pack_size') is not None else None
        gst_rate     = float(data['gst_rate'])
        purchase_rate = float(data['purchase_rate']) if data.get('purchase_rate') is not None else None
        selling_rate = float(data['selling_rate'])
        barcode      = str(data['barcode']).strip() if data.get('barcode') else None
    except (ValueError, TypeError) as e:
        return jsonify({'message': f'Invalid data type: {e}'}), 400

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            UPDATE products
            SET name=%s, pack_size=%s, sku=%s, gst_rate=%s, purchase_rate=%s, selling_rate=%s, barcode=%s
            WHERE product_id=%s
        """, (name, pack_size, sku, gst_rate, purchase_rate, selling_rate, barcode, product_id))

        if cur.rowcount == 0:
            return jsonify({'message': 'Product not found'}), 404
        conn.commit()

        cur.execute("""
            SELECT p.*, i.stock_quantity
            FROM products p
            LEFT JOIN inventory i ON p.product_id = i.product_id
            WHERE p.product_id = %s
        """, (product_id,))
        return jsonify(cur.fetchone()), 200
    except Exception as e:
        conn.rollback()
        logger.exception("Failed to update product %s", product_id)
        return jsonify({'message': 'Failed to update product'}), 500  # BUG-010 fixed
    finally:
        cur.close()
        release_db_connection(conn)


@products_bp.route('/products/<int:product_id>', methods=['DELETE'])
@admin_required   # BUG-008 fixed
def delete_product(payload, product_id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT COUNT(*) FROM sales_invoice_items WHERE product_id=%s", (product_id,))
        sales_count = cur.fetchone()[0]
        if sales_count > 0:
            return jsonify({'message': f'Cannot delete: referenced in {sales_count} sales invoice(s)'}), 400

        cur.execute("SELECT COUNT(*) FROM purchase_order_items WHERE product_id=%s", (product_id,))
        if cur.fetchone()[0] > 0:
            return jsonify({'message': 'Cannot delete: referenced in purchase orders'}), 400

        cur.execute("DELETE FROM inventory WHERE product_id=%s", (product_id,))
        cur.execute("DELETE FROM products WHERE product_id=%s", (product_id,))
        if cur.rowcount == 0:
            return jsonify({'message': 'Product not found'}), 404
        conn.commit()
        return jsonify({'message': 'Product deleted successfully'}), 200
    except Exception as e:
        conn.rollback()
        logger.exception("Failed to delete product %s", product_id)
        return jsonify({'message': 'Failed to delete product'}), 500
    finally:
        cur.close()
        release_db_connection(conn)


@products_bp.route('/products/<int:product_id>/deactivate', methods=['PUT'])
@admin_required   # BUG-008 fixed
def deactivate_product(payload, product_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("UPDATE products SET status='Inactive' WHERE product_id=%s", (product_id,))
        if cur.rowcount == 0:
            return jsonify({'message': 'Product not found'}), 404
        conn.commit()

        cur.execute("""
            SELECT p.*, i.stock_quantity
            FROM products p
            LEFT JOIN inventory i ON p.product_id = i.product_id
            WHERE p.product_id = %s
        """, (product_id,))
        return jsonify(cur.fetchone()), 200
    except Exception as e:
        conn.rollback()
        logger.exception("Failed to deactivate product %s", product_id)
        return jsonify({'message': 'Failed to deactivate product'}), 500
    finally:
        cur.close()
        release_db_connection(conn)
