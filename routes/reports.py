from flask import Blueprint, request, jsonify, Response
from db import get_db_connection, release_db_connection
from auth import token_required
from psycopg2.extras import RealDictCursor
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/reports/sales', methods=['GET'])
@token_required
def get_sales_report(payload):
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Build query based on date filters with proper date handling
        base_query = """
            SELECT 
                DATE(si.invoice_date) as sale_date,
                COUNT(si.invoice_id) as total_invoices,
                SUM(si.total_amount) as total_sales,
                SUM(si.total_gst) as total_gst_collected
            FROM sales_invoices si
            WHERE si.status = 'Completed'
        """
        
        params = []
        if start_date:
            base_query += " AND DATE(si.invoice_date) >= %s"
            params.append(start_date)
            
        if end_date:
            base_query += " AND DATE(si.invoice_date) <= %s"
            params.append(end_date)
            
        base_query += " GROUP BY DATE(si.invoice_date) ORDER BY sale_date"
        
        cur.execute(base_query, params)
        daily_sales = cur.fetchall()
        
        # Calculate overall totals
        total_sales = sum(row['total_sales'] if row['total_sales'] else 0 for row in daily_sales)
        total_gst = sum(row['total_gst_collected'] if row['total_gst_collected'] else 0 for row in daily_sales)
        total_invoices = sum(row['total_invoices'] for row in daily_sales)
        
        # Get top selling products with proper date filtering
        product_query = """
            SELECT 
                p.name,
                SUM(sii.quantity) as total_quantity_sold,
                SUM(sii.total_line_amount) as total_value_sold
            FROM sales_invoice_items sii
            JOIN products p ON sii.product_id = p.product_id
            JOIN sales_invoices si ON sii.invoice_id = si.invoice_id
            WHERE si.status = 'Completed'
        """
        
        product_params = []
        if start_date:
            product_query += " AND DATE(si.invoice_date) >= %s"
            product_params.append(start_date)
            
        if end_date:
            product_query += " AND DATE(si.invoice_date) <= %s"
            product_params.append(end_date)
            
        product_query += " GROUP BY p.product_id, p.name ORDER BY total_quantity_sold DESC LIMIT 10"
        
        cur.execute(product_query, product_params)
        top_products = cur.fetchall()
        
        # Get detailed invoice data
        include_cancelled = request.args.get('include_cancelled', 'false').lower() == 'true'
        
        status_filter = '' if include_cancelled else "WHERE si.status = 'Completed'"
        status_column = ', si.status' if include_cancelled else ''
        status_group = ', si.status' if include_cancelled else ''
        
        detailed_query = f"""
            SELECT 
                si.invoice_id,
                si.invoice_number,
                si.invoice_date,
                si.customer_name,
                si.customer_contact,
                si.total_amount,
                si.total_gst,
                si.mode_of_payment,
                COUNT(sii.item_id) as item_count{status_column}
            FROM sales_invoices si
            LEFT JOIN sales_invoice_items sii ON si.invoice_id = sii.invoice_id
            {status_filter}
        """
        
        detailed_params = params.copy()  # Use the same date filters
        if start_date:
            if include_cancelled:
                detailed_query += " WHERE DATE(si.invoice_date) >= %s"
            else:
                detailed_query += " AND DATE(si.invoice_date) >= %s"
            
        if end_date:
            detailed_query += " AND DATE(si.invoice_date) <= %s"
            
        detailed_query += f" GROUP BY si.invoice_id, si.invoice_number, si.invoice_date, si.customer_name, si.customer_contact, si.total_amount, si.total_gst, si.mode_of_payment{status_group} ORDER BY si.invoice_date DESC"
        
        cur.execute(detailed_query, detailed_params)
        detailed_invoices = cur.fetchall()
        
        # Format the response to match what the frontend expects
        return jsonify({
            'total_sales': float(total_sales) if total_sales else 0.0,
            'total_gst': float(total_gst) if total_gst else 0.0,
            'total_invoices': int(total_invoices) if total_invoices else 0,
            'sales_trend': [
                {
                    'date': str(row['sale_date']),
                    'amount': float(row['total_sales'] if row['total_sales'] else 0),
                    'invoices': int(row['total_invoices'])
                } for row in daily_sales
            ],
            'top_products': [
                {
                    'name': row['name'] or 'Unknown Product',
                    'quantity_sold': int(row['total_quantity_sold'] if row['total_quantity_sold'] else 0),
                    'total_value': float(row['total_value_sold'] if row['total_value_sold'] else 0)
                } for row in top_products
            ],
            'detailed_invoices': detailed_invoices
        }), 200
    except Exception as e:
        print(f"Error generating sales report: {str(e)}")  # Log the error for debugging
        return jsonify({'message': 'Failed to generate sales report', 'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)

@reports_bp.route('/reports/sales/export', methods=['GET'])
@token_required
def export_sales_report_excel(payload):
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Sales Report"
        
        # Add header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # Report title
        if ws is not None:
            ws['A1'] = 'SALES REPORT'
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:H1')
            ws['A1'].alignment = center_alignment
        
        # Date range
        date_range = f"From {start_date or 'Beginning'} to {end_date or 'Today'}"
        if ws is not None:
            ws['A2'] = date_range
            ws['A2'].font = Font(bold=True)
            ws.merge_cells('A2:H2')
            ws['A2'].alignment = center_alignment
        
        # Add some spacing
        if ws is not None:
            ws.append([''])
        
        # Get detailed sales data
        detailed_query = """
            SELECT 
                si.invoice_number,
                si.invoice_date,
                si.customer_name,
                si.customer_contact,
                si.total_amount,
                si.total_gst,
                si.discount_amount,
                si.discount_percentage,
                u.username as cashier_name
            FROM sales_invoices si
            JOIN users u ON si.user_id = u.user_id
            WHERE si.status = 'Completed'
        """
        
        detailed_params = []
        if start_date:
            detailed_query += " AND DATE(si.invoice_date) >= %s"
            detailed_params.append(start_date)
            
        if end_date:
            detailed_query += " AND DATE(si.invoice_date) <= %s"
            detailed_params.append(end_date)
            
        detailed_query += " ORDER BY si.invoice_date DESC"
        
        cur.execute(detailed_query, detailed_params)
        detailed_sales = cur.fetchall()
        
        # Detailed sales section header
        if ws is not None:
            ws.append([''])
            ws.append(['INVOICE DETAILS'])
            ws['A4'].font = Font(bold=True, size=14)
            ws.merge_cells('A4:I4')
        
        # Add column headers for detailed sales
        headers = ['Invoice Number', 'Date', 'Customer Name', 'Customer Contact', 
                  'Total Amount (₹)', 'Total GST (₹)', 'Discount (₹)', 'Discount (%)', 'Cashier']
        if ws is not None:
            ws.append(headers)
        
        # Style the headers
        if ws is not None:
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
        
        # Add detailed sales data
        row_num = 6
        for row_data in detailed_sales:
            if ws is not None:
                ws.append([
                    row_data['invoice_number'],
                    row_data['invoice_date'].strftime('%Y-%m-%d %H:%M:%S') if row_data['invoice_date'] else '',
                    row_data['customer_name'] or '',
                    row_data['customer_contact'] or '',
                    float(row_data['total_amount']) if row_data['total_amount'] else 0.0,
                    float(row_data['total_gst']) if row_data['total_gst'] else 0.0,
                    float(row_data['discount_amount']) if row_data['discount_amount'] else 0.0,
                    float(row_data['discount_percentage']) if row_data['discount_percentage'] else 0.0,
                    row_data['cashier_name'] or ''
                ])
                row_num += 1
        
        # Add some spacing
        if ws is not None:
            ws.append([''])
            row_num += 1
        
        # Get itemized sales data
        itemized_query = """
            SELECT 
                si.invoice_number,
                si.invoice_date,
                p.name as product_name,
                p.pack_size,
                sii.quantity,
                sii.rate_at_sale,
                sii.gst_rate_at_sale,
                sii.exclusive_gst_amount,
                sii.sgst,
                sii.cgst,
                sii.total_line_amount,
                sii.discount_percentage as item_discount_percentage
            FROM sales_invoice_items sii
            JOIN sales_invoices si ON sii.invoice_id = si.invoice_id
            JOIN products p ON sii.product_id = p.product_id
            WHERE si.status = 'Completed'
        """
        
        itemized_params = []
        if start_date:
            itemized_query += " AND DATE(si.invoice_date) >= %s"
            itemized_params.append(start_date)
            
        if end_date:
            itemized_query += " AND DATE(si.invoice_date) <= %s"
            itemized_params.append(end_date)
            
        itemized_query += " ORDER BY si.invoice_date DESC, sii.item_id"
        
        cur.execute(itemized_query, itemized_params)
        itemized_sales = cur.fetchall()
        
        # Itemized sales section header
        if ws is not None:
            ws.append([''])
            ws.append(['ITEMIZED SALES DETAILS'])
            ws.cell(row=row_num+2, column=1).font = Font(bold=True, size=14)
            ws.merge_cells(start_row=row_num+2, start_column=1, end_row=row_num+2, end_column=12)
            row_num += 2
        
        # Add column headers for itemized sales
        item_headers = ['Invoice Number', 'Date', 'Product Name', 'Pack Size', 'Quantity', 
                       'Rate (₹)', 'GST Rate (%)', 'Taxable Value (₹)', 'SGST (₹)', 'CGST (₹)', 'Total Amount (₹)', 'Discount (%)']
        if ws is not None:
            ws.append(item_headers)
            row_num += 1
        
        # Style the item headers
        if ws is not None:
            for col_num, header in enumerate(item_headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
        
        # Add itemized sales data
        for row_data in itemized_sales:
            if ws is not None:
                ws.append([
                    row_data['invoice_number'],
                    row_data['invoice_date'].strftime('%Y-%m-%d') if row_data['invoice_date'] else '',
                    row_data['product_name'] or '',
                    row_data['pack_size'] or '',
                    int(row_data['quantity']) if row_data['quantity'] else 0,
                    float(row_data['rate_at_sale']) if row_data['rate_at_sale'] else 0.0,
                    float(row_data['gst_rate_at_sale']) if row_data['gst_rate_at_sale'] else 0.0,
                    float(row_data['exclusive_gst_amount']) if row_data['exclusive_gst_amount'] else 0.0,
                    float(row_data['sgst']) if row_data['sgst'] else 0.0,
                    float(row_data['cgst']) if row_data['cgst'] else 0.0,
                    float(row_data['total_line_amount']) if row_data['total_line_amount'] else 0.0,
                    float(row_data['item_discount_percentage']) if row_data['item_discount_percentage'] else 0.0
                ])
                row_num += 1
        
        # Auto-adjust column widths
        if ws is not None:
            for col_num in range(1, len(item_headers) + 1):
                max_length = 0
                column_letter = chr(64 + col_num)  # Convert 1-26 to A-Z
                for row_num in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook to a bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create response
        filename = f"sales_report_{start_date or 'beginning'}_to_{end_date or 'today'}.xlsx"
        response = Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
        return response
        
    except Exception as e:
        print(f"Error exporting sales report to Excel: {str(e)}")
        return jsonify({'message': 'Failed to export sales report', 'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)

@reports_bp.route('/reports/purchase', methods=['GET'])
@token_required
def get_purchase_report(payload):
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Get detailed purchase data
        detailed_query = """
            SELECT 
                po.purchase_order_number,
                po.purchase_date,
                s.supplier_name,
                s.supplier_gst_number,
                po.total_amount,
                u.username as created_by
            FROM purchase_orders po
            JOIN suppliers s ON po.supplier_id = s.supplier_id
            JOIN users u ON po.user_id = u.user_id
            WHERE po.status = 'Completed'
        """
        
        detailed_params = []
        if start_date:
            detailed_query += " AND DATE(po.purchase_date) >= %s"
            detailed_params.append(start_date)
            
        if end_date:
            detailed_query += " AND DATE(po.purchase_date) <= %s"
            detailed_params.append(end_date)
            
        detailed_query += " ORDER BY po.purchase_date DESC"
        
        cur.execute(detailed_query, detailed_params)
        detailed_purchases = cur.fetchall()
        
        # Get itemized purchase data
        itemized_query = """
            SELECT 
                po.purchase_order_number,
                po.purchase_date,
                s.supplier_name,
                p.name as product_name,
                p.pack_size,
                poi.quantity,
                poi.purchase_rate,
                poi.gst_rate,
                poi.taxable_value,
                poi.sgst,
                poi.cgst,
                poi.total_amount as item_total_amount
            FROM purchase_order_items poi
            JOIN purchase_orders po ON poi.purchase_order_id = po.purchase_order_id
            JOIN suppliers s ON po.supplier_id = s.supplier_id
            JOIN products p ON poi.product_id = p.product_id
            WHERE po.status = 'Completed'
        """
        
        itemized_params = []
        if start_date:
            itemized_query += " AND DATE(po.purchase_date) >= %s"
            itemized_params.append(start_date)
            
        if end_date:
            itemized_query += " AND DATE(po.purchase_date) <= %s"
            itemized_params.append(end_date)
            
        itemized_query += " ORDER BY po.purchase_date DESC, poi.item_id"
        
        cur.execute(itemized_query, itemized_params)
        itemized_purchases = cur.fetchall()
        
        # Calculate overall totals
        total_purchase_amount = sum(row['total_amount'] if row['total_amount'] else 0 for row in detailed_purchases)
        total_purchase_orders = len(detailed_purchases)
        
        # Format the response
        return jsonify({
            'total_purchase_amount': float(total_purchase_amount) if total_purchase_amount else 0.0,
            'total_purchase_orders': int(total_purchase_orders) if total_purchase_orders else 0,
            'purchase_orders': detailed_purchases,
            'purchase_items': itemized_purchases
        }), 200
    except Exception as e:
        print(f"Error generating purchase report: {str(e)}")
        return jsonify({'message': 'Failed to generate purchase report', 'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)

@reports_bp.route('/reports/purchase/export', methods=['GET'])
@token_required
def export_purchase_report_excel(payload):
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Purchase Report"
        
        # Add header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # Report title
        if ws is not None:
            ws['A1'] = 'PURCHASE REPORT'
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:H1')
            ws['A1'].alignment = center_alignment
        
        # Date range
        date_range = f"From {start_date or 'Beginning'} to {end_date or 'Today'}"
        if ws is not None:
            ws['A2'] = date_range
            ws['A2'].font = Font(bold=True)
            ws.merge_cells('A2:H2')
            ws['A2'].alignment = center_alignment
        
        # Add some spacing
        if ws is not None:
            ws.append([''])
        
        # Get detailed purchase data
        detailed_query = """
            SELECT 
                po.purchase_order_number,
                po.purchase_date,
                s.supplier_name,
                s.supplier_gst_number,
                po.total_amount,
                u.username as created_by
            FROM purchase_orders po
            JOIN suppliers s ON po.supplier_id = s.supplier_id
            JOIN users u ON po.user_id = u.user_id
            WHERE po.status = 'Completed'
        """
        
        detailed_params = []
        if start_date:
            detailed_query += " AND DATE(po.purchase_date) >= %s"
            detailed_params.append(start_date)
            
        if end_date:
            detailed_query += " AND DATE(po.purchase_date) <= %s"
            detailed_params.append(end_date)
            
        detailed_query += " ORDER BY po.purchase_date DESC"
        
        cur.execute(detailed_query, detailed_params)
        detailed_purchases = cur.fetchall()
        
        # Detailed purchase section header
        if ws is not None:
            ws.append([''])
            ws.append(['PURCHASE ORDER DETAILS'])
            ws['A4'].font = Font(bold=True, size=14)
            ws.merge_cells('A4:G4')
        
        # Add column headers for detailed purchases
        headers = ['Purchase Order Number', 'Date', 'Supplier Name', 'Supplier GST Number', 
                  'Total Amount (₹)', 'Created By']
        if ws is not None:
            ws.append(headers)
        
        # Style the headers
        if ws is not None:
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
        
        # Add detailed purchase data
        row_num = 6
        for row_data in detailed_purchases:
            if ws is not None:
                ws.append([
                    row_data['purchase_order_number'],
                    row_data['purchase_date'].strftime('%Y-%m-%d %H:%M:%S') if row_data['purchase_date'] else '',
                    row_data['supplier_name'] or '',
                    row_data['supplier_gst_number'] or '',
                    float(row_data['total_amount']) if row_data['total_amount'] else 0.0,
                    row_data['created_by'] or ''
                ])
                row_num += 1
        
        # Add some spacing
        if ws is not None:
            ws.append([''])
            row_num += 1
        
        # Get itemized purchase data
        itemized_query = """
            SELECT 
                po.purchase_order_number,
                po.purchase_date,
                s.supplier_name,
                p.name as product_name,
                p.pack_size,
                poi.quantity,
                poi.purchase_rate,
                poi.gst_rate,
                poi.taxable_value,
                poi.sgst,
                poi.cgst,
                poi.total_amount as item_total_amount
            FROM purchase_order_items poi
            JOIN purchase_orders po ON poi.purchase_order_id = po.purchase_order_id
            JOIN suppliers s ON po.supplier_id = s.supplier_id
            JOIN products p ON poi.product_id = p.product_id
            WHERE po.status = 'Completed'
        """
        
        itemized_params = []
        if start_date:
            itemized_query += " AND DATE(po.purchase_date) >= %s"
            itemized_params.append(start_date)
            
        if end_date:
            itemized_query += " AND DATE(po.purchase_date) <= %s"
            itemized_params.append(end_date)
            
        itemized_query += " ORDER BY po.purchase_date DESC, poi.item_id"
        
        cur.execute(itemized_query, itemized_params)
        itemized_purchases = cur.fetchall()
        
        # Itemized purchase section header
        if ws is not None:
            ws.append([''])
            ws.append(['ITEMIZED PURCHASE DETAILS'])
            ws.cell(row=row_num+2, column=1).font = Font(bold=True, size=14)
            ws.merge_cells(start_row=row_num+2, start_column=1, end_row=row_num+2, end_column=12)
            row_num += 2
        
        # Add column headers for itemized purchases
        item_headers = ['Purchase Order Number', 'Date', 'Supplier Name', 'Product Name', 'Pack Size', 'Quantity', 
                       'Purchase Rate (₹)', 'GST Rate (%)', 'Taxable Value (₹)', 'SGST (₹)', 'CGST (₹)', 'Total Amount (₹)']
        if ws is not None:
            ws.append(item_headers)
            row_num += 1
        
        # Style the item headers
        if ws is not None:
            for col_num, header in enumerate(item_headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
        
        # Add itemized purchase data
        for row_data in itemized_purchases:
            if ws is not None:
                ws.append([
                    row_data['purchase_order_number'],
                    row_data['purchase_date'].strftime('%Y-%m-%d') if row_data['purchase_date'] else '',
                    row_data['supplier_name'] or '',
                    row_data['product_name'] or '',
                    row_data['pack_size'] or '',
                    int(row_data['quantity']) if row_data['quantity'] else 0,
                    float(row_data['purchase_rate']) if row_data['purchase_rate'] else 0.0,
                    float(row_data['gst_rate']) if row_data['gst_rate'] else 0.0,
                    float(row_data['taxable_value']) if row_data['taxable_value'] else 0.0,
                    float(row_data['sgst']) if row_data['sgst'] else 0.0,
                    float(row_data['cgst']) if row_data['cgst'] else 0.0,
                    float(row_data['item_total_amount']) if row_data['item_total_amount'] else 0.0
                ])
                row_num += 1
        
        # Auto-adjust column widths
        if ws is not None:
            for col_num in range(1, len(item_headers) + 1):
                max_length = 0
                column_letter = chr(64 + col_num)  # Convert 1-26 to A-Z
                for row_num in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook to a bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create response
        filename = f"purchase_report_{start_date or 'beginning'}_to_{end_date or 'today'}.xlsx"
        response = Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
        return response
        
    except Exception as e:
        print(f"Error exporting purchase report to Excel: {str(e)}")
        return jsonify({'message': 'Failed to export purchase report', 'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)

@reports_bp.route('/reports/inventory', methods=['GET'])
@token_required
def get_inventory_report(payload):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Get detailed inventory data
        cur.execute("""
            SELECT 
                p.product_id,
                p.name,
                p.sku,
                p.pack_size,
                p.selling_rate,
                p.gst_rate,
                i.stock_quantity,
                (p.selling_rate * i.stock_quantity) as subtotal
            FROM products p
            JOIN inventory i ON p.product_id = i.product_id
            ORDER BY p.name
        """)
        
        inventory_data = cur.fetchall()
        
        # Calculate overall totals
        total_products = len(inventory_data)
        total_stock_value = sum(row['subtotal'] if row['subtotal'] else 0 for row in inventory_data)
        total_stock_quantity = sum(row['stock_quantity'] if row['stock_quantity'] else 0 for row in inventory_data)
        
        # Format the response
        return jsonify({
            'total_products': int(total_products) if total_products else 0,
            'total_stock_value': float(total_stock_value) if total_stock_value else 0.0,
            'total_stock_quantity': int(total_stock_quantity) if total_stock_quantity else 0,
            'inventory_items': inventory_data
        }), 200
    except Exception as e:
        print(f"Error generating inventory report: {str(e)}")
        return jsonify({'message': 'Failed to generate inventory report', 'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)

@reports_bp.route('/reports/inventory/export', methods=['GET'])
@token_required
def export_inventory_report_excel(payload):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Inventory Report"
        
        # Add header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # Report title
        if ws is not None:
            ws['A1'] = 'INVENTORY REPORT'
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:H1')
            ws['A1'].alignment = center_alignment
        
        # Add some spacing
        if ws is not None:
            ws.append([''])
        
        # Get detailed inventory data
        cur.execute("""
            SELECT 
                p.product_id,
                p.name,
                p.sku,
                p.pack_size,
                p.selling_rate,
                p.gst_rate,
                i.stock_quantity,
                (p.selling_rate * i.stock_quantity) as subtotal
            FROM products p
            JOIN inventory i ON p.product_id = i.product_id
            ORDER BY p.name
        """)
        
        inventory_data = cur.fetchall()
        
        # Inventory section header
        if ws is not None:
            ws.append([''])
            ws.append(['INVENTORY DETAILS'])
            ws['A4'].font = Font(bold=True, size=14)
            ws.merge_cells('A4:H4')
        
        # Add column headers for inventory
        headers = ['Product ID', 'Product Name', 'SKU', 'Pack Size', 'Selling Rate (₹)', 
                  'GST Rate (%)', 'Stock Quantity', 'Subtotal (₹)']
        if ws is not None:
            ws.append(headers)
        
        # Style the headers
        if ws is not None:
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
        
        # Add inventory data
        row_num = 6
        for row_data in inventory_data:
            if ws is not None:
                ws.append([
                    row_data['product_id'],
                    row_data['name'] or '',
                    row_data['sku'] or '',
                    row_data['pack_size'] or '',
                    float(row_data['selling_rate']) if row_data['selling_rate'] else 0.0,
                    float(row_data['gst_rate']) if row_data['gst_rate'] else 0.0,
                    int(row_data['stock_quantity']) if row_data['stock_quantity'] else 0,
                    float(row_data['subtotal']) if row_data['subtotal'] else 0.0
                ])
                row_num += 1
        
        # Add summary row
        total_products = len(inventory_data)
        total_stock_value = sum(row['subtotal'] if row['subtotal'] else 0 for row in inventory_data)
        total_stock_quantity = sum(row['stock_quantity'] if row['stock_quantity'] else 0 for row in inventory_data)
        
        if ws is not None:
            ws.append([''])
            ws.append(['SUMMARY'])
            ws.cell(row=row_num+2, column=1).font = Font(bold=True, size=14)
            ws.merge_cells(start_row=row_num+2, start_column=1, end_row=row_num+2, end_column=8)
            
            ws.append(['Total Products', total_products, '', '', '', '', 'Total Stock Quantity', total_stock_quantity])
            ws.append(['', '', '', '', '', '', 'Total Stock Value (₹)', total_stock_value])
            
            # Style summary
            summary_row = row_num + 4
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
            ws.cell(row=summary_row, column=2).font = Font(bold=True)
            ws.cell(row=summary_row, column=7).font = Font(bold=True)
            ws.cell(row=summary_row, column=8).font = Font(bold=True)
            ws.cell(row=summary_row+1, column=7).font = Font(bold=True)
            ws.cell(row=summary_row+1, column=8).font = Font(bold=True)
        
        # Auto-adjust column widths
        if ws is not None:
            for col_num in range(1, len(headers) + 1):
                max_length = 0
                column_letter = chr(64 + col_num)  # Convert 1-26 to A-Z
                for row_num in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook to a bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create response
        filename = f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
        return response

    except Exception as e:
        print(f"Error exporting inventory report to Excel: {str(e)}")
        return jsonify({'message': 'Failed to export inventory report', 'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)


# ── Profit & Loss Report ──────────────────────────────────────────────────────

@reports_bp.route('/reports/profit-loss', methods=['GET'])
@token_required
def get_profit_loss_report(payload):
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        params = []
        date_filter = ""
        if start_date:
            date_filter += " AND DATE(si.invoice_date) >= %s"
            params.append(start_date)
        if end_date:
            date_filter += " AND DATE(si.invoice_date) <= %s"
            params.append(end_date)

        # Revenue: Total sales
        revenue_query = f"""
            SELECT
                COALESCE(SUM(si.total_amount), 0) as revenue,
                COALESCE(SUM(si.total_gst), 0) as tax_collected
            FROM sales_invoices si
            WHERE si.status = 'Completed' {date_filter}
        """
        cur.execute(revenue_query, params)
        revenue_data = cur.fetchone()
        total_revenue = float(revenue_data['revenue']) if revenue_data['revenue'] else 0.0
        tax_collected = float(revenue_data['tax_collected']) if revenue_data['tax_collected'] else 0.0

        # Cost of Goods Sold: Sum of purchase costs for sold items
        cogs_query = f"""
            SELECT
                COALESCE(SUM(sii.quantity * p.purchase_rate), 0) as cogs
            FROM sales_invoice_items sii
            JOIN products p ON sii.product_id = p.product_id
            JOIN sales_invoices si ON sii.invoice_id = si.invoice_id
            WHERE si.status = 'Completed' {date_filter}
        """
        cur.execute(cogs_query, params)
        cogs_data = cur.fetchone()
        cost_of_goods_sold = float(cogs_data['cogs']) if cogs_data['cogs'] else 0.0

        # Gross Profit
        gross_profit = total_revenue - cost_of_goods_sold
        gross_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0

        # Operating Expenses (approximated from returns/discounts)
        expenses_query = f"""
            SELECT
                COALESCE(SUM(si.discount_amount), 0) as discounts,
                COALESCE(SUM(sr.total_amount), 0) as returns
            FROM sales_invoices si
            LEFT JOIN sales_returns sr ON si.invoice_id = sr.original_invoice_id
            WHERE si.status = 'Completed' {date_filter}
        """
        cur.execute(expenses_query, params)
        expenses_data = cur.fetchone()
        discounts = float(expenses_data['discounts']) if expenses_data['discounts'] else 0.0
        returns = float(expenses_data['returns']) if expenses_data['returns'] else 0.0
        total_expenses = discounts + returns

        # Net Profit
        net_profit = gross_profit - total_expenses
        net_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0

        # Daily breakdown
        daily_query = f"""
            SELECT
                DATE(si.invoice_date) as report_date,
                COALESCE(SUM(si.total_amount), 0) as daily_revenue,
                COALESCE(SUM(sii.quantity * p.purchase_rate), 0) as daily_cogs,
                COALESCE(SUM(si.discount_amount), 0) as daily_discounts
            FROM sales_invoices si
            LEFT JOIN sales_invoice_items sii ON si.invoice_id = sii.invoice_id
            LEFT JOIN products p ON sii.product_id = p.product_id
            WHERE si.status = 'Completed' {date_filter}
            GROUP BY DATE(si.invoice_date)
            ORDER BY report_date DESC
        """
        cur.execute(daily_query, params)
        daily_breakdown = cur.fetchall()

        # Process daily breakdown
        daily_data = []
        for row in daily_breakdown:
            daily_revenue = float(row['daily_revenue']) if row['daily_revenue'] else 0.0
            daily_cogs = float(row['daily_cogs']) if row['daily_cogs'] else 0.0
            daily_discounts = float(row['daily_discounts']) if row['daily_discounts'] else 0.0
            daily_gross_profit = daily_revenue - daily_cogs
            daily_net_profit = daily_gross_profit - daily_discounts

            daily_data.append({
                'date': str(row['report_date']),
                'revenue': round(daily_revenue, 2),
                'cogs': round(daily_cogs, 2),
                'gross_profit': round(daily_gross_profit, 2),
                'expenses': round(daily_discounts, 2),
                'net_profit': round(daily_net_profit, 2),
                'margin': round((daily_net_profit / daily_revenue * 100) if daily_revenue > 0 else 0, 2)
            })

        return jsonify({
            'summary': {
                'revenue': round(total_revenue, 2),
                'tax_collected': round(tax_collected, 2),
                'cogs': round(cost_of_goods_sold, 2),
                'gross_profit': round(gross_profit, 2),
                'gross_margin': round(gross_margin, 2),
                'expenses': round(total_expenses, 2),
                'net_profit': round(net_profit, 2),
                'net_margin': round(net_margin, 2),
            },
            'breakdown': {
                'discounts': round(discounts, 2),
                'returns': round(returns, 2),
            },
            'daily': daily_data
        })

    except Exception as e:
        print(f"Error generating P&L report: {str(e)}")
        return jsonify({'message': 'Failed to generate report', 'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)


@reports_bp.route('/reports/profit-loss/export', methods=['GET'])
@token_required
def export_profit_loss_report(payload):
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        params = []
        date_filter = ""
        if start_date:
            date_filter += " AND DATE(si.invoice_date) >= %s"
            params.append(start_date)
        if end_date:
            date_filter += " AND DATE(si.invoice_date) <= %s"
            params.append(end_date)

        # Revenue
        revenue_query = f"""
            SELECT
                COALESCE(SUM(si.total_amount), 0) as revenue,
                COALESCE(SUM(si.total_gst), 0) as tax_collected
            FROM sales_invoices si
            WHERE si.status = 'Completed' {date_filter}
        """
        cur.execute(revenue_query, params)
        revenue_data = cur.fetchone()
        total_revenue = float(revenue_data['revenue']) if revenue_data['revenue'] else 0.0
        tax_collected = float(revenue_data['tax_collected']) if revenue_data['tax_collected'] else 0.0

        # COGS
        cogs_query = f"""
            SELECT
                COALESCE(SUM(sii.quantity * p.purchase_rate), 0) as cogs
            FROM sales_invoice_items sii
            JOIN products p ON sii.product_id = p.product_id
            JOIN sales_invoices si ON sii.invoice_id = si.invoice_id
            WHERE si.status = 'Completed' {date_filter}
        """
        cur.execute(cogs_query, params)
        cogs_data = cur.fetchone()
        cost_of_goods_sold = float(cogs_data['cogs']) if cogs_data['cogs'] else 0.0

        # Gross Profit
        gross_profit = total_revenue - cost_of_goods_sold
        gross_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0

        # Expenses
        expenses_query = f"""
            SELECT
                COALESCE(SUM(si.discount_amount), 0) as discounts,
                COALESCE(SUM(sr.total_amount), 0) as returns
            FROM sales_invoices si
            LEFT JOIN sales_returns sr ON si.invoice_id = sr.original_invoice_id
            WHERE si.status = 'Completed' {date_filter}
        """
        cur.execute(expenses_query, params)
        expenses_data = cur.fetchone()
        discounts = float(expenses_data['discounts']) if expenses_data['discounts'] else 0.0
        returns = float(expenses_data['returns']) if expenses_data['returns'] else 0.0
        total_expenses = discounts + returns

        # Net Profit
        net_profit = gross_profit - total_expenses
        net_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0

        # Daily breakdown with detailed metrics
        daily_query = f"""
            SELECT
                DATE(si.invoice_date) as report_date,
                COUNT(DISTINCT si.invoice_id) as invoice_count,
                COALESCE(SUM(si.total_amount), 0) as daily_revenue,
                COALESCE(SUM(si.total_gst), 0) as daily_gst,
                COALESCE(SUM(sii.quantity * p.purchase_rate), 0) as daily_cogs,
                COALESCE(SUM(si.discount_amount), 0) as daily_discounts,
                COALESCE(SUM(sr.total_amount), 0) as daily_returns
            FROM sales_invoices si
            LEFT JOIN sales_invoice_items sii ON si.invoice_id = sii.invoice_id
            LEFT JOIN products p ON sii.product_id = p.product_id
            LEFT JOIN sales_returns sr ON si.invoice_id = sr.original_invoice_id
            WHERE si.status = 'Completed' {date_filter}
            GROUP BY DATE(si.invoice_date)
            ORDER BY report_date DESC
        """
        cur.execute(daily_query, params)
        daily_breakdown = cur.fetchall()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "P&L Report"

        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        summary_font = Font(bold=True, size=11)
        summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        total_font = Font(bold=True, color="FFFFFF")
        currency_format = '₹#,##0.00'
        percent_format = '0.00"%"'
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Title
        ws['A1'] = "PROFIT & LOSS STATEMENT"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')

        # Date range
        ws['A2'] = f"Period: {start_date} to {end_date}"
        ws.merge_cells('A2:D2')

        # P&L Statement
        ws['A4'] = "P&L STATEMENT"
        ws['A4'].font = header_font
        ws['A4'].fill = header_fill
        ws.merge_cells('A4:D4')

        row = 5
        # Revenue
        ws[f'A{row}'] = "REVENUE"
        ws[f'B{row}'] = total_revenue
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = summary_font
        row += 1

        ws[f'A{row}'] = "Tax Collected (GST)"
        ws[f'B{row}'] = tax_collected
        ws[f'B{row}'].number_format = currency_format
        row += 1

        # COGS
        ws[f'A{row}'] = "COST OF GOODS SOLD"
        ws[f'B{row}'] = cost_of_goods_sold
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = summary_font
        ws[f'A{row}'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        row += 1

        # Gross Profit
        ws[f'A{row}'] = "GROSS PROFIT"
        ws[f'B{row}'] = gross_profit
        ws[f'B{row}'].number_format = currency_format
        ws[f'C{row}'] = gross_margin
        ws[f'C{row}'].number_format = percent_format
        ws[f'A{row}'].font = summary_font
        ws[f'A{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        row += 1

        # Expenses breakdown
        ws[f'A{row}'] = "OPERATING EXPENSES"
        ws[f'A{row}'].font = summary_font
        row += 1

        ws[f'A{row}'] = "  Discounts Given"
        ws[f'B{row}'] = discounts
        ws[f'B{row}'].number_format = currency_format
        row += 1

        ws[f'A{row}'] = "  Returns & Refunds"
        ws[f'B{row}'] = returns
        ws[f'B{row}'].number_format = currency_format
        row += 1

        # Net Profit
        row += 1
        ws[f'A{row}'] = "NET PROFIT"
        ws[f'B{row}'] = net_profit
        ws[f'B{row}'].number_format = currency_format
        ws[f'C{row}'] = net_margin
        ws[f'C{row}'].number_format = percent_format
        ws[f'A{row}'].font = total_font
        ws[f'A{row}'].fill = total_fill
        ws[f'B{row}'].font = total_font
        ws[f'B{row}'].fill = total_fill
        ws[f'C{row}'].font = total_font
        ws[f'C{row}'].fill = total_fill

        # Daily Breakdown Sheet
        if daily_breakdown:
            ws2 = wb.create_sheet("Daily Breakdown")
            ws2['A1'] = "DAILY PROFIT & LOSS BREAKDOWN"
            ws2['A1'].font = Font(bold=True, size=12)
            ws2.merge_cells('A1:L1')

            # Headers
            headers = ['Date', 'Invoices', 'Revenue', 'GST', 'COGS', 'Gross Profit', 'Gross %', 'Discounts', 'Returns', 'Expenses', 'Net Profit', 'Net %']
            ws2.append(headers)
            for col_num, header in enumerate(headers, 1):
                cell = ws2.cell(row=2, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

            # Daily data
            row_num = 3
            for row_data in daily_breakdown:
                daily_revenue = float(row_data['daily_revenue']) if row_data['daily_revenue'] else 0.0
                daily_gst = float(row_data['daily_gst']) if row_data['daily_gst'] else 0.0
                daily_cogs = float(row_data['daily_cogs']) if row_data['daily_cogs'] else 0.0
                daily_discounts = float(row_data['daily_discounts']) if row_data['daily_discounts'] else 0.0
                daily_returns = float(row_data['daily_returns']) if row_data['daily_returns'] else 0.0
                daily_gross_profit = daily_revenue - daily_cogs
                daily_gross_margin = (daily_gross_profit / daily_revenue * 100) if daily_revenue > 0 else 0
                daily_expenses = daily_discounts + daily_returns
                daily_net_profit = daily_gross_profit - daily_expenses
                daily_net_margin = (daily_net_profit / daily_revenue * 100) if daily_revenue > 0 else 0
                invoice_count = int(row_data['invoice_count']) if row_data['invoice_count'] else 0

                ws2[f'A{row_num}'] = str(row_data['report_date'])
                ws2[f'B{row_num}'] = invoice_count
                ws2[f'C{row_num}'] = daily_revenue
                ws2[f'D{row_num}'] = daily_gst
                ws2[f'E{row_num}'] = daily_cogs
                ws2[f'F{row_num}'] = daily_gross_profit
                ws2[f'G{row_num}'] = daily_gross_margin
                ws2[f'H{row_num}'] = daily_discounts
                ws2[f'I{row_num}'] = daily_returns
                ws2[f'J{row_num}'] = daily_expenses
                ws2[f'K{row_num}'] = daily_net_profit
                ws2[f'L{row_num}'] = daily_net_margin

                # Format columns
                ws2[f'B{row_num}'].number_format = '0'
                for col in ['C', 'D', 'E', 'F', 'H', 'I', 'J', 'K']:
                    ws2[f'{col}{row_num}'].number_format = currency_format
                for col in ['G', 'L']:
                    ws2[f'{col}{row_num}'].number_format = percent_format

                # Highlight net profit
                if daily_net_profit >= 0:
                    ws2[f'K{row_num}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    ws2[f'K{row_num}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                row_num += 1

            # Auto-adjust column widths
            col_widths = {'A': 12, 'B': 10, 'C': 14, 'D': 12, 'E': 14, 'F': 14, 'G': 10, 'H': 12, 'I': 12, 'J': 12, 'K': 14, 'L': 10}
            for col, width in col_widths.items():
                ws2.column_dimensions[col].width = width

        # Auto-adjust column widths for main sheet
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        # Generate file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"profit_loss_{start_date}_to_{end_date}.xlsx"
        response = Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

        return response

    except Exception as e:
        print(f"Error exporting P&L report: {str(e)}")
        return jsonify({'message': 'Failed to export P&L report', 'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)