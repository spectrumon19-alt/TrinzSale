from flask import Blueprint, request, jsonify, send_file
from auth import token_required, permission_required
import os, io, uuid, json, csv, re, base64, traceback, requests
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

ocr_excel_bp = Blueprint('ocr_excel', __name__)

def _app_base() -> str:
    # TRINTZ_APP_BASE is set by launcher.py before any imports when running from zip.
    # Falls back to the project root in normal dev mode.
    base = os.environ.get('TRINTZ_APP_BASE')
    if base:
        return base
    f = __file__
    if '.zip' in f.replace('\\', '/'):
        idx = f.lower().find('.zip')
        return os.path.dirname(f[:idx + 4])
    return os.path.dirname(os.path.dirname(os.path.abspath(f)))

UPLOAD_DIR = os.path.join(_app_base(), 'tmp_ocr')
os.makedirs(UPLOAD_DIR, exist_ok=True)

MAX_IMAGE_BYTES = 20 * 1024 * 1024   # 20 MB
MAX_EXCEL_BYTES = 15 * 1024 * 1024   # 15 MB

ALLOWED_IMAGE = {'png', 'jpg', 'jpeg', 'webp', 'gif', 'bmp', 'tiff', 'tif'}
ALLOWED_EXCEL = {'xlsx', 'xls', 'csv'}
ALLOWED_ALL   = ALLOWED_IMAGE | ALLOWED_EXCEL

MIME_MAP = {
    'png': 'image/png', 'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
    'webp': 'image/webp', 'gif': 'image/gif',
    'bmp': 'image/bmp',  'tiff': 'image/tiff', 'tif': 'image/tiff',
}

# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _ext(filename):
    return filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''


def _safe_token():
    return uuid.uuid4().hex


def _cleanup_old(max_age_seconds=3600):
    """Remove temp files older than max_age_seconds."""
    try:
        now = datetime.now().timestamp()
        for fname in os.listdir(UPLOAD_DIR):
            fpath = os.path.join(UPLOAD_DIR, fname)
            if os.path.isfile(fpath) and (now - os.path.getmtime(fpath)) > max_age_seconds:
                os.remove(fpath)
    except Exception:
        pass


def _is_numeric(val):
    try:
        float(str(val).replace(',', '').replace('₹', '').replace('$', '').replace('%', '').strip())
        return True
    except (ValueError, TypeError):
        return False


def _to_numeric(val):
    try:
        return float(str(val).replace(',', '').replace('₹', '').replace('$', '').replace('%', '').strip())
    except (ValueError, TypeError):
        return None


def _detect_col_types(headers, rows):
    """Return list of 'number' | 'text' for each column."""
    types = []
    for ci in range(len(headers)):
        vals = [row[ci] for row in rows if ci < len(row) and str(row[ci]).strip()]
        numeric_count = sum(1 for v in vals if _is_numeric(v))
        types.append('number' if vals and numeric_count / len(vals) >= 0.7 else 'text')
    return types


# ──────────────────────────────────────────────────────────────────────────────
# Multi-provider vision OCR
# ──────────────────────────────────────────────────────────────────────────────

_OCR_PROMPT = (
    'Extract ALL tabular data visible in this image.\n'
    'Rules:\n'
    '- Return ONLY a JSON object — no markdown, no explanation.\n'
    '- If multiple tables exist, merge them or return the largest/most complete one.\n'
    '- Clean obvious OCR artifacts and fix spelling in column headers.\n'
    '- Numbers: return clean numeric strings (strip stray spaces, keep decimal points).\n'
    '- Preserve currency symbols or units as part of the value (e.g. "₹1200", "5 kg").\n'
    '- If a cell is empty, use an empty string "".\n'
    'Required JSON format:\n'
    '{"headers": ["Col1", "Col2", ...], "rows": [["val", "val", ...], ...]}\n'
)

# Which models support vision per provider
_VISION_ALL      = {'anthropic', 'gemini', 'minimax'}
_VISION_KEYWORDS = {
    'openai':   ('gpt-4o', 'gpt-4-turbo', 'gpt-4-vision'),
    'groq':     ('vision',),
    'ollama':   ('llava', 'bakllava', 'moondream', 'minicpm-v',
                 'llama3.2-vision', 'llama-3.2-vision', 'llama3.2:'),
    'mimo':     ('mimo-v2.5', 'mimo-v2-omni',
                 'xiaomi/mimo-v2.5', 'xiaomi/mimo-v2-omni'),
    'mistral':  ('pixtral',),
    'together': ('llava', 'vision', 'llama-3.2', 'llama3.2'),
}


def _is_vision_capable(provider: str, model: str) -> bool:
    if provider in _VISION_ALL:
        return True
    kws = _VISION_KEYWORDS.get(provider)
    if kws is None:
        return False
    ml = model.lower()
    return any(k in ml for k in kws)


def _get_active_ocr_config() -> dict:
    """Return the active AI config from DB; fall back to ANTHROPIC_API_KEY env var."""
    try:
        from routes.ai_settings import _ensure_table
        from db import get_db_connection as _gc, release_db_connection as _rc
        from psycopg2.extras import RealDictCursor as _RDC
        conn = _gc()
        try:
            with conn.cursor(cursor_factory=_RDC) as cur:
                _ensure_table(conn, cur)
                cur.execute('SELECT * FROM ai_settings WHERE is_active=TRUE LIMIT 1')
                row = cur.fetchone()
        finally:
            _rc(conn)
        if row:
            return dict(row)
    except Exception:
        pass

    key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
    if not key:
        raise ValueError(
            'No AI provider is configured. '
            'Go to Admin → AI Settings to configure a vision-capable AI provider, '
            'or add ANTHROPIC_API_KEY to your .env file.'
        )
    return {
        'provider': 'anthropic',
        'api_key': key,
        'model': 'claude-opus-4-5',
        'api_base_url': '',
        'extra_config': {},
    }


def _ocr_image(image_path: str, ext: str):
    """OCR an image using the active AI provider configured in AI Settings."""
    config   = _get_active_ocr_config()
    provider = config.get('provider', 'anthropic')
    model    = config.get('model', '')

    if not _is_vision_capable(provider, model):
        from routes.ai_settings import PROVIDERS
        prov_name = PROVIDERS.get(provider, {}).get('name', provider)
        raise ValueError(
            f'The active AI model "{model}" ({prov_name}) does not support image OCR. '
            f'Please select a vision-capable model in Admin → AI Settings. '
            f'Vision-capable options: any Anthropic Claude, GPT-4o, Gemini, MiMo v2.5/omni, '
            f'Groq vision preview, Ollama llava/moondream, or Mistral pixtral.'
        )

    api_key  = config.get('api_key', '')
    base_url = (config.get('api_base_url') or '').rstrip('/')

    with open(image_path, 'rb') as f:
        image_b64 = base64.standard_b64encode(f.read()).decode('utf-8')
    media_type = MIME_MAP.get(ext, 'image/jpeg')

    raw = _call_vision(provider, api_key, model, base_url, image_b64, media_type)

    # Strip markdown fences
    raw = re.sub(r'^```(?:json)?\s*', '', raw, flags=re.MULTILINE)
    raw = re.sub(r'```\s*$', '', raw, flags=re.MULTILINE)
    raw = raw.strip()

    # Extract just the outermost {...} block in case the model added commentary
    m = re.search(r'\{[\s\S]*\}', raw)
    if m:
        raw = m.group(0)

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ValueError(
            f'AI returned unreadable data (JSON parse error: {exc}). '
            f'The table may be too large — try cropping the image or using a model with a '
            f'higher token limit, then re-upload.'
        ) from exc

    headers = [str(h).strip() for h in data.get('headers', [])]
    rows    = [[str(c).strip() for c in row] for row in data.get('rows', [])]
    if not headers:
        raise ValueError('No table headers detected in the image.')
    return headers, rows


def _call_vision(provider, api_key, model, base_url, image_b64, media_type) -> str:
    """Dispatch vision/image call to the correct provider API."""

    if provider == 'anthropic':
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model=model, max_tokens=8192,
            messages=[{'role': 'user', 'content': [
                {'type': 'image',
                 'source': {'type': 'base64', 'media_type': media_type, 'data': image_b64}},
                {'type': 'text', 'text': _OCR_PROMPT},
            ]}],
        )
        return msg.content[0].text.strip()

    elif provider in ('openai', 'groq', 'mistral', 'mimo', 'together'):
        _base = {
            'openai':   base_url or 'https://api.openai.com/v1',
            'groq':     'https://api.groq.com/openai/v1',
            'mistral':  'https://api.mistral.ai/v1',
            'mimo':     base_url or 'https://api.xiaomimimo.com/v1',
            'together': 'https://api.together.xyz/v1',
        }[provider]
        r = requests.post(
            f'{_base}/chat/completions',
            headers={'Authorization': f'Bearer {api_key}',
                     'Content-Type': 'application/json'},
            json={
                'model': model,
                'messages': [{'role': 'user', 'content': [
                    {'type': 'image_url',
                     'image_url': {'url': f'data:{media_type};base64,{image_b64}'}},
                    {'type': 'text', 'text': _OCR_PROMPT},
                ]}],
                'max_tokens': 8192,
            },
            timeout=90,
        )
        r.raise_for_status()
        return r.json()['choices'][0]['message']['content'].strip()

    elif provider == 'gemini':
        r = requests.post(
            f'https://generativelanguage.googleapis.com/v1beta/models/'
            f'{model}:generateContent?key={api_key}',
            json={
                'contents': [{'parts': [
                    {'inline_data': {'mime_type': media_type, 'data': image_b64}},
                    {'text': _OCR_PROMPT},
                ]}],
                'generationConfig': {'maxOutputTokens': 4096, 'temperature': 0.1},
            },
            timeout=90,
        )
        r.raise_for_status()
        return r.json()['candidates'][0]['content']['parts'][0]['text'].strip()

    elif provider == 'ollama':
        base = base_url or 'http://localhost:11434'
        hdrs = {'Content-Type': 'application/json'}
        if api_key:
            hdrs['Authorization'] = f'Bearer {api_key}'
        r = requests.post(
            f'{base}/api/generate',
            headers=hdrs,
            json={
                'model': model,
                'prompt': _OCR_PROMPT,
                'images': [image_b64],
                'stream': False,
            },
            timeout=180,
        )
        r.raise_for_status()
        return r.json()['response'].strip()

    elif provider == 'minimax':
        r = requests.post(
            'https://api.minimax.chat/v1/text/chatcompletion_v2',
            headers={'Authorization': f'Bearer {api_key}',
                     'Content-Type': 'application/json'},
            json={
                'model': model,
                'messages': [{'role': 'user', 'content': [
                    {'type': 'image_url',
                     'image_url': {'url': f'data:{media_type};base64,{image_b64}'}},
                    {'type': 'text', 'text': _OCR_PROMPT},
                ]}],
                'max_tokens': 8192,
            },
            timeout=90,
        )
        r.raise_for_status()
        return r.json()['choices'][0]['message']['content'].strip()

    else:
        raise ValueError(f'Provider "{provider}" is not supported for image OCR.')


# ──────────────────────────────────────────────────────────────────────────────
# Excel / CSV reading
# ──────────────────────────────────────────────────────────────────────────────

def _read_excel(path, ext):
    if ext == 'csv':
        with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
            reader = csv.reader(f)
            all_rows = [row for row in reader if any(c.strip() for c in row)]
        if not all_rows:
            raise ValueError('CSV file is empty.')
        headers = [str(h).strip() for h in all_rows[0]]
        rows    = [[str(c).strip() for c in r] for r in all_rows[1:]]
    else:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else '' for c in row]
            if any(c for c in cells):
                all_rows.append(cells)
        wb.close()
        if not all_rows:
            raise ValueError('Excel file is empty.')
        headers = all_rows[0]
        rows    = all_rows[1:]

    # Pad short rows
    ncols = len(headers)
    rows = [(r + [''] * ncols)[:ncols] for r in rows]
    return headers, rows


# ──────────────────────────────────────────────────────────────────────────────
# Professional Excel generation
# ──────────────────────────────────────────────────────────────────────────────

# Palette
C_TITLE_BG   = '1D4ED8'   # rich blue   — company / title bar
C_TITLE_FG   = 'FFFFFF'
C_HEADER_BG  = '1E3A5F'   # deep navy   — column headers
C_HEADER_FG  = 'FFFFFF'
C_ALT        = 'EEF2FF'   # lavender-white — alternating rows
C_WHITE      = 'FFFFFF'
C_TOTAL_BG   = 'DBEAFE'   # light blue   — total / summary row
C_TOTAL_FG   = '1E3A5F'
C_BORDER     = 'CBD5E1'   # slate-200   — cell borders
C_HEADER_BDR = '1E3A5F'   # navy        — heavy header border


def _side(color, style='thin'):
    return Side(style=style, color=color)


def _make_professional_excel(headers, rows, cfg):
    """
    Generate a professionally styled .xlsx workbook.
    cfg keys: title, company, date, sheet_name, show_total, author
    """
    wb = Workbook()
    ws = wb.active
    ws.title = (cfg.get('sheet_name') or 'Data')[:31]

    ncols       = len(headers)
    col_types   = _detect_col_types(headers, rows)

    thin   = _side(C_BORDER)
    medium = _side(C_HEADER_BDR, 'medium')
    thick  = _side(C_HEADER_BDR, 'thick')

    header_border = Border(left=thin, right=thin, top=thick,   bottom=thick)
    cell_border   = Border(left=thin, right=thin, top=thin,    bottom=thin)
    total_border  = Border(left=thin, right=thin, top=medium,  bottom=thick)

    r = 1   # current write row

    # ── Company banner ────────────────────────────────────────────────────────
    company = (cfg.get('company') or '').strip()
    if company:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(row=r, column=1, value=company)
        c.font      = Font(name='Calibri', bold=True, size=15, color=C_TITLE_FG)
        c.fill      = PatternFill('solid', fgColor=C_TITLE_BG)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 30
        r += 1

    # ── Report title ──────────────────────────────────────────────────────────
    title = (cfg.get('title') or '').strip()
    if title:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(row=r, column=1, value=title)
        c.font      = Font(name='Calibri', bold=True, size=12, color=C_HEADER_BG)
        c.fill      = PatternFill('solid', fgColor='DBEAFE')
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 26
        r += 1

    # ── Meta row (date / generated by) ────────────────────────────────────────
    date_str  = (cfg.get('date') or datetime.now().strftime('%d %B %Y')).strip()
    author    = (cfg.get('author') or '').strip()
    meta_parts = [f'Date: {date_str}']
    if author:
        meta_parts.append(f'Prepared by: {author}')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value='   '.join(meta_parts))
    c.font      = Font(name='Calibri', italic=True, size=9, color='64748B')
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.fill      = PatternFill('solid', fgColor='F8FAFC')
    ws.row_dimensions[r].height = 16
    r += 1

    # ── Spacer ────────────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 5
    r += 1

    header_row = r   # for freeze_panes

    # ── Column headers ────────────────────────────────────────────────────────
    header_fill = PatternFill('solid', fgColor=C_HEADER_BG)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=ci, value=str(h))
        c.font      = Font(name='Calibri', bold=True, size=10, color=C_HEADER_FG)
        c.fill      = header_fill
        c.border    = header_border
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 22
    r += 1

    # ── Data rows ─────────────────────────────────────────────────────────────
    col_totals = {ci: 0.0 for ci, t in enumerate(col_types) if t == 'number'}
    any_totals = False

    for ri, row in enumerate(rows):
        is_alt  = ri % 2 == 1
        row_bg  = C_ALT if is_alt else C_WHITE
        row_fill = PatternFill('solid', fgColor=row_bg)

        for ci, val in enumerate(row[:ncols], 1):
            c = ws.cell(row=r, column=ci)
            c.fill   = row_fill
            c.border = cell_border
            c.font   = Font(name='Calibri', size=10)

            if col_types[ci - 1] == 'number':
                num = _to_numeric(val)
                if num is not None:
                    c.value          = num
                    c.number_format  = '#,##0.00'
                    c.alignment      = Alignment(horizontal='right',  vertical='center')
                    col_totals[ci - 1] += num
                    any_totals = True
                else:
                    c.value     = val
                    c.alignment = Alignment(horizontal='right', vertical='center')
            else:
                c.value     = val
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

        ws.row_dimensions[r].height = 18
        r += 1

    # ── Total / summary row ───────────────────────────────────────────────────
    if cfg.get('show_total') and any_totals:
        total_fill = PatternFill('solid', fgColor=C_TOTAL_BG)
        for ci in range(1, ncols + 1):
            c = ws.cell(row=r, column=ci)
            c.fill   = total_fill
            c.border = total_border
            c.font   = Font(name='Calibri', bold=True, size=10, color=C_TOTAL_FG)

            if ci == 1:
                c.value     = 'TOTAL'
                c.alignment = Alignment(horizontal='left', vertical='center')
            elif col_types[ci - 1] == 'number' and (ci - 1) in col_totals:
                c.value         = col_totals[ci - 1]
                c.number_format = '#,##0.00'
                c.alignment     = Alignment(horizontal='right', vertical='center')

        ws.row_dimensions[r].height = 22
        r += 1

    # ── Column widths (auto-fit, capped) ─────────────────────────────────────
    for ci, h in enumerate(headers, 1):
        col_vals  = [str(h)] + [str(row[ci - 1]) for row in rows if ci - 1 < len(row)]
        max_chars = max((len(v) for v in col_vals), default=8)
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_chars + 3, 10), 42)

    # ── Freeze panes (keep headers visible while scrolling) ──────────────────
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── Print settings ────────────────────────────────────────────────────────
    ws.print_title_rows = f'1:{header_row}'
    ws.page_setup.orientation    = 'landscape'
    ws.page_setup.fitToPage      = True
    ws.page_setup.fitToWidth     = 1
    ws.page_setup.fitToHeight    = 0
    ws.page_margins.left         = 0.5
    ws.page_margins.right        = 0.5
    ws.page_margins.top          = 0.75
    ws.page_margins.bottom       = 0.75

    # ── Workbook metadata ─────────────────────────────────────────────────────
    wb.properties.title   = title or 'Export'
    wb.properties.creator = author or 'TrintzPOS'
    wb.properties.created = datetime.now()

    return wb


# ──────────────────────────────────────────────────────────────────────────────
# Routes
# ──────────────────────────────────────────────────────────────────────────────

@ocr_excel_bp.route('/ocr-excel/ai-status', methods=['GET'])
@permission_required('ocr-excel')
def ai_status(payload):
    """Return the active AI config's provider, model, and vision capability."""
    try:
        config   = _get_active_ocr_config()
        provider = config.get('provider', 'anthropic')
        model    = config.get('model', '')
        vision   = _is_vision_capable(provider, model)
        source   = 'db' if config.get('id') else 'env'

        from routes.ai_settings import PROVIDERS
        meta = PROVIDERS.get(provider, {})

        return jsonify({
            'configured':     True,
            'provider':       provider,
            'provider_name':  meta.get('name', provider),
            'provider_icon':  meta.get('icon', '🤖'),
            'model':          model,
            'vision_capable': vision,
            'source':         source,
        })
    except Exception as e:
        return jsonify({
            'configured':     False,
            'vision_capable': False,
            'error':          str(e),
        })


@ocr_excel_bp.route('/ocr-excel/process', methods=['POST'])
@permission_required('ocr-excel')
def process_file(payload):
    """
    Upload a file (image or Excel/CSV).
    Returns extracted table data as JSON so the frontend can preview/edit it.
    """
    _cleanup_old()

    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400

    f    = request.files['file']
    name = f.filename or ''
    ext  = _ext(name)

    if ext not in ALLOWED_ALL:
        return jsonify({'success': False,
                        'message': f'Unsupported file type ".{ext}". Allowed: {", ".join(sorted(ALLOWED_ALL))}'}), 400

    raw = f.read()
    limit = MAX_IMAGE_BYTES if ext in ALLOWED_IMAGE else MAX_EXCEL_BYTES
    if len(raw) > limit:
        return jsonify({'success': False,
                        'message': f'File too large (max {limit // 1024 // 1024} MB)'}), 413

    # Save to temp dir
    job_id   = _safe_token()
    tmp_path = os.path.join(UPLOAD_DIR, f'{job_id}.{ext}')
    with open(tmp_path, 'wb') as out:
        out.write(raw)

    try:
        if ext in ALLOWED_IMAGE:
            headers, rows = _ocr_image(tmp_path, ext)
            source = 'ocr'
        else:
            headers, rows = _read_excel(tmp_path, ext)
            source = 'excel'
    except Exception as e:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        print(f'[ocr_excel] process error: {e}\n{traceback.format_exc()}')
        return jsonify({'success': False, 'message': str(e)}), 422

    col_types = _detect_col_types(headers, rows)

    return jsonify({
        'success':   True,
        'job_id':    job_id,
        'source':    source,
        'file_name': name,
        'headers':   headers,
        'rows':      rows,
        'col_types': col_types,
        'row_count': len(rows),
        'col_count': len(headers),
    }), 200


@ocr_excel_bp.route('/ocr-excel/generate', methods=['POST'])
@permission_required('ocr-excel')
def generate_excel(payload):
    """
    Accept edited table data + styling config, generate a professional .xlsx,
    save it to tmp, return a download token.
    """
    data = request.get_json(force=True)
    if not data:
        return jsonify({'success': False, 'message': 'No data'}), 400

    headers = data.get('headers') or []
    rows    = data.get('rows')    or []
    cfg     = data.get('config')  or {}

    if not headers:
        return jsonify({'success': False, 'message': 'No headers provided'}), 400

    try:
        wb = _make_professional_excel(headers, rows, cfg)
    except Exception as e:
        print(f'[ocr_excel] generate error: {e}\n{traceback.format_exc()}')
        return jsonify({'success': False, 'message': str(e)}), 500

    token     = _safe_token()
    safe_name = re.sub(r'[^\w\-]', '_', cfg.get('title') or cfg.get('sheet_name') or 'export')
    filename  = f'{safe_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    out_path  = os.path.join(UPLOAD_DIR, f'{token}_out.xlsx')

    wb.save(out_path)

    return jsonify({
        'success':  True,
        'token':    token,
        'filename': filename,
    }), 200


@ocr_excel_bp.route('/ocr-excel/download/<token>', methods=['GET'])
@permission_required('ocr-excel')
def download_excel(payload, token):
    """Serve a previously generated Excel file by token."""
    if not re.fullmatch(r'[0-9a-f]{32}', token):
        return jsonify({'success': False, 'message': 'Invalid token'}), 400

    path = os.path.join(UPLOAD_DIR, f'{token}_out.xlsx')
    if not os.path.exists(path):
        return jsonify({'success': False, 'message': 'File not found or expired'}), 404

    filename = request.args.get('filename', 'export.xlsx')

    return send_file(
        path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
