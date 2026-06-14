from flask import Blueprint, request, jsonify
from db import get_db_connection, release_db_connection
from auth import token_required, admin_required
from psycopg2.extras import RealDictCursor
from openpyxl import load_workbook
from io import BytesIO
import traceback

purchase_bp = Blueprint('purchase', __name__)

@purchase_bp.route('/purchase', methods=['POST'])
@admin_required
def create_purchase(payload):
    data = request.get_json(silent=True)
    user_id = payload.get('user_id')

    # ── Input validation (BUG-013) ───────────────────────────────────────────
    # A purchase order must have a body, at least one line item, and a supplier
    # reference that resolves to a real supplier. Without these the PO is
    # meaningless and silently corrupts inventory/payables data.
    if not data:
        return jsonify({'message': 'Request body is required'}), 400

    items = data.get('items')
    if not items or not isinstance(items, list) or len(items) == 0:
        return jsonify({'message': 'At least one item is required'}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Resolve and validate the supplier. The order may reference a supplier
        # by supplier_id or by supplier_gst_number; either way it must exist.
        supplier_id_in = data.get('supplier_id')
        supplier_gst = data.get('supplier_gst_number')
        if supplier_id_in is None and not supplier_gst:
            return jsonify({'message': 'A supplier (supplier_id or supplier_gst_number) is required'}), 400

        if supplier_id_in is not None:
            cur.execute("SELECT supplier_id FROM suppliers WHERE supplier_id = %s", (supplier_id_in,))
            if not cur.fetchone():
                return jsonify({'message': f'Supplier {supplier_id_in} does not exist'}), 404
        elif supplier_gst:
            cur.execute("SELECT supplier_id FROM suppliers WHERE supplier_gst_number = %s", (supplier_gst,))
            if not cur.fetchone():
                return jsonify({'message': f'Supplier with GST {supplier_gst} does not exist'}), 404

        # Check if purchase order number already exists
        purchase_order_number = data.get('purchase_order_number')
        if purchase_order_number and purchase_order_number.strip():
            cur.execute("""
                SELECT purchase_order_id FROM purchase_orders 
                WHERE purchase_order_number = %s
            """, (purchase_order_number,))
            
            existing_order = cur.fetchone()
            if existing_order:
                return jsonify({'message': f'Purchase order number {purchase_order_number} already exists. Please use a unique purchase order number.'}), 400
        else:
            # If no purchase order number provided, auto-generate one
            from datetime import datetime
            date_prefix = datetime.now().strftime('%Y%m%d')
            cur.execute("""
                SELECT MAX(CAST(SUBSTRING(purchase_order_number FROM 9) AS INTEGER)) as max_num
                FROM purchase_orders 
                WHERE purchase_order_number LIKE %s
            """, (f'{date_prefix}%',))
            
            result = cur.fetchone()
            next_num = (result['max_num'] or 0) + 1 if result else 1
            purchase_order_number = f"{date_prefix}{next_num:04d}"
        
        # Resolve supplier_id (validated above to exist). Prefer an explicit
        # supplier_id, else look up by GST number.
        supplier_id = supplier_id_in
        if supplier_id is None and supplier_gst:
            cur.execute("""
                SELECT supplier_id FROM suppliers WHERE supplier_gst_number = %s
            """, (supplier_gst,))
            supplier_result = cur.fetchone()
            if supplier_result:
                supplier_id = supplier_result['supplier_id']

        # Create purchase order
        cur.execute("""
            INSERT INTO purchase_orders (
                supplier_id, supplier_name, supplier_gst_number, purchase_order_number, purchase_date,
                user_id, total_amount, status
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING purchase_order_id
        """, (
            supplier_id,
            data.get('supplier_name'),
            data.get('supplier_gst_number'),
            purchase_order_number,
            data.get('purchase_date'),
            user_id,
            0,  # Will be updated after calculating items
            'Completed'
        ))
        
        purchase_order = cur.fetchone()
        if not purchase_order:
            raise Exception('Failed to create purchase order')
        purchase_order_id = purchase_order['purchase_order_id']
        
        # Process each item
        items = data.get('items', [])
        total_purchase_amount = 0
        
        for item in items:
            product_id = item.get('product_id')
            quantity = item.get('quantity')
            purchase_rate = item.get('purchase_rate')
            gst_rate = item.get('gst_rate', 0)
            taxable_value = item.get('taxable_value', 0)
            sgst = item.get('sgst', 0)
            cgst = item.get('cgst', 0)
            
            # Calculate item amount
            item_amount = quantity * purchase_rate
            total_purchase_amount += item_amount
            
            # Insert purchase item
            cur.execute("""
                INSERT INTO purchase_order_items (
                    purchase_order_id, product_id, quantity, purchase_rate, gst_rate, 
                    taxable_value, sgst, cgst, total_amount
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                purchase_order_id, product_id, quantity, purchase_rate, gst_rate,
                taxable_value, sgst, cgst, item_amount
            ))
            
            # Update inventory (increment stock)
            cur.execute("""
                UPDATE inventory 
                SET stock_quantity = stock_quantity + %s 
                WHERE product_id = %s
            """, (quantity, product_id))
            
            # Check if update was successful
            if cur.rowcount == 0:
                # If product doesn't exist in inventory, create entry
                cur.execute("""
                    INSERT INTO inventory (product_id, stock_quantity)
                    VALUES (%s, %s)
                """, (product_id, quantity))
        
        # Update purchase order with calculated total
        cur.execute("""
            UPDATE purchase_orders 
            SET total_amount = %s
            WHERE purchase_order_id = %s
        """, (total_purchase_amount, purchase_order_id))
        
        # Commit transaction
        conn.commit()
        
        # Fetch the complete purchase order with items (simplified query)
        cur.execute("""
            SELECT 
                po.*,
                u.username as created_by
            FROM purchase_orders po
            JOIN users u ON po.user_id = u.user_id
            WHERE po.purchase_order_id = %s
        """, (purchase_order_id,))
        
        purchase_order_result = cur.fetchone()
        
        # Fetch items separately
        cur.execute("""
            SELECT 
                poi.*,
                p.name
            FROM purchase_order_items poi
            JOIN products p ON poi.product_id = p.product_id
            WHERE poi.purchase_order_id = %s
            ORDER BY poi.created_at
        """, (purchase_order_id,))
        
        items_result = cur.fetchall()
        
        # Combine results
        if purchase_order_result:
            purchase_order_result['items'] = items_result
        
        return jsonify(purchase_order_result), 201
    except Exception as e:
        if conn:
            conn.rollback()
        # Log the full traceback for debugging
        print(f"Error creating purchase order: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'message': 'Failed to create purchase order', 'error': str(e)}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            release_db_connection(conn)

@purchase_bp.route('/purchase/upload', methods=['POST'])
@admin_required
def upload_purchase_excel(payload):
    try:
        # Check if file is present in request
        if 'file' not in request.files:
            return jsonify({'message': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'message': 'No file selected'}), 400
        
        # Read Excel file using openpyxl instead of pandas
        file_content = file.read()
        workbook = load_workbook(filename=BytesIO(file_content), read_only=True)
        worksheet = workbook.active
        
        if worksheet is None:
            return jsonify({'message': 'Could not read Excel file'}), 400
        
        # Get headers from first row
        header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        if not header_row:
            return jsonify({'message': 'Excel file is empty'}), 400
        
        headers = list(header_row[0])
        
        # Validate required columns
        required_columns = ['Supplier Name', 'Supplier GST Number', 'Purchase Order Number', 'Purchase Date',
                           'Product Name', 'Quantity', 'Purchase Rate']

        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            return jsonify({'message': f'Missing required columns: {missing_columns}'}), 400

        # Get optional column indices
        gst_rate_idx = headers.index('GST %') if 'GST %' in headers else None
        sgst_idx = headers.index('SGST') if 'SGST' in headers else None
        cgst_idx = headers.index('CGST') if 'CGST' in headers else None

        # Group data by purchase order
        purchase_orders = {}

        # Process each row (skip header row)
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None for cell in row):  # Skip empty rows
                continue

            # Extract data using column indices
            try:
                supplier_name_idx = headers.index('Supplier Name')
                supplier_gst_idx = headers.index('Supplier GST Number')
                po_number_idx = headers.index('Purchase Order Number')
                purchase_date_idx = headers.index('Purchase Date')
                product_name_idx = headers.index('Product Name')
                quantity_idx = headers.index('Quantity')
                purchase_rate_idx = headers.index('Purchase Rate')
            except ValueError as e:
                return jsonify({'message': f'Error finding column index: {str(e)}'}), 400

            # Extract required data
            supplier_name = row[supplier_name_idx] if supplier_name_idx < len(row) else None
            supplier_gst_number = row[supplier_gst_idx] if supplier_gst_idx < len(row) else None
            po_number = row[po_number_idx] if po_number_idx < len(row) else None
            purchase_date = row[purchase_date_idx] if purchase_date_idx < len(row) else None
            product_name = row[product_name_idx] if product_name_idx < len(row) else None
            quantity = row[quantity_idx] if quantity_idx < len(row) else None
            purchase_rate = row[purchase_rate_idx] if purchase_rate_idx < len(row) else None

            # Extract optional data
            gst_rate = float(row[gst_rate_idx]) if gst_rate_idx is not None and gst_rate_idx < len(row) and row[gst_rate_idx] else 0
            sgst = float(row[sgst_idx]) if sgst_idx is not None and sgst_idx < len(row) and row[sgst_idx] else 0
            cgst = float(row[cgst_idx]) if cgst_idx is not None and cgst_idx < len(row) and row[cgst_idx] else 0

            # Skip rows with missing required data
            if not all([po_number, product_name, quantity, purchase_rate]):
                continue

            if po_number not in purchase_orders:
                purchase_orders[po_number] = {
                    'supplier_name': supplier_name,
                    'supplier_gst_number': supplier_gst_number,
                    'purchase_order_number': po_number,
                    'purchase_date': purchase_date,
                    'items': []
                }

            # Add item to purchase order
            purchase_orders[po_number]['items'].append({
                'product_name': product_name,
                'quantity': quantity,
                'purchase_rate': purchase_rate,
                'gst_rate': gst_rate,
                'sgst': sgst,
                'cgst': cgst
            })
        
        # Process each purchase order
        results = []
        user_id = payload.get('user_id')
        
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        try:
            for po_data in purchase_orders.values():
                # Get or create supplier
                cur.execute("""
                    INSERT INTO suppliers (supplier_name, supplier_gst_number)
                    VALUES (%s, %s)
                    ON CONFLICT (supplier_gst_number) 
                    DO UPDATE SET supplier_name = EXCLUDED.supplier_name
                    RETURNING supplier_id
                """, (po_data['supplier_name'], po_data['supplier_gst_number']))
                
                supplier = cur.fetchone()
                supplier_id = supplier['supplier_id'] if supplier else None
                
                # Create purchase order
                cur.execute("""
                    INSERT INTO purchase_orders (
                        supplier_id, supplier_name, supplier_gst_number, purchase_order_number, purchase_date,
                        user_id, total_amount, status
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING purchase_order_id
                """, (
                    supplier_id,
                    po_data['supplier_name'],
                    po_data['supplier_gst_number'],
                    po_data['purchase_order_number'],
                    po_data['purchase_date'],
                    user_id,
                    0,  # Will be updated after calculating items
                    'Completed'
                ))
                
                purchase_order = cur.fetchone()
                if not purchase_order:
                    raise Exception(f'Failed to create purchase order {po_data["purchase_order_number"]}')
                purchase_order_id = purchase_order['purchase_order_id']
                
                # Process each item
                total_purchase_amount = 0
                items_data = []
                
                for item in po_data['items']:
                    # Get product by name
                    cur.execute("""
                        SELECT product_id FROM products WHERE name = %s
                    """, (item['product_name'],))
                    
                    product = cur.fetchone()
                    if not product:
                        raise Exception(f'Product {item["product_name"]} not found')
                    
                    product_id = product['product_id']
                    quantity = item['quantity']
                    purchase_rate = item['purchase_rate']
                    gst_rate = item.get('gst_rate', 0)
                    sgst = item.get('sgst', 0)
                    cgst = item.get('cgst', 0)

                    # Calculate item amount
                    item_amount = quantity * purchase_rate
                    total_purchase_amount += item_amount

                    # Insert purchase item with GST details
                    cur.execute("""
                        INSERT INTO purchase_order_items (
                            purchase_order_id, product_id, quantity, purchase_rate, gst_rate,
                            sgst, cgst, total_amount
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING item_id
                    """, (
                        purchase_order_id, product_id, quantity, purchase_rate, gst_rate,
                        sgst, cgst, item_amount
                    ))

                    item_result = cur.fetchone()
                    items_data.append({
                        'item_id': item_result['item_id'] if item_result else None,
                        'product_id': product_id,
                        'product_name': item['product_name'],
                        'quantity': quantity,
                        'purchase_rate': purchase_rate,
                        'gst_rate': gst_rate,
                        'sgst': sgst,
                        'cgst': cgst,
                        'total_amount': item_amount
                    })
                    
                    # Update inventory (increment stock)
                    cur.execute("""
                        UPDATE inventory 
                        SET stock_quantity = stock_quantity + %s 
                        WHERE product_id = %s
                    """, (quantity, product_id))
                    
                    # Check if update was successful
                    if cur.rowcount == 0:
                        # If product doesn't exist in inventory, create entry
                        cur.execute("""
                            INSERT INTO inventory (product_id, stock_quantity)
                            VALUES (%s, %s)
                        """, (product_id, quantity))
                
                # Update purchase order with calculated total
                cur.execute("""
                    UPDATE purchase_orders 
                    SET total_amount = %s
                    WHERE purchase_order_id = %s
                """, (total_purchase_amount, purchase_order_id))
                
                results.append({
                    'purchase_order_id': purchase_order_id,
                    'purchase_order_number': po_data['purchase_order_number'],
                    'supplier_name': po_data['supplier_name'],
                    'supplier_gst_number': po_data['supplier_gst_number'],
                    'total_amount': total_purchase_amount,
                    'items_count': len(items_data)
                })
            
            # Commit transaction
            conn.commit()
            
            return jsonify({
                'message': f'Successfully uploaded {len(results)} purchase orders',
                'purchase_orders': results
            }), 201
            
        except Exception as e:
            if conn:
                conn.rollback()
            return jsonify({'message': 'Failed to process purchase orders', 'error': str(e)}), 500
        finally:
            if cur:
                cur.close()
            if conn:
                release_db_connection(conn)
                
    except Exception as e:
        return jsonify({'message': 'Failed to upload Excel file', 'error': str(e)}), 500

@purchase_bp.route('/purchase/<int:purchase_order_id>/cancel', methods=['PUT'])
@admin_required
def cancel_purchase(payload, purchase_order_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Check if purchase order exists and is not already cancelled
        cur.execute("""
            SELECT status FROM purchase_orders WHERE purchase_order_id = %s
        """, (purchase_order_id,))
        
        purchase_order = cur.fetchone()
        if not purchase_order:
            return jsonify({'message': 'Purchase order not found'}), 404
            
        if purchase_order['status'] == 'Cancelled':
            return jsonify({'message': 'Purchase order is already cancelled'}), 400
        
        # Get all items in the purchase order
        cur.execute("""
            SELECT product_id, quantity FROM purchase_order_items WHERE purchase_order_id = %s
        """, (purchase_order_id,))
        
        items = cur.fetchall()
        
        # Update inventory (decrement stock for each item)
        for item in items:
            cur.execute("""
                UPDATE inventory 
                SET stock_quantity = stock_quantity - %s 
                WHERE product_id = %s
            """, (item['quantity'], item['product_id']))
            
            # Check if update was successful
            if cur.rowcount == 0:
                # If no rows were updated, it means there's no inventory record for this product
                # Let's check if the product exists
                cur.execute("SELECT 1 FROM products WHERE product_id = %s", (item['product_id'],))
                product_exists = cur.fetchone()
                
                if not product_exists:
                    raise Exception(f"Product with ID {item['product_id']} does not exist")
                
                # If product exists but no inventory record, create one with 0 stock and then update
                try:
                    cur.execute("""
                        INSERT INTO inventory (product_id, stock_quantity) 
                        VALUES (%s, %s)
                    """, (item['product_id'], 0))  # Start with 0 stock
                    
                    # Now update the inventory
                    cur.execute("""
                        UPDATE inventory 
                        SET stock_quantity = stock_quantity - %s 
                        WHERE product_id = %s
                    """, (item['quantity'], item['product_id']))
                except Exception as insert_error:
                    # If insert fails due to constraint, try update again
                    cur.execute("""
                        UPDATE inventory 
                        SET stock_quantity = stock_quantity - %s 
                        WHERE product_id = %s
                    """, (item['quantity'], item['product_id']))
                    
                    if cur.rowcount == 0:
                        raise Exception(f"Failed to update inventory for product {item['product_id']}. No inventory record found and unable to create one.")
        
        # Update purchase order status to cancelled
        cur.execute("""
            UPDATE purchase_orders 
            SET status = 'Cancelled' 
            WHERE purchase_order_id = %s
        """, (purchase_order_id,))
        
        # Commit transaction
        conn.commit()
        
        return jsonify({'message': 'Purchase order cancelled successfully'}), 200
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'message': 'Failed to cancel purchase order', 'error': str(e)}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            release_db_connection(conn)

@purchase_bp.route('/purchase/<int:purchase_order_id>', methods=['PUT'])
@admin_required
def update_purchase(payload, purchase_order_id):
    data = request.get_json()
    user_id = payload.get('user_id')
    
    # Log the incoming data for debugging
    print(f"Update purchase data received: {data}")
    print(f"Items: {data.get('items', [])}")
    
    conn = None
    cur = None
    
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Check if purchase order exists
        cur.execute("""
            SELECT * FROM purchase_orders WHERE purchase_order_id = %s
        """, (purchase_order_id,))
        
        existing_order = cur.fetchone()
        if not existing_order:
            return jsonify({'message': 'Purchase order not found'}), 404
        
        # Check if purchase order number already exists (for other orders)
        purchase_order_number = data.get('purchase_order_number')
        if purchase_order_number:
            cur.execute("""
                SELECT purchase_order_id FROM purchase_orders 
                WHERE purchase_order_number = %s AND purchase_order_id != %s
            """, (purchase_order_number, purchase_order_id))
            
            duplicate_order = cur.fetchone()
            if duplicate_order:
                return jsonify({'message': f'Purchase order number {purchase_order_number} already exists. Please use a unique purchase order number.'}), 400
        else:
            # Use existing purchase order number if not provided
            purchase_order_number = existing_order['purchase_order_number']
        
        # Get supplier_id from suppliers table
        supplier_id = None
        if data.get('supplier_gst_number'):
            cur.execute("""
                SELECT supplier_id FROM suppliers WHERE supplier_gst_number = %s
            """, (data.get('supplier_gst_number'),))
            supplier_result = cur.fetchone()
            if supplier_result:
                supplier_id = supplier_result['supplier_id']
        
        # Get existing items to calculate inventory differences
        cur.execute("""
            SELECT product_id, quantity FROM purchase_order_items WHERE purchase_order_id = %s
        """, (purchase_order_id,))
        
        existing_items = cur.fetchall()
        existing_items_dict = {item['product_id']: item['quantity'] for item in existing_items}
        
        # Delete existing items
        try:
            cur.execute("""
                DELETE FROM purchase_order_items WHERE purchase_order_id = %s
            """, (purchase_order_id,))
        except Exception as e:
            raise Exception(f"Failed to delete existing items: {str(e)}")
        
        # Handle empty purchase_date - preserve existing date if empty
        purchase_date = data.get('purchase_date')
        if purchase_date == '' or purchase_date is None:
            purchase_date = existing_order.get('purchase_date')
        
        # Update purchase order
        try:
            cur.execute("""
                UPDATE purchase_orders 
                SET supplier_id = %s, supplier_name = %s, supplier_gst_number = %s, 
                    purchase_order_number = %s, purchase_date = %s, user_id = %s, total_amount = %s
                WHERE purchase_order_id = %s
            """, (
                supplier_id,
                data.get('supplier_name'),
                data.get('supplier_gst_number'),
                purchase_order_number,
                purchase_date,
                user_id,
                0,  # Will be updated after calculating items
                purchase_order_id
            ))
        except Exception as e:
            raise Exception(f"Failed to update purchase order header: {str(e)}")
        
        # Process each new item
        items = data.get('items', [])
        if not items:
            return jsonify({'message': 'At least one item is required'}), 400
            
        total_purchase_amount = 0
        new_items_dict = {}
        
        for idx, item in enumerate(items):
            # Validate and convert product_id
            product_id_raw = item.get('product_id')
            if product_id_raw is None or product_id_raw == '':
                return jsonify({'message': f'Item {idx + 1}: Product ID is required'}), 400
            try:
                product_id = int(product_id_raw)
            except (ValueError, TypeError):
                return jsonify({'message': f'Item {idx + 1}: Invalid Product ID: {product_id_raw}'}), 400
            
            # Validate quantity
            quantity = item.get('quantity')
            if quantity is None:
                return jsonify({'message': f'Item {idx + 1}: Quantity is required'}), 400
            try:
                quantity = float(quantity)
                if quantity <= 0:
                    return jsonify({'message': f'Item {idx + 1}: Quantity must be greater than 0'}), 400
            except (ValueError, TypeError):
                return jsonify({'message': f'Item {idx + 1}: Invalid quantity: {quantity}'}), 400
            
            # Validate purchase_rate
            purchase_rate = item.get('purchase_rate')
            if purchase_rate is None:
                return jsonify({'message': f'Item {idx + 1}: Purchase rate is required'}), 400
            try:
                purchase_rate = float(purchase_rate)
                if purchase_rate < 0:
                    return jsonify({'message': f'Item {idx + 1}: Purchase rate cannot be negative'}), 400
            except (ValueError, TypeError):
                return jsonify({'message': f'Item {idx + 1}: Invalid purchase rate: {purchase_rate}'}), 400
            
            gst_rate = float(item.get('gst_rate', 0) or 0)
            taxable_value = float(item.get('taxable_value', 0) or 0)
            sgst = float(item.get('sgst', 0) or 0)
            cgst = float(item.get('cgst', 0) or 0)
            
            # Calculate item amount
            item_amount = quantity * purchase_rate
            total_purchase_amount += item_amount
            
            # Store for inventory update (use integer product_id for consistent lookups)
            new_items_dict[product_id] = quantity
            
            # Insert purchase item
            try:
                cur.execute("""
                    INSERT INTO purchase_order_items (
                        purchase_order_id, product_id, quantity, purchase_rate, gst_rate, 
                        taxable_value, sgst, cgst, total_amount
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    purchase_order_id, product_id, quantity, purchase_rate, gst_rate,
                    taxable_value, sgst, cgst, item_amount
                ))
            except Exception as e:
                raise Exception(f"Failed to insert item for product {product_id}: {str(e)}")
            
            # Update inventory based on difference
            existing_quantity = existing_items_dict.get(product_id, 0)
            quantity_difference = quantity - existing_quantity
            
            try:
                if quantity_difference > 0:
                    # Increase inventory
                    cur.execute("""
                        UPDATE inventory 
                        SET stock_quantity = stock_quantity + %s 
                        WHERE product_id = %s
                    """, (quantity_difference, product_id))
                    
                    # Check if update was successful
                    if cur.rowcount == 0:
                        # If product doesn't exist in inventory, create entry
                        cur.execute("""
                            INSERT INTO inventory (product_id, stock_quantity)
                            VALUES (%s, %s)
                        """, (product_id, quantity_difference))
                elif quantity_difference < 0:
                    # Decrease inventory
                    cur.execute("""
                        UPDATE inventory 
                        SET stock_quantity = stock_quantity - %s 
                        WHERE product_id = %s
                    """, (abs(quantity_difference), product_id))
            except Exception as e:
                raise Exception(f"Failed to update inventory for product {product_id}: {str(e)}")
        
        # Handle items that were removed (in existing but not in new)
        try:
            for product_id, existing_quantity in existing_items_dict.items():
                if product_id not in new_items_dict:
                    # Remove all quantity from inventory
                    cur.execute("""
                        UPDATE inventory 
                        SET stock_quantity = stock_quantity - %s 
                        WHERE product_id = %s
                    """, (existing_quantity, product_id))
        except Exception as e:
            raise Exception(f"Failed to update inventory for removed items: {str(e)}")
        
        # Update purchase order with calculated total
        try:
            cur.execute("""
                UPDATE purchase_orders 
                SET total_amount = %s
                WHERE purchase_order_id = %s
            """, (total_purchase_amount, purchase_order_id))
        except Exception as e:
            raise Exception(f"Failed to update total amount: {str(e)}")
        
        # Commit transaction
        try:
            conn.commit()
        except Exception as e:
            raise Exception(f"Failed to commit transaction: {str(e)}")
        
        # Fetch the complete purchase order with items (simplified query)
        try:
            cur.execute("""
                SELECT 
                    po.*,
                    u.username as created_by
                FROM purchase_orders po
                JOIN users u ON po.user_id = u.user_id
                WHERE po.purchase_order_id = %s
            """, (purchase_order_id,))
            
            purchase_order_result = cur.fetchone()
            
            # Fetch items separately
            cur.execute("""
                SELECT 
                    poi.*,
                    p.name
                FROM purchase_order_items poi
                JOIN products p ON poi.product_id = p.product_id
                WHERE poi.purchase_order_id = %s
                ORDER BY poi.created_at
            """, (purchase_order_id,))
            
            items_result = cur.fetchall()
            
            # Combine results
            if purchase_order_result:
                purchase_order_result['items'] = items_result
            
            return jsonify(purchase_order_result), 200
        except Exception as e:
            raise Exception(f"Failed to fetch updated order: {str(e)}")
    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except Exception as rollback_error:
                print(f"Rollback error: {rollback_error}")
        # Log the full traceback for debugging
        print(f"Error updating purchase order: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'message': 'Failed to update purchase order', 'error': str(e)}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            release_db_connection(conn)

@purchase_bp.route('/purchase/<int:purchase_order_id>', methods=['DELETE'])
@admin_required
def delete_purchase(payload, purchase_order_id):
    conn = None
    cur = None
    
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Check if purchase order exists
        cur.execute("""
            SELECT status FROM purchase_orders WHERE purchase_order_id = %s
        """, (purchase_order_id,))
        
        purchase_order = cur.fetchone()
        if not purchase_order:
            return jsonify({'message': 'Purchase order not found'}), 404
            
        if purchase_order['status'] == 'Cancelled':
            return jsonify({'message': 'Purchase order is already cancelled'}), 400
        
        # Get all items in the purchase order
        cur.execute("""
            SELECT product_id, quantity FROM purchase_order_items WHERE purchase_order_id = %s
        """, (purchase_order_id,))
        
        items = cur.fetchall()
        
        # Update inventory (decrement stock for each item)
        for item in items:
            try:
                cur.execute("""
                    UPDATE inventory 
                    SET stock_quantity = stock_quantity - %s 
                    WHERE product_id = %s
                """, (item['quantity'], item['product_id']))
                
                # Check if update was successful
                if cur.rowcount == 0:
                    # If no rows were updated, it means there's no inventory record for this product
                    # Let's check if the product exists
                    cur.execute("SELECT 1 FROM products WHERE product_id = %s", (item['product_id'],))
                    product_exists = cur.fetchone()
                    
                    if not product_exists:
                        raise Exception(f"Product with ID {item['product_id']} does not exist")
                    
                    # If product exists but no inventory record, create one with 0 stock and then update
                    try:
                        cur.execute("""
                            INSERT INTO inventory (product_id, stock_quantity) 
                            VALUES (%s, %s)
                        """, (item['product_id'], 0))  # Start with 0 stock
                        
                        # Now update the inventory
                        cur.execute("""
                            UPDATE inventory 
                            SET stock_quantity = stock_quantity - %s 
                            WHERE product_id = %s
                        """, (item['quantity'], item['product_id']))
                    except Exception as insert_error:
                        # If insert fails due to constraint, try update again
                        cur.execute("""
                            UPDATE inventory 
                            SET stock_quantity = stock_quantity - %s 
                            WHERE product_id = %s
                        """, (item['quantity'], item['product_id']))
                        
                        if cur.rowcount == 0:
                            raise Exception(f"Failed to update inventory for product {item['product_id']}. No inventory record found and unable to create one.")
            except Exception as item_error:
                raise Exception(f"Error processing item with product ID {item['product_id']}: {str(item_error)}")
        
        # Delete purchase order (cascades to items due to foreign key constraint)
        cur.execute("""
            DELETE FROM purchase_orders 
            WHERE purchase_order_id = %s
        """, (purchase_order_id,))
        
        # Check if any rows were affected
        if cur.rowcount == 0:
            raise Exception("No purchase order was deleted. This may indicate a database constraint issue.")
        
        # Commit transaction
        conn.commit()
        
        return jsonify({'message': 'Purchase order deleted successfully'}), 200
    except Exception as e:
        if conn:
            conn.rollback()
        # Log the full error for debugging
        print(f"Error deleting purchase order {purchase_order_id}: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'message': 'Failed to delete purchase order', 'error': str(e)}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            release_db_connection(conn)

@purchase_bp.route('/purchase', methods=['GET'])
@token_required
def get_purchase_orders(payload):
    conn = None
    cur = None
    
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Optimized query that fetches purchase orders with items in a single query
        cur.execute("""
            SELECT 
                po.*,
                u.username as created_by,
                s.supplier_name,
                s.supplier_gst_number,
                json_agg(
                    json_build_object(
                        'item_id', poi.item_id,
                        'product_id', poi.product_id,
                        'product_name', p.name,
                        'quantity', poi.quantity,
                        'purchase_rate', poi.purchase_rate,
                        'gst_rate', poi.gst_rate,
                        'taxable_value', poi.taxable_value,
                        'sgst', poi.sgst,
                        'cgst', poi.cgst,
                        'total_amount', poi.total_amount,
                        'created_at', poi.created_at
                    ) ORDER BY poi.created_at
                ) FILTER (WHERE poi.item_id IS NOT NULL) as items
            FROM purchase_orders po
            JOIN users u ON po.user_id = u.user_id
            LEFT JOIN suppliers s ON po.supplier_id = s.supplier_id
            LEFT JOIN purchase_order_items poi ON po.purchase_order_id = poi.purchase_order_id
            LEFT JOIN products p ON poi.product_id = p.product_id
            GROUP BY po.purchase_order_id, u.username, s.supplier_name, s.supplier_gst_number
            ORDER BY po.purchase_date DESC, po.created_at DESC
        """)
        
        purchase_orders = cur.fetchall()
        
        # Process the results to ensure items is always an array
        for purchase_order in purchase_orders:
            if purchase_order['items'] is None:
                purchase_order['items'] = []
        
        return jsonify(purchase_orders), 200
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'message': 'Failed to fetch purchase orders', 'error': str(e)}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            release_db_connection(conn)
