from flask import Blueprint, request, jsonify, current_app
from db import get_db_connection, release_db_connection
from auth import token_required, admin_required
from psycopg2.extras import RealDictCursor
import os
import uuid
from datetime import datetime
from openpyxl import load_workbook
import tempfile

data_upload_bp = Blueprint('data_upload', __name__)

@data_upload_bp.route('/admin/upload-sales', methods=['POST'])
@admin_required
def upload_sales_data(payload):
    # Check if file is present in request
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file provided'}), 400
    
    file = request.files['file']
    
    # Check if file is selected
    if not file or not file.filename:
        return jsonify({'success': False, 'message': 'No file selected'}), 400
    
    # Check file extension
    filename = file.filename
    if not filename or not isinstance(filename, str) or not filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'}), 400
    
    temp_filename = None
    try:
        # Save file temporarily
        temp_dir = tempfile.gettempdir()
        file_extension = os.path.splitext(filename)[1] if filename else ''
        temp_filename = os.path.join(temp_dir, str(uuid.uuid4()) + file_extension)
        file.save(temp_filename)
        
        # Process the Excel file
        record_count = process_sales_excel(temp_filename)
        
        # Remove temporary file
        if temp_filename and os.path.exists(temp_filename):
            try:
                os.remove(temp_filename)
            except PermissionError:
                # On Windows, the file might still be locked by the system
                # This is a known issue, but it doesn't affect the upload process
                print(f"Warning: Could not delete temporary file {temp_filename} (file may be locked)")
            except Exception as e:
                print(f"Error cleaning up temp file: {str(e)}")
        
        return jsonify({
            'success': True, 
            'message': 'Sales data uploaded successfully',
            'record_count': record_count
        }), 200
        
    except Exception as e:
        # Remove temporary file if it exists
        if temp_filename and os.path.exists(temp_filename):
            try:
                os.remove(temp_filename)
            except PermissionError:
                # On Windows, the file might still be locked by the system
                # This is a known issue, but it doesn't affect the upload process
                print(f"Warning: Could not delete temporary file {temp_filename} (file may be locked)")
            except Exception as cleanup_error:
                print(f"Error cleaning up temp file: {str(cleanup_error)}")
        return jsonify({'success': False, 'message': f'Error processing file: {str(e)}'}), 500

@data_upload_bp.route('/admin/upload-purchase', methods=['POST'])
@admin_required
def upload_purchase_data(payload):
    # Check if file is present in request
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file provided'}), 400
    
    file = request.files['file']
    
    # Check if file is selected
    if not file or not file.filename:
        return jsonify({'success': False, 'message': 'No file selected'}), 400
    
    # Check file extension
    filename = file.filename
    if not filename or not isinstance(filename, str) or not filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'}), 400
    
    temp_filename = None
    try:
        # Save file temporarily
        temp_dir = tempfile.gettempdir()
        file_extension = os.path.splitext(filename)[1] if filename else ''
        temp_filename = os.path.join(temp_dir, str(uuid.uuid4()) + file_extension)
        file.save(temp_filename)
        
        # Process the Excel file
        record_count = process_purchase_excel(temp_filename)
        
        # Remove temporary file
        if temp_filename and os.path.exists(temp_filename):
            try:
                os.remove(temp_filename)
            except PermissionError:
                # On Windows, the file might still be locked by the system
                # This is a known issue, but it doesn't affect the upload process
                print(f"Warning: Could not delete temporary file {temp_filename} (file may be locked)")
            except Exception as e:
                print(f"Error cleaning up temp file: {str(e)}")
        
        return jsonify({
            'success': True, 
            'message': 'Purchase data uploaded successfully',
            'record_count': record_count
        }), 200
        
    except Exception as e:
        # Log the error for debugging
        print(f"Error in purchase upload: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Remove temporary file if it exists
        if temp_filename and os.path.exists(temp_filename):
            try:
                os.remove(temp_filename)
            except PermissionError:
                # On Windows, the file might still be locked by the system
                # This is a known issue, but it doesn't affect the upload process
                print(f"Warning: Could not delete temporary file {temp_filename} (file may be locked)")
            except Exception as cleanup_error:
                print(f"Error cleaning up temp file: {str(cleanup_error)}")
        
        # Return detailed error message
        error_message = str(e)
        if "Missing required column" in error_message:
            error_message = f"Excel format error: {error_message}"
        elif "Missing required data" in error_message:
            error_message = f"Data validation error: {error_message}"
        elif "not found" in error_message:
            error_message = f"Reference error: {error_message}"
        
        return jsonify({'success': False, 'message': error_message}), 500

def safe_float_convert(value, default=0.0):
    """Safely convert a value to float, handling various Excel cell types"""
    if value is None:
        return default
    try:
        # Handle various Excel cell types
        if hasattr(value, 'value'):
            value = value.value
        if value is None:
            return default
        return float(value)
    except (ValueError, TypeError):
        return default

def safe_int_convert(value, default=0):
    """Safely convert a value to int, handling various Excel cell types"""
    if value is None:
        return default
    try:
        # Handle various Excel cell types
        if hasattr(value, 'value'):
            value = value.value
        if value is None:
            return default
        return int(float(value))  # Convert to float first to handle decimal strings
    except (ValueError, TypeError):
        return default

def safe_str_convert(value, default=""):
    """Safely convert a value to string, handling various Excel cell types"""
    if value is None:
        return default
    try:
        # Handle various Excel cell types
        if hasattr(value, 'value'):
            value = value.value
        if value is None:
            return default
        return str(value)
    except (ValueError, TypeError):
        return default

def process_sales_excel(file_path):
    """Process sales Excel file and insert data into database"""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    record_count = 0
    workbook = None
    
    try:
        # Load workbook
        workbook = load_workbook(file_path, read_only=True)
        worksheet = workbook.active
        
        if worksheet is None:
            raise ValueError("Could not load worksheet from Excel file")
        
        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            headers.append(safe_str_convert(cell, ""))
        
        # Required columns
        required_columns = ['Invoice Number', 'Date', 'Customer Name', 'Product Name', 'Quantity', 'Rate', 'GST Rate']
        
        # Check if all required columns are present
        for col in required_columns:
            if col not in headers:
                raise ValueError(f"Missing required column: {col}")
        
        # Get column indices
        invoice_col = headers.index('Invoice Number')
        date_col = headers.index('Date')
        customer_col = headers.index('Customer Name')
        product_col = headers.index('Product Name')
        quantity_col = headers.index('Quantity')
        rate_col = headers.index('Rate')
        gst_rate_col = headers.index('GST Rate')
        
        # Process each row
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None for cell in row):  # Skip empty rows
                continue
                
            # Extract data
            invoice_number = safe_str_convert(row[invoice_col])
            date_str = safe_str_convert(row[date_col])
            customer_name = safe_str_convert(row[customer_col])
            product_name = safe_str_convert(row[product_col])
            quantity = safe_float_convert(row[quantity_col])
            rate = safe_float_convert(row[rate_col])
            gst_rate = safe_float_convert(row[gst_rate_col])
            
            # Validate required fields
            if not all([invoice_number.strip(), date_str.strip(), product_name.strip()]):
                raise ValueError(f"Missing required data in row {row_num}")
            
            # Convert date
            if isinstance(date_str, str):
                try:
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                except ValueError:
                    raise ValueError(f"Invalid date format in row {row_num}. Expected YYYY-MM-DD")
            else:
                date_obj = date_str
            
            # Get or create customer user (using admin user for now)
            cur.execute("SELECT user_id FROM users WHERE role = 'Admin' LIMIT 1")
            user_result = cur.fetchone()
            if not user_result:
                raise ValueError("No admin user found")
            user_id = user_result['user_id']
            
            # Get product
            cur.execute("SELECT product_id, gst_rate as product_gst_rate FROM products WHERE name = %s", (product_name,))
            product_result = cur.fetchone()
            if not product_result:
                raise ValueError(f"Product '{product_name}' not found in row {row_num}")
            
            product_id = product_result['product_id']
            product_gst_rate = safe_float_convert(product_result['product_gst_rate'])
            
            # Validate GST rate matches
            if gst_rate != product_gst_rate:
                raise ValueError(f"GST rate mismatch for product '{product_name}' in row {row_num}")
            
            # Calculate amounts
            line_amount = quantity * rate
            taxable_value = line_amount / (1 + (gst_rate / 100)) if gst_rate != 0 else line_amount
            gst_amount = line_amount - taxable_value
            sgst = gst_amount / 2
            cgst = gst_amount / 2
            
            # Check if invoice exists
            # Instead of using the uploaded invoice number directly, we'll generate a proper one
            # First, let's generate a proper invoice number using the same logic as in sales.py
            def generate_invoice_number_for_upload(conn, cur, date_obj):
                from datetime import datetime
                # Get date in YYMMDD format
                date_str = date_obj.strftime('%y%m%d')
                
                # Find all invoices with new format (DXXPYYY_YYMMDD) regardless of date
                cur.execute("SELECT invoice_number FROM sales_invoices")
                all_invoices = cur.fetchall()
                
                # Filter for invoices with new format (DXXPYYY_YYMMDD)
                matching_invoices = []
                for row in all_invoices:
                    invoice = row['invoice_number']
                    if (invoice and invoice.startswith('D') and 'P' in invoice and '_' in invoice and 
                        len(invoice) >= 11):
                        # Extract the numeric parts for comparison
                        try:
                            # Extract DXX (positions 1-3)
                            d_part = invoice[1:3]
                            # Extract PYYY (positions 4-7) 
                            p_part = invoice[4:7]
                            
                            d_value = int(d_part)
                            p_value = int(p_part)
                            
                            # Create a sortable key: D * 1000 + P (to sort by sequence)
                            sort_key = d_value * 1000 + p_value
                            matching_invoices.append((sort_key, invoice))
                        except (ValueError, IndexError):
                            # Skip invalid formats
                            continue
                
                # Sort matching invoices to find the highest
                if matching_invoices:
                    matching_invoices.sort(reverse=True)
                    last_invoice = matching_invoices[0][1]  # Get the invoice number part
                    
                    try:
                        # Parse the existing format DXXPYYY_YYMMDD
                        if '_' in last_invoice and last_invoice.startswith('D') and 'P' in last_invoice:
                            # Extract DXX (positions 1-3)
                            d_part = last_invoice[1:3]
                            # Extract PYYY (positions 4-7) 
                            p_part = last_invoice[4:7]
                            
                            d_value = int(d_part)
                            p_value = int(p_part)
                            
                            # Increment logic
                            # When P hits 100 -> D increments by 1 and P resets to 001
                            if p_value >= 100:
                                d_value += 1
                                # Optional wraparound: When D reaches 99 -> it rolls back to 00
                                if d_value > 99:
                                    d_value = 0
                                p_value = 1  # Reset to 001
                            else:
                                p_value += 1
                            
                            # Format with leading zeros
                            d_str = str(d_value).zfill(2)
                            p_str = str(p_value).zfill(3)
                            
                            return f'D{d_str}P{p_str}_{date_str}'
                    except (ValueError, IndexError, Exception):
                        # If parsing fails, start with D00P001
                        pass
                
                # Default to D00P001 if no previous invoice or parsing failed
                return f'D00P001_{date_str}'
            
            # Generate proper invoice number
            proper_invoice_number = generate_invoice_number_for_upload(conn, cur, date_obj)
            
            cur.execute("SELECT invoice_id FROM sales_invoices WHERE invoice_number = %s", (proper_invoice_number,))
            invoice_result = cur.fetchone()
            
            if not invoice_result:
                # Create new invoice with proper invoice number
                cur.execute("""
                    INSERT INTO sales_invoices 
                    (invoice_number, invoice_date, customer_name, user_id, mode_of_payment, 
                     total_amount, total_gst, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING invoice_id
                """, (proper_invoice_number, date_obj, customer_name, user_id, 'Cash', 
                      line_amount, gst_amount, 'Completed'))
                invoice_id_result = cur.fetchone()
                if invoice_id_result is None:
                    raise ValueError(f"Failed to create invoice in row {row_num}")
                invoice_id = invoice_id_result['invoice_id']
            else:
                invoice_id = invoice_result['invoice_id']
                # Update invoice totals
                cur.execute("""
                    UPDATE sales_invoices 
                    SET total_amount = total_amount + %s, total_gst = total_gst + %s
                    WHERE invoice_id = %s
                """, (line_amount, gst_amount, invoice_id))
            
            # Insert invoice item
            cur.execute("""
                INSERT INTO sales_invoice_items 
                (invoice_id, product_id, quantity, rate_at_sale, gst_rate_at_sale, 
                 exclusive_gst_amount, sgst, cgst, total_line_amount)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (invoice_id, product_id, quantity, rate, gst_rate, 
                  taxable_value, sgst, cgst, line_amount))
            
            # Update inventory
            cur.execute("""
                UPDATE inventory 
                SET stock_quantity = stock_quantity - %s 
                WHERE product_id = %s
            """, (int(quantity), product_id))
            
            record_count += 1
        
        conn.commit()
        return record_count
        
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        if cur:
            cur.close()
        if conn:
            release_db_connection(conn)
        # Close workbook if it was opened
        if workbook:
            workbook.close()

def process_purchase_excel(file_path):
    """Process purchase Excel file and insert data into database"""
    conn = None
    cur = None
    workbook = None
    try:
        # Load workbook
        workbook = load_workbook(file_path, read_only=True)
        worksheet = workbook.active
        
        if worksheet is None:
            raise ValueError("Could not load worksheet from Excel file")
        
        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            headers.append(safe_str_convert(cell, ""))
        
        # Required columns (based on the actual template)
        required_columns = ['Purchase Order Number', 'Purchase Date', 'Supplier Name', 'Product Name', 'Quantity', 'Purchase Rate']
        
        # Check if all required columns are present
        for col in required_columns:
            if col not in headers:
                raise ValueError(f"Missing required column: {col}")
        
        # Get column indices
        po_col = headers.index('Purchase Order Number')
        date_col = headers.index('Purchase Date')
        supplier_col = headers.index('Supplier Name')
        supplier_gst_col = headers.index('Supplier GST Number') if 'Supplier GST Number' in headers else None
        product_col = headers.index('Product Name')
        pack_size_col = headers.index('Product Pack Size') if 'Product Pack Size' in headers else None
        quantity_col = headers.index('Quantity')
        rate_col = headers.index('Purchase Rate')
        gst_rate_col = headers.index('GST %') if 'GST %' in headers else None
        
        # Connect to database
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        record_count = 0
        
        # Process each row
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None for cell in row):  # Skip empty rows
                continue
            
            try:
                # Extract data
                po_number = safe_str_convert(row[po_col])
                date_str = safe_str_convert(row[date_col])
                supplier_name = safe_str_convert(row[supplier_col])
                supplier_gst_number = safe_str_convert(row[supplier_gst_col]) if supplier_gst_col is not None and supplier_gst_col < len(row) else None
                product_name = safe_str_convert(row[product_col])
                pack_size = safe_str_convert(row[pack_size_col]) if pack_size_col is not None and pack_size_col < len(row) else None
                quantity = safe_float_convert(row[quantity_col])
                purchase_rate = safe_float_convert(row[rate_col])
                gst_rate = safe_float_convert(row[gst_rate_col]) if gst_rate_col is not None and gst_rate_col < len(row) else 0.0
                
                # Validate required fields
                if not all([supplier_name.strip(), po_number.strip(), date_str.strip(), product_name.strip()]):
                    raise ValueError(f"Missing required data in row {row_num}")
                
                # Convert date
                if isinstance(date_str, str):
                    try:
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                    except ValueError:
                        raise ValueError(f"Invalid date format in row {row_num}. Expected YYYY-MM-DD")
                else:
                    date_obj = date_str
                
                # Get or create admin user
                cur.execute("SELECT user_id FROM users WHERE role = 'Admin' LIMIT 1")
                user_result = cur.fetchone()
                if not user_result:
                    raise ValueError("No admin user found")
                user_id = user_result['user_id']
                
                # Get or create supplier
                supplier_id = None
                if supplier_gst_number:
                    # Try to find supplier by GST number first
                    cur.execute("SELECT supplier_id FROM suppliers WHERE supplier_gst_number = %s", (supplier_gst_number,))
                    supplier_result = cur.fetchone()
                    
                    if not supplier_result:
                        # Create supplier with GST number
                        cur.execute("""
                            INSERT INTO suppliers (supplier_name, supplier_gst_number)
                            VALUES (%s, %s)
                            RETURNING supplier_id
                        """, (supplier_name, supplier_gst_number))
                        supplier_id_result = cur.fetchone()
                        if supplier_id_result is None:
                            raise ValueError(f"Failed to create supplier in row {row_num}")
                        supplier_id = supplier_id_result['supplier_id']
                    else:
                        supplier_id = supplier_result['supplier_id']
                        # Update supplier name if different
                        cur.execute("""
                            UPDATE suppliers 
                            SET supplier_name = %s 
                            WHERE supplier_id = %s
                        """, (supplier_name, supplier_id))
                else:
                    # Try to find supplier by name
                    cur.execute("SELECT supplier_id FROM suppliers WHERE supplier_name = %s", (supplier_name,))
                    supplier_result = cur.fetchone()
                    
                    if not supplier_result:
                        # Create supplier without GST number
                        cur.execute("""
                            INSERT INTO suppliers (supplier_name)
                            VALUES (%s)
                            RETURNING supplier_id
                        """, (supplier_name,))
                        supplier_id_result = cur.fetchone()
                        if supplier_id_result is None:
                            raise ValueError(f"Failed to create supplier in row {row_num}")
                        supplier_id = supplier_id_result['supplier_id']
                    else:
                        supplier_id = supplier_result['supplier_id']
                
                # Get or create product
                cur.execute("SELECT product_id FROM products WHERE name = %s", (product_name,))
                product_result = cur.fetchone()
                
                if not product_result:
                    # Create product if not exists (with default values)
                    cur.execute("""
                        INSERT INTO products (name, sku, gst_rate, purchase_rate, selling_rate)
                        VALUES (%s, %s, %s, %s, %s)
                        RETURNING product_id
                    """, (
                        product_name,
                        f"SKU-{product_name.replace(' ', '-')[:10]}-{int(datetime.now().timestamp()) % 10000}",
                        gst_rate,  # Use GST rate from template
                        purchase_rate,
                        purchase_rate * 1.2  # Default selling rate 20% higher
                    ))
                    product_id_result = cur.fetchone()
                    if product_id_result is None:
                        raise ValueError(f"Failed to create product in row {row_num}")
                    product_id = product_id_result['product_id']
                    
                    # Initialize inventory for the new product
                    cur.execute("""
                        INSERT INTO inventory (product_id, stock_quantity)
                        VALUES (%s, %s)
                    """, (product_id, 0))
                else:
                    product_id = product_result['product_id']
                    # Update product purchase rate and GST rate if provided
                    if purchase_rate > 0 or gst_rate > 0:
                        update_fields = []
                        update_values = []
                        
                        if purchase_rate > 0:
                            update_fields.append("purchase_rate = %s")
                            update_values.append(purchase_rate)
                        
                        if gst_rate > 0:
                            update_fields.append("gst_rate = %s")
                            update_values.append(gst_rate)
                        
                        update_values.append(product_id)
                        cur.execute(f"""
                            UPDATE products 
                            SET {', '.join(update_fields)} 
                            WHERE product_id = %s
                        """, update_values)
                
                # Calculate amounts with GST
                taxable_value = quantity * purchase_rate
                gst_amount = taxable_value * (gst_rate / 100)
                total_amount = taxable_value + gst_amount
                
                # Check if purchase order exists
                cur.execute("SELECT purchase_order_id FROM purchase_orders WHERE purchase_order_number = %s", (po_number,))
                po_result = cur.fetchone()
                
                purchase_order_id = None
                if not po_result:
                    # Create new purchase order
                    cur.execute("""
                        INSERT INTO purchase_orders 
                        (purchase_order_number, purchase_date, supplier_id, supplier_name, supplier_gst_number, user_id, total_amount, status)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING purchase_order_id
                    """, (po_number, date_obj, supplier_id, supplier_name, supplier_gst_number, user_id, 0, 'Completed'))
                    po_id_result = cur.fetchone()
                    if po_id_result is None:
                        raise ValueError(f"Failed to create purchase order in row {row_num}")
                    purchase_order_id = po_id_result['purchase_order_id']
                else:
                    purchase_order_id = po_result['purchase_order_id']
                
                # Insert purchase order item with GST calculations
                cur.execute("""
                    INSERT INTO purchase_order_items 
                    (purchase_order_id, product_id, quantity, purchase_rate, gst_rate, taxable_value, sgst, cgst, total_amount)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    purchase_order_id, product_id, quantity, purchase_rate, gst_rate,
                    taxable_value, gst_amount/2, gst_amount/2, total_amount  # Split GST into SGST and CGST
                ))
                
                # Update purchase order total
                cur.execute("""
                    UPDATE purchase_orders 
                    SET total_amount = total_amount + %s
                    WHERE purchase_order_id = %s
                """, (total_amount, purchase_order_id))
                
                # Update inventory
                cur.execute("""
                    UPDATE inventory 
                    SET stock_quantity = stock_quantity + %s 
                    WHERE product_id = %s
                """, (int(quantity), product_id))
                
                record_count += 1
            
            except Exception as row_error:
                raise ValueError(f"Error processing row {row_num}: {str(row_error)}")
        
        conn.commit()
        return record_count
        
    except Exception as e:
        if conn:
            conn.rollback()
        raise e
    finally:
        if cur:
            cur.close()
        if conn:
            release_db_connection(conn)
        # Close workbook if it was opened
        if workbook:
            workbook.close()

def process_product_excel(file_path):
    """Process product Excel file and insert data into database"""
    conn = None
    cur = None
    workbook = None
    try:
        # Load workbook
        workbook = load_workbook(file_path, read_only=True)
        worksheet = workbook.active
        
        if worksheet is None:
            raise ValueError("Could not load worksheet from Excel file")
        
        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            headers.append(safe_str_convert(cell, ""))
        
        # Required columns
        required_columns = ['Product Name', 'SKU', 'GST Rate', 'Selling Rate']
        
        # Check if all required columns are present
        for col in required_columns:
            if col not in headers:
                raise ValueError(f"Missing required column: {col}")
        
        # Get column indices
        name_col = headers.index('Product Name')
        sku_col = headers.index('SKU')
        pack_size_col = headers.index('Pack Size') if 'Pack Size' in headers else None
        gst_rate_col = headers.index('GST Rate')
        purchase_rate_col = headers.index('Purchase Rate') if 'Purchase Rate' in headers else None
        selling_rate_col = headers.index('Selling Rate')
        initial_stock_col = headers.index('Initial Stock') if 'Initial Stock' in headers else None
        
        # Connect to database
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        record_count = 0
        
        # Process each row
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None for cell in row):  # Skip empty rows
                continue
            
            try:
                # Extract data
                product_name = safe_str_convert(row[name_col])
                sku = safe_str_convert(row[sku_col])
                pack_size = safe_str_convert(row[pack_size_col]) if pack_size_col is not None and len(row) > pack_size_col else None
                gst_rate = safe_float_convert(row[gst_rate_col])
                purchase_rate = safe_float_convert(row[purchase_rate_col]) if purchase_rate_col is not None and len(row) > purchase_rate_col else None
                selling_rate = safe_float_convert(row[selling_rate_col])
                initial_stock = safe_int_convert(row[initial_stock_col]) if initial_stock_col is not None and len(row) > initial_stock_col else 0
                
                # Validate required fields
                if not all([product_name.strip(), sku.strip()]):
                    raise ValueError(f"Missing required data in row {row_num}")
                
                # Check if product with same SKU already exists
                cur.execute("SELECT product_id FROM products WHERE sku = %s", (sku,))
                existing_product = cur.fetchone()
                
                if existing_product:
                    # Update existing product
                    cur.execute("""
                        UPDATE products 
                        SET name = %s, pack_size = %s, gst_rate = %s, purchase_rate = %s, selling_rate = %s
                        WHERE product_id = %s
                    """, (
                        product_name,
                        pack_size,
                        gst_rate,
                        purchase_rate,
                        selling_rate,
                        existing_product['product_id']
                    ))
                    
                    # Update inventory if initial stock is provided
                    if initial_stock > 0:
                        cur.execute("""
                            UPDATE inventory 
                            SET stock_quantity = %s 
                            WHERE product_id = %s
                        """, (initial_stock, existing_product['product_id']))
                else:
                    # Create new product
                    cur.execute("""
                        INSERT INTO products (name, pack_size, sku, gst_rate, purchase_rate, selling_rate)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        RETURNING product_id
                    """, (
                        product_name,
                        pack_size,
                        sku,
                        gst_rate,
                        purchase_rate,
                        selling_rate
                    ))
                    
                    result = cur.fetchone()
                    if result is None:
                        raise Exception("Failed to insert product")
                    product_id = result['product_id']
                    
                    # Initialize inventory for the new product
                    cur.execute("""
                        INSERT INTO inventory (product_id, stock_quantity)
                        VALUES (%s, %s)
                    """, (product_id, initial_stock))
                
                record_count += 1
            
            except Exception as row_error:
                raise ValueError(f"Error processing row {row_num}: {str(row_error)}")
        
        conn.commit()
        return record_count
        
    except Exception as e:
        if conn:
            conn.rollback()
        raise e
    finally:
        if cur:
            cur.close()
        if conn:
            release_db_connection(conn)
        # Close workbook if it was opened
        if workbook:
            workbook.close()

@data_upload_bp.route('/admin/upload-products', methods=['POST'])
@admin_required
def upload_products_data(payload):
    # Check if file is present in request
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file provided'}), 400
    
    file = request.files['file']
    
    # Check if file is selected
    if not file or not file.filename:
        return jsonify({'success': False, 'message': 'No file selected'}), 400
    
    # Check file extension
    filename = file.filename
    if not filename or not isinstance(filename, str) or not filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'}), 400
    
    temp_filename = None
    try:
        # Save file temporarily
        temp_dir = tempfile.gettempdir()
        file_extension = os.path.splitext(filename)[1] if filename else ''
        temp_filename = os.path.join(temp_dir, str(uuid.uuid4()) + file_extension)
        file.save(temp_filename)
        
        # Process the Excel file
        record_count = process_product_excel(temp_filename)
        
        # Remove temporary file
        if temp_filename and os.path.exists(temp_filename):
            try:
                os.remove(temp_filename)
            except PermissionError:
                # On Windows, the file might still be locked by the system
                # This is a known issue, but it doesn't affect the upload process
                print(f"Warning: Could not delete temporary file {temp_filename} (file may be locked)")
            except Exception as e:
                print(f"Error cleaning up temp file: {str(e)}")
        
        return jsonify({
            'success': True, 
            'message': 'Product data uploaded successfully',
            'record_count': record_count
        }), 200
        
    except Exception as e:
        # Log the error for debugging
        print(f"Error in product upload: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Remove temporary file if it exists
        if temp_filename and os.path.exists(temp_filename):
            try:
                os.remove(temp_filename)
            except PermissionError:
                # On Windows, the file might still be locked by the system
                # This is a known issue, but it doesn't affect the upload process
                print(f"Warning: Could not delete temporary file {temp_filename} (file may be locked)")
            except Exception as cleanup_error:
                print(f"Error cleaning up temp file: {str(cleanup_error)}")
        
        # Return detailed error message
        return jsonify({'success': False, 'message': f'Error processing file: {str(e)}'}), 500
