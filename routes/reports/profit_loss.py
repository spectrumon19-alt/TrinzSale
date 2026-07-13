from flask import Blueprint, request, jsonify, Response
from db import get_db_connection, release_db_connection
from auth import token_required
from psycopg2.extras import RealDictCursor
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

reports_pl_bp = Blueprint('reports_pl', __name__)

@reports_pl_bp.route('/reports/profit-loss', methods=['GET'])
@token_required
def get_profit_loss_report(payload):
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        params = []
        date_filter = ""
        if start_date:
            date_filter += " AND DATE(si.invoice_date) >= %s"
            params.append(start_date)
        if end_date:
            date_filter += " AND DATE(si.invoice_date) <= %s"
            params.append(end_date)

        # ── P&L design notes ────────────────────────────────────────────────
        # Each metric is aggregated in ITS OWN query so that a one-to-many join
        # (an invoice has many line items, and can have many returns) can never
        # multiply another metric's rows (fan-out). Mixing line-items and
        # returns in a single query would inflate revenue/COGS/discounts.
        #
        #   revenue / GST / discounts : from sales_invoices alone (no joins).
        #   COGS                      : sales_invoice_items ⋈ products (1:1 per
        #                               line — safe; NOT joined to returns).
        #   returns                   : from sales_returns alone.
        #
        # Money basis: sales_invoices.total_amount is the EX-GST subtotal (GST is
        # tracked separately), so revenue, COGS and returns are all handled on an
        # ex-GST basis for a consistent P&L.
        #
        # Discounts: the sale's total_amount is ALREADY net of the discount
        # applied at billing, so revenue already reflects discounts. Discounts
        # are therefore reported for visibility only and are NOT subtracted again
        # (doing so would double-count them).
        #
        # NOTE (COGS cost basis): line items do not capture a cost-at-sale, so
        # COGS uses products.purchase_rate (current cost). If a product's
        # purchase cost changes later, historical COGS shifts accordingly. To make
        # COGS historically stable a cost_at_sale column would need to be captured
        # at sale time (schema change — intentionally not done here).

        # Revenue & tax — sales_invoices only (no joins → no fan-out).
        revenue_query = f"""
            SELECT
                COALESCE(SUM(si.total_amount), 0) AS revenue,
                COALESCE(SUM(si.total_gst), 0)    AS tax_collected,
                COALESCE(SUM(si.discount_amount), 0) AS discounts
            FROM sales_invoices si
            WHERE si.status = 'Completed' {date_filter}
        """
        cur.execute(revenue_query, params)
        revenue_data = cur.fetchone()
        total_revenue = float(revenue_data['revenue']) if revenue_data['revenue'] else 0.0
        tax_collected = float(revenue_data['tax_collected']) if revenue_data['tax_collected'] else 0.0
        discounts = float(revenue_data['discounts']) if revenue_data['discounts'] else 0.0

        # Cost of Goods Sold — line items joined to products (1:1 per line, safe).
        cogs_query = f"""
            SELECT
                COALESCE(SUM(sii.quantity * p.purchase_rate), 0) AS cogs
            FROM sales_invoice_items sii
            JOIN products p ON sii.product_id = p.product_id
            JOIN sales_invoices si ON sii.invoice_id = si.invoice_id
            WHERE si.status = 'Completed' {date_filter}
        """
        cur.execute(cogs_query, params)
        cogs_data = cur.fetchone()
        cost_of_goods_sold = float(cogs_data['cogs']) if cogs_data['cogs'] else 0.0

        # Gross Profit
        gross_profit = total_revenue - cost_of_goods_sold
        gross_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0

        # Returns — sales_returns only (no fan-out), completed returns whose
        # original invoice is completed and within the date window. Uses the
        # ex-GST subtotal so it matches the ex-GST revenue/COGS basis.
        returns_query = f"""
            SELECT COALESCE(SUM(sr.subtotal), 0) AS returns
            FROM sales_returns sr
            JOIN sales_invoices si ON sr.original_invoice_id = si.invoice_id
            WHERE sr.status = 'Completed'
              AND si.status = 'Completed' {date_filter}
        """
        cur.execute(returns_query, params)
        returns_data = cur.fetchone()
        returns = float(returns_data['returns']) if returns_data['returns'] else 0.0

        # Net Profit = Gross Profit − Returns.
        # Discounts are already reflected in revenue and are NOT subtracted again.
        total_expenses = returns
        net_profit = gross_profit - total_expenses
        net_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0

        # Daily breakdown — computed as THREE independent per-day aggregates
        # (invoice-level, line-item-level, return-level) that are merged in
        # Python by date. This avoids joining line items and returns together,
        # which would fan out and inflate the daily figures.

        # (a) Per-day revenue / GST / discounts from sales_invoices alone.
        daily_inv_query = f"""
            SELECT
                DATE(si.invoice_date) AS report_date,
                COALESCE(SUM(si.total_amount), 0)    AS daily_revenue,
                COALESCE(SUM(si.total_gst), 0)       AS daily_gst,
                COALESCE(SUM(si.discount_amount), 0) AS daily_discounts
            FROM sales_invoices si
            WHERE si.status = 'Completed' {date_filter}
            GROUP BY DATE(si.invoice_date)
        """
        cur.execute(daily_inv_query, params)
        inv_by_date = {str(r['report_date']): r for r in cur.fetchall()}

        # (b) Per-day COGS from line items (1:1 per line — safe).
        daily_cogs_query = f"""
            SELECT
                DATE(si.invoice_date) AS report_date,
                COALESCE(SUM(sii.quantity * p.purchase_rate), 0) AS daily_cogs
            FROM sales_invoice_items sii
            JOIN products p ON sii.product_id = p.product_id
            JOIN sales_invoices si ON sii.invoice_id = si.invoice_id
            WHERE si.status = 'Completed' {date_filter}
            GROUP BY DATE(si.invoice_date)
        """
        cur.execute(daily_cogs_query, params)
        cogs_by_date = {str(r['report_date']): float(r['daily_cogs'] or 0) for r in cur.fetchall()}

        # (c) Per-day returns from sales_returns alone (ex-GST subtotal).
        daily_ret_query = f"""
            SELECT
                DATE(si.invoice_date) AS report_date,
                COALESCE(SUM(sr.subtotal), 0) AS daily_returns
            FROM sales_returns sr
            JOIN sales_invoices si ON sr.original_invoice_id = si.invoice_id
            WHERE sr.status = 'Completed'
              AND si.status = 'Completed' {date_filter}
            GROUP BY DATE(si.invoice_date)
        """
        cur.execute(daily_ret_query, params)
        ret_by_date = {str(r['report_date']): float(r['daily_returns'] or 0) for r in cur.fetchall()}

        # Merge by date (invoice dates are the master set), newest first.
        daily_data = []
        for date_key in sorted(inv_by_date.keys(), reverse=True):
            r = inv_by_date[date_key]
            daily_revenue = float(r['daily_revenue'] or 0)
            daily_discounts = float(r['daily_discounts'] or 0)
            daily_cogs = cogs_by_date.get(date_key, 0.0)
            daily_returns = ret_by_date.get(date_key, 0.0)
            daily_gross_profit = daily_revenue - daily_cogs
            # Net = gross − returns. Discounts are already in revenue (not re-subtracted).
            daily_net_profit = daily_gross_profit - daily_returns

            daily_data.append({
                'date': date_key,
                'revenue': round(daily_revenue, 2),
                'cogs': round(daily_cogs, 2),
                'gross_profit': round(daily_gross_profit, 2),
                'discounts': round(daily_discounts, 2),
                'returns': round(daily_returns, 2),
                'expenses': round(daily_returns, 2),
                'net_profit': round(daily_net_profit, 2),
                'margin': round((daily_net_profit / daily_revenue * 100) if daily_revenue > 0 else 0, 2)
            })

        return jsonify({
            'summary': {
                'revenue': round(total_revenue, 2),
                'tax_collected': round(tax_collected, 2),
                'cogs': round(cost_of_goods_sold, 2),
                'gross_profit': round(gross_profit, 2),
                'gross_margin': round(gross_margin, 2),
                'expenses': round(total_expenses, 2),
                'net_profit': round(net_profit, 2),
                'net_margin': round(net_margin, 2),
            },
            'breakdown': {
                'discounts': round(discounts, 2),
                'returns': round(returns, 2),
            },
            'daily': daily_data
        })

    except Exception as e:
        print(f"Error generating P&L report: {str(e)}")
        return jsonify({'message': 'Failed to generate report', 'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)


@reports_pl_bp.route('/reports/profit-loss/export', methods=['GET'])
@token_required
def export_profit_loss_report(payload):
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        params = []
        date_filter = ""
        if start_date:
            date_filter += " AND DATE(si.invoice_date) >= %s"
            params.append(start_date)
        if end_date:
            date_filter += " AND DATE(si.invoice_date) <= %s"
            params.append(end_date)

        # The export MUST use the same fan-out-free aggregation as the on-screen
        # report, otherwise the two disagree. See the design notes in
        # get_profit_loss_report(). Each metric is aggregated independently.

        # Revenue / GST / discounts — sales_invoices only.
        revenue_query = f"""
            SELECT
                COALESCE(SUM(si.total_amount), 0)    AS revenue,
                COALESCE(SUM(si.total_gst), 0)       AS tax_collected,
                COALESCE(SUM(si.discount_amount), 0) AS discounts
            FROM sales_invoices si
            WHERE si.status = 'Completed' {date_filter}
        """
        cur.execute(revenue_query, params)
        revenue_data = cur.fetchone()
        total_revenue = float(revenue_data['revenue']) if revenue_data['revenue'] else 0.0
        tax_collected = float(revenue_data['tax_collected']) if revenue_data['tax_collected'] else 0.0
        discounts = float(revenue_data['discounts']) if revenue_data['discounts'] else 0.0

        # COGS — line items ⋈ products (1:1 per line, safe).
        cogs_query = f"""
            SELECT
                COALESCE(SUM(sii.quantity * p.purchase_rate), 0) AS cogs
            FROM sales_invoice_items sii
            JOIN products p ON sii.product_id = p.product_id
            JOIN sales_invoices si ON sii.invoice_id = si.invoice_id
            WHERE si.status = 'Completed' {date_filter}
        """
        cur.execute(cogs_query, params)
        cogs_data = cur.fetchone()
        cost_of_goods_sold = float(cogs_data['cogs']) if cogs_data['cogs'] else 0.0

        # Gross Profit
        gross_profit = total_revenue - cost_of_goods_sold
        gross_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0

        # Returns — sales_returns only, completed, ex-GST subtotal.
        returns_query = f"""
            SELECT COALESCE(SUM(sr.subtotal), 0) AS returns
            FROM sales_returns sr
            JOIN sales_invoices si ON sr.original_invoice_id = si.invoice_id
            WHERE sr.status = 'Completed'
              AND si.status = 'Completed' {date_filter}
        """
        cur.execute(returns_query, params)
        returns = float((cur.fetchone() or {}).get('returns') or 0)

        # Net Profit = Gross Profit − Returns. Discounts already in revenue.
        total_expenses = returns
        net_profit = gross_profit - total_expenses
        net_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0

        # ── Daily breakdown: three independent per-day aggregates, merged. ──
        daily_inv_query = f"""
            SELECT
                DATE(si.invoice_date) AS report_date,
                COUNT(DISTINCT si.invoice_id)        AS invoice_count,
                COALESCE(SUM(si.total_amount), 0)    AS daily_revenue,
                COALESCE(SUM(si.total_gst), 0)       AS daily_gst,
                COALESCE(SUM(si.discount_amount), 0) AS daily_discounts
            FROM sales_invoices si
            WHERE si.status = 'Completed' {date_filter}
            GROUP BY DATE(si.invoice_date)
        """
        cur.execute(daily_inv_query, params)
        _inv = {str(r['report_date']): r for r in cur.fetchall()}

        daily_cogs_query = f"""
            SELECT
                DATE(si.invoice_date) AS report_date,
                COALESCE(SUM(sii.quantity * p.purchase_rate), 0) AS daily_cogs
            FROM sales_invoice_items sii
            JOIN products p ON sii.product_id = p.product_id
            JOIN sales_invoices si ON sii.invoice_id = si.invoice_id
            WHERE si.status = 'Completed' {date_filter}
            GROUP BY DATE(si.invoice_date)
        """
        cur.execute(daily_cogs_query, params)
        _cogs = {str(r['report_date']): float(r['daily_cogs'] or 0) for r in cur.fetchall()}

        daily_ret_query = f"""
            SELECT
                DATE(si.invoice_date) AS report_date,
                COALESCE(SUM(sr.subtotal), 0) AS daily_returns
            FROM sales_returns sr
            JOIN sales_invoices si ON sr.original_invoice_id = si.invoice_id
            WHERE sr.status = 'Completed'
              AND si.status = 'Completed' {date_filter}
            GROUP BY DATE(si.invoice_date)
        """
        cur.execute(daily_ret_query, params)
        _ret = {str(r['report_date']): float(r['daily_returns'] or 0) for r in cur.fetchall()}

        # Merge into a list of dict rows (newest first) mirroring the DB rows the
        # rest of the export expects.
        daily_breakdown = []
        for date_key in sorted(_inv.keys(), reverse=True):
            r = _inv[date_key]
            daily_breakdown.append({
                'report_date':    r['report_date'],
                'invoice_count':  r['invoice_count'],
                'daily_revenue':  r['daily_revenue'],
                'daily_gst':      r['daily_gst'],
                'daily_cogs':     _cogs.get(date_key, 0.0),
                'daily_discounts': r['daily_discounts'],
                'daily_returns':  _ret.get(date_key, 0.0),
            })

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "P&L Report"

        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        summary_font = Font(bold=True, size=11)
        summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        total_font = Font(bold=True, color="FFFFFF")
        currency_format = '₹#,##0.00'
        percent_format = '0.00"%"'
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Title
        ws['A1'] = "PROFIT & LOSS STATEMENT"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')

        # Date range
        ws['A2'] = f"Period: {start_date} to {end_date}"
        ws.merge_cells('A2:D2')

        # P&L Statement
        ws['A4'] = "P&L STATEMENT"
        ws['A4'].font = header_font
        ws['A4'].fill = header_fill
        ws.merge_cells('A4:D4')

        row = 5
        # Revenue
        ws[f'A{row}'] = "REVENUE"
        ws[f'B{row}'] = total_revenue
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = summary_font
        row += 1

        ws[f'A{row}'] = "Tax Collected (GST)"
        ws[f'B{row}'] = tax_collected
        ws[f'B{row}'].number_format = currency_format
        row += 1

        # COGS
        ws[f'A{row}'] = "COST OF GOODS SOLD"
        ws[f'B{row}'] = cost_of_goods_sold
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = summary_font
        ws[f'A{row}'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        row += 1

        # Gross Profit
        ws[f'A{row}'] = "GROSS PROFIT"
        ws[f'B{row}'] = gross_profit
        ws[f'B{row}'].number_format = currency_format
        ws[f'C{row}'] = gross_margin
        ws[f'C{row}'].number_format = percent_format
        ws[f'A{row}'].font = summary_font
        ws[f'A{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        row += 1

        # Deductions (returns reduce net profit; discounts are informational —
        # they are already reflected in Revenue and are NOT subtracted again).
        ws[f'A{row}'] = "LESS: RETURNS & REFUNDS"
        ws[f'B{row}'] = returns
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = summary_font
        row += 1

        ws[f'A{row}'] = "  Discounts Given (already in Revenue — info only)"
        ws[f'B{row}'] = discounts
        ws[f'B{row}'].number_format = currency_format
        row += 1

        # Net Profit
        row += 1
        ws[f'A{row}'] = "NET PROFIT"
        ws[f'B{row}'] = net_profit
        ws[f'B{row}'].number_format = currency_format
        ws[f'C{row}'] = net_margin
        ws[f'C{row}'].number_format = percent_format
        ws[f'A{row}'].font = total_font
        ws[f'A{row}'].fill = total_fill
        ws[f'B{row}'].font = total_font
        ws[f'B{row}'].fill = total_fill
        ws[f'C{row}'].font = total_font
        ws[f'C{row}'].fill = total_fill

        # Daily Breakdown Sheet
        if daily_breakdown:
            ws2 = wb.create_sheet("Daily Breakdown")
            ws2['A1'] = "DAILY PROFIT & LOSS BREAKDOWN"
            ws2['A1'].font = Font(bold=True, size=12)
            ws2.merge_cells('A1:L1')

            # Headers
            headers = ['Date', 'Invoices', 'Revenue', 'GST', 'COGS', 'Gross Profit', 'Gross %', 'Discounts', 'Returns', 'Expenses', 'Net Profit', 'Net %']
            ws2.append(headers)
            for col_num, header in enumerate(headers, 1):
                cell = ws2.cell(row=2, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

            # Daily data
            row_num = 3
            for row_data in daily_breakdown:
                daily_revenue = float(row_data['daily_revenue']) if row_data['daily_revenue'] else 0.0
                daily_gst = float(row_data['daily_gst']) if row_data['daily_gst'] else 0.0
                daily_cogs = float(row_data['daily_cogs']) if row_data['daily_cogs'] else 0.0
                daily_discounts = float(row_data['daily_discounts']) if row_data['daily_discounts'] else 0.0
                daily_returns = float(row_data['daily_returns']) if row_data['daily_returns'] else 0.0
                daily_gross_profit = daily_revenue - daily_cogs
                daily_gross_margin = (daily_gross_profit / daily_revenue * 100) if daily_revenue > 0 else 0
                # Only returns reduce net profit; discounts are already in revenue.
                daily_expenses = daily_returns
                daily_net_profit = daily_gross_profit - daily_expenses
                daily_net_margin = (daily_net_profit / daily_revenue * 100) if daily_revenue > 0 else 0
                invoice_count = int(row_data['invoice_count']) if row_data['invoice_count'] else 0

                ws2[f'A{row_num}'] = str(row_data['report_date'])
                ws2[f'B{row_num}'] = invoice_count
                ws2[f'C{row_num}'] = daily_revenue
                ws2[f'D{row_num}'] = daily_gst
                ws2[f'E{row_num}'] = daily_cogs
                ws2[f'F{row_num}'] = daily_gross_profit
                ws2[f'G{row_num}'] = daily_gross_margin
                ws2[f'H{row_num}'] = daily_discounts
                ws2[f'I{row_num}'] = daily_returns
                ws2[f'J{row_num}'] = daily_expenses
                ws2[f'K{row_num}'] = daily_net_profit
                ws2[f'L{row_num}'] = daily_net_margin

                # Format columns
                ws2[f'B{row_num}'].number_format = '0'
                for col in ['C', 'D', 'E', 'F', 'H', 'I', 'J', 'K']:
                    ws2[f'{col}{row_num}'].number_format = currency_format
                for col in ['G', 'L']:
                    ws2[f'{col}{row_num}'].number_format = percent_format

                # Highlight net profit
                if daily_net_profit >= 0:
                    ws2[f'K{row_num}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    ws2[f'K{row_num}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                row_num += 1

            # Auto-adjust column widths
            col_widths = {'A': 12, 'B': 10, 'C': 14, 'D': 12, 'E': 14, 'F': 14, 'G': 10, 'H': 12, 'I': 12, 'J': 12, 'K': 14, 'L': 10}
            for col, width in col_widths.items():
                ws2.column_dimensions[col].width = width

        # Auto-adjust column widths for main sheet
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        # Generate file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"profit_loss_{start_date}_to_{end_date}.xlsx"
        response = Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

        return response

    except Exception as e:
        print(f"Error exporting P&L report: {str(e)}")
        return jsonify({'message': 'Failed to export P&L report', 'error': str(e)}), 500
    finally:
        cur.close()
