from flask import Blueprint, request, jsonify, Response
from db import get_db_connection, release_db_connection
from auth import token_required
from psycopg2.extras import RealDictCursor
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

reports_sales_bp = Blueprint('reports_sales', __name__)

@reports_sales_bp.route('/reports/sales', methods=['GET'])
@token_required
def get_sales_report(payload):
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Build query based on date filters with proper date handling
        base_query = """
            SELECT 
                DATE(si.invoice_date) as sale_date,
                COUNT(si.invoice_id) as total_invoices,
                SUM(si.total_amount) as total_sales,
                SUM(si.total_gst) as total_gst_collected
            FROM sales_invoices si
            WHERE si.status = 'Completed'
        """
        
        params = []
        if start_date:
            base_query += " AND DATE(si.invoice_date) >= %s"
            params.append(start_date)
            
        if end_date:
            base_query += " AND DATE(si.invoice_date) <= %s"
            params.append(end_date)
            
        base_query += " GROUP BY DATE(si.invoice_date) ORDER BY sale_date"
        
        cur.execute(base_query, params)
        daily_sales = cur.fetchall()
        
        # Calculate overall totals
        total_sales = sum(row['total_sales'] if row['total_sales'] else 0 for row in daily_sales)
        total_gst = sum(row['total_gst_collected'] if row['total_gst_collected'] else 0 for row in daily_sales)
        total_invoices = sum(row['total_invoices'] for row in daily_sales)
        
        # Get top selling products with proper date filtering
        product_query = """
            SELECT 
                p.name,
                SUM(sii.quantity) as total_quantity_sold,
                SUM(sii.total_line_amount) as total_value_sold
            FROM sales_invoice_items sii
            JOIN products p ON sii.product_id = p.product_id
            JOIN sales_invoices si ON sii.invoice_id = si.invoice_id
            WHERE si.status = 'Completed'
        """
        
        product_params = []
        if start_date:
            product_query += " AND DATE(si.invoice_date) >= %s"
            product_params.append(start_date)
            
        if end_date:
            product_query += " AND DATE(si.invoice_date) <= %s"
            product_params.append(end_date)
            
        product_query += " GROUP BY p.product_id, p.name ORDER BY total_quantity_sold DESC LIMIT 10"
        
        cur.execute(product_query, product_params)
        top_products = cur.fetchall()
        
        # Get detailed invoice data
        include_cancelled = request.args.get('include_cancelled', 'false').lower() == 'true'
        
        status_filter = '' if include_cancelled else "WHERE si.status = 'Completed'"
        status_column = ', si.status' if include_cancelled else ''
        status_group = ', si.status' if include_cancelled else ''
        
        detailed_query = f"""
            SELECT 
                si.invoice_id,
                si.invoice_number,
                si.invoice_date,
                si.customer_name,
                si.customer_contact,
                si.total_amount,
                si.total_gst,
                si.mode_of_payment,
                COUNT(sii.item_id) as item_count{status_column}
            FROM sales_invoices si
            LEFT JOIN sales_invoice_items sii ON si.invoice_id = sii.invoice_id
            {status_filter}
        """
        
        detailed_params = params.copy()  # Use the same date filters
        if start_date:
            if include_cancelled:
                detailed_query += " WHERE DATE(si.invoice_date) >= %s"
            else:
                detailed_query += " AND DATE(si.invoice_date) >= %s"
            
        if end_date:
            detailed_query += " AND DATE(si.invoice_date) <= %s"
            
        detailed_query += f" GROUP BY si.invoice_id, si.invoice_number, si.invoice_date, si.customer_name, si.customer_contact, si.total_amount, si.total_gst, si.mode_of_payment{status_group} ORDER BY si.invoice_date DESC"
        
        cur.execute(detailed_query, detailed_params)
        detailed_invoices = cur.fetchall()

        # ── Returns for the same period (net them out of sales & GST) ───────────
        returns_query = """
            SELECT
                COALESCE(SUM(sr.total_amount), 0) AS returns_amount,
                COALESCE(SUM(sr.total_gst), 0)    AS returns_gst,
                COUNT(*)                          AS returns_count
            FROM sales_returns sr
            WHERE sr.status = 'Completed'
        """
        returns_params = []
        if start_date:
            returns_query += " AND sr.return_date >= %s"
            returns_params.append(start_date)
        if end_date:
            returns_query += " AND sr.return_date <= %s"
            returns_params.append(end_date)

        cur.execute(returns_query, returns_params)
        returns_row = cur.fetchone()
        returns_amount = float(returns_row['returns_amount']) if returns_row and returns_row['returns_amount'] else 0.0
        returns_gst    = float(returns_row['returns_gst']) if returns_row and returns_row['returns_gst'] else 0.0
        returns_count  = int(returns_row['returns_count']) if returns_row and returns_row['returns_count'] else 0

        gross_sales = float(total_sales) if total_sales else 0.0
        gross_gst   = float(total_gst) if total_gst else 0.0
        net_sales   = round(gross_sales - returns_amount, 2)
        net_gst     = round(gross_gst - returns_gst, 2)

        # Net returns out of the daily sales trend (match on return_date)
        returns_by_day = {}
        if daily_sales:
            daily_returns_query = """
                SELECT sr.return_date AS d,
                       COALESCE(SUM(sr.total_amount), 0) AS returns_amount
                FROM sales_returns sr
                WHERE sr.status = 'Completed'
            """
            if start_date:
                daily_returns_query += " AND sr.return_date >= %s"
            if end_date:
                daily_returns_query += " AND sr.return_date <= %s"
            daily_returns_query += " GROUP BY sr.return_date"
            cur.execute(daily_returns_query, returns_params)
            for r in cur.fetchall():
                returns_by_day[str(r['d'])] = float(r['returns_amount'] or 0)

        # Format the response to match what the frontend expects
        return jsonify({
            # Net figures (sales minus returns) — headline numbers
            'total_sales': net_sales,
            'total_gst': net_gst,
            # Gross + returns breakdown (for transparency)
            'gross_sales': round(gross_sales, 2),
            'gross_gst': round(gross_gst, 2),
            'returns_amount': round(returns_amount, 2),
            'returns_gst': round(returns_gst, 2),
            'returns_count': returns_count,
            'net_sales': net_sales,
            'net_gst': net_gst,
            'total_invoices': int(total_invoices) if total_invoices else 0,
            'sales_trend': [
                {
                    'date': str(row['sale_date']),
                    'amount': round(float(row['total_sales'] if row['total_sales'] else 0)
                                    - returns_by_day.get(str(row['sale_date']), 0.0), 2),
                    'invoices': int(row['total_invoices'])
                } for row in daily_sales
            ],
            'top_products': [
                {
                    'name': row['name'] or 'Unknown Product',
                    'quantity_sold': int(row['total_quantity_sold'] if row['total_quantity_sold'] else 0),
                    'total_value': float(row['total_value_sold'] if row['total_value_sold'] else 0)
                } for row in top_products
            ],
            'detailed_invoices': [
                {
                    'invoice_id':      row['invoice_id'],
                    'invoice_number':  row['invoice_number'],
                    'invoice_date':    str(row['invoice_date']) if row['invoice_date'] else None,
                    'customer_name':   row['customer_name'],
                    'customer_contact':row['customer_contact'],
                    'total_amount':    float(row['total_amount'] or 0),
                    'total_gst':       float(row['total_gst']    or 0),
                    'grand_total':     float(row['total_amount'] or 0) + float(row['total_gst'] or 0),
                    'mode_of_payment': row['mode_of_payment'],
                    'item_count':      int(row['item_count'] or 0),
                    'status':          row.get('status', 'Completed'),
                }
                for row in detailed_invoices
            ]
        }), 200
    except Exception as e:
        print(f"Error generating sales report: {str(e)}")  # Log the error for debugging
        return jsonify({'message': 'Failed to generate sales report', 'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)

@reports_sales_bp.route('/reports/sales/export', methods=['GET'])
@token_required
def export_sales_report_excel(payload):
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Sales Report"
        
        # Add header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # Report title
        if ws is not None:
            ws['A1'] = 'SALES REPORT'
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:H1')
            ws['A1'].alignment = center_alignment
        
        # Date range
        date_range = f"From {start_date or 'Beginning'} to {end_date or 'Today'}"
        if ws is not None:
            ws['A2'] = date_range
            ws['A2'].font = Font(bold=True)
            ws.merge_cells('A2:H2')
            ws['A2'].alignment = center_alignment
        
        # Add some spacing
        if ws is not None:
            ws.append([''])
        
        # Get detailed sales data
        detailed_query = """
            SELECT 
                si.invoice_number,
                si.invoice_date,
                si.customer_name,
                si.customer_contact,
                si.total_amount,
                si.total_gst,
                si.discount_amount,
                si.discount_percentage,
                u.username as cashier_name
            FROM sales_invoices si
            JOIN users u ON si.user_id = u.user_id
            WHERE si.status = 'Completed'
        """
        
        detailed_params = []
        if start_date:
            detailed_query += " AND DATE(si.invoice_date) >= %s"
            detailed_params.append(start_date)
            
        if end_date:
            detailed_query += " AND DATE(si.invoice_date) <= %s"
            detailed_params.append(end_date)
            
        detailed_query += " ORDER BY si.invoice_date DESC"
        
        cur.execute(detailed_query, detailed_params)
        detailed_sales = cur.fetchall()
        
        # Detailed sales section header
        if ws is not None:
            ws.append([''])
            ws.append(['INVOICE DETAILS'])
            ws['A4'].font = Font(bold=True, size=14)
            ws.merge_cells('A4:I4')
        
        # Add column headers for detailed sales
        headers = ['Invoice Number', 'Date', 'Customer Name', 'Customer Contact', 
                  'Total Amount (₹)', 'Total GST (₹)', 'Discount (₹)', 'Discount (%)', 'Cashier']
        if ws is not None:
            ws.append(headers)
        
        # Style the headers
        if ws is not None:
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
        
        # Add detailed sales data
        row_num = 6
        for row_data in detailed_sales:
            if ws is not None:
                ws.append([
                    row_data['invoice_number'],
                    row_data['invoice_date'].strftime('%Y-%m-%d %H:%M:%S') if row_data['invoice_date'] else '',
                    row_data['customer_name'] or '',
                    row_data['customer_contact'] or '',
                    float(row_data['total_amount']) if row_data['total_amount'] else 0.0,
                    float(row_data['total_gst']) if row_data['total_gst'] else 0.0,
                    float(row_data['discount_amount']) if row_data['discount_amount'] else 0.0,
                    float(row_data['discount_percentage']) if row_data['discount_percentage'] else 0.0,
                    row_data['cashier_name'] or ''
                ])
                row_num += 1
        
        # Add some spacing
        if ws is not None:
            ws.append([''])
            row_num += 1
        
        # Get itemized sales data
        itemized_query = """
            SELECT 
                si.invoice_number,
                si.invoice_date,
                p.name as product_name,
                p.pack_size,
                sii.quantity,
                sii.rate_at_sale,
                sii.gst_rate_at_sale,
                sii.exclusive_gst_amount,
                sii.sgst,
                sii.cgst,
                sii.total_line_amount,
                sii.discount_percentage as item_discount_percentage,
                COALESCE(sii.rebate_amount, 0) as item_rebate_amount
            FROM sales_invoice_items sii
            JOIN sales_invoices si ON sii.invoice_id = si.invoice_id
            JOIN products p ON sii.product_id = p.product_id
            WHERE si.status = 'Completed'
        """
        
        itemized_params = []
        if start_date:
            itemized_query += " AND DATE(si.invoice_date) >= %s"
            itemized_params.append(start_date)
            
        if end_date:
            itemized_query += " AND DATE(si.invoice_date) <= %s"
            itemized_params.append(end_date)
            
        itemized_query += " ORDER BY si.invoice_date DESC, sii.item_id"
        
        cur.execute(itemized_query, itemized_params)
        itemized_sales = cur.fetchall()
        
        # Itemized sales section header
        if ws is not None:
            ws.append([''])
            ws.append(['ITEMIZED SALES DETAILS'])
            ws.cell(row=row_num+2, column=1).font = Font(bold=True, size=14)
            ws.merge_cells(start_row=row_num+2, start_column=1, end_row=row_num+2, end_column=12)
            row_num += 2
        
        # Add column headers for itemized sales
        item_headers = ['Invoice Number', 'Date', 'Product Name', 'Pack Size', 'Quantity', 
                       'Rate (₹)', 'GST Rate (%)', 'Taxable Value (₹)', 'SGST (₹)', 'CGST (₹)', 'Total Amount (₹)', 'Discount (%)', 'Rebate (₹)']
        if ws is not None:
            ws.append(item_headers)
            row_num += 1
        
        # Style the item headers
        if ws is not None:
            for col_num, header in enumerate(item_headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
        
        # Add itemized sales data
        for row_data in itemized_sales:
            if ws is not None:
                ws.append([
                    row_data['invoice_number'],
                    row_data['invoice_date'].strftime('%Y-%m-%d') if row_data['invoice_date'] else '',
                    row_data['product_name'] or '',
                    row_data['pack_size'] or '',
                    int(row_data['quantity']) if row_data['quantity'] else 0,
                    float(row_data['rate_at_sale']) if row_data['rate_at_sale'] else 0.0,
                    float(row_data['gst_rate_at_sale']) if row_data['gst_rate_at_sale'] else 0.0,
                    float(row_data['exclusive_gst_amount']) if row_data['exclusive_gst_amount'] else 0.0,
                    float(row_data['sgst']) if row_data['sgst'] else 0.0,
                    float(row_data['cgst']) if row_data['cgst'] else 0.0,
                    float(row_data['total_line_amount']) if row_data['total_line_amount'] else 0.0,
                    float(row_data['item_discount_percentage']) if row_data['item_discount_percentage'] else 0.0,
                    float(row_data['item_rebate_amount']) if row_data['item_rebate_amount'] else 0.0
                ])
                row_num += 1
        
        # Auto-adjust column widths
        if ws is not None:
            for col_num in range(1, len(item_headers) + 1):
                max_length = 0
                column_letter = chr(64 + col_num)  # Convert 1-26 to A-Z
                for row_num in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook to a bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create response
        filename = f"sales_report_{start_date or 'beginning'}_to_{end_date or 'today'}.xlsx"
        response = Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
        return response
        
    except Exception as e:
        print(f"Error exporting sales report to Excel: {str(e)}")
        return jsonify({'message': 'Failed to export sales report', 'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)
