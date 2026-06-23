from flask import Blueprint, request, jsonify, Response
from db import get_db_connection, release_db_connection
from auth import token_required
from psycopg2.extras import RealDictCursor
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

reports_inventory_bp = Blueprint('reports_inventory', __name__)

@reports_inventory_bp.route('/reports/inventory', methods=['GET'])
@token_required
def get_inventory_report(payload):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Get detailed inventory data
        cur.execute("""
            SELECT 
                p.product_id,
                p.name,
                p.sku,
                p.pack_size,
                p.selling_rate,
                p.gst_rate,
                i.stock_quantity,
                (p.selling_rate * i.stock_quantity) as subtotal
            FROM products p
            JOIN inventory i ON p.product_id = i.product_id
            ORDER BY p.name
        """)
        
        inventory_data = cur.fetchall()
        
        # Calculate overall totals
        total_products = len(inventory_data)
        total_stock_value = sum(row['subtotal'] if row['subtotal'] else 0 for row in inventory_data)
        total_stock_quantity = sum(row['stock_quantity'] if row['stock_quantity'] else 0 for row in inventory_data)
        
        # Format the response
        return jsonify({
            'total_products': int(total_products) if total_products else 0,
            'total_stock_value': float(total_stock_value) if total_stock_value else 0.0,
            'total_stock_quantity': int(total_stock_quantity) if total_stock_quantity else 0,
            'inventory_items': inventory_data
        }), 200
    except Exception as e:
        print(f"Error generating inventory report: {str(e)}")
        return jsonify({'message': 'Failed to generate inventory report', 'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)

@reports_inventory_bp.route('/reports/inventory/export', methods=['GET'])
@token_required
def export_inventory_report_excel(payload):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Inventory Report"
        
        # Add header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # Report title
        if ws is not None:
            ws['A1'] = 'INVENTORY REPORT'
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:H1')
            ws['A1'].alignment = center_alignment
        
        # Add some spacing
        if ws is not None:
            ws.append([''])
        
        # Get detailed inventory data
        cur.execute("""
            SELECT 
                p.product_id,
                p.name,
                p.sku,
                p.pack_size,
                p.selling_rate,
                p.gst_rate,
                i.stock_quantity,
                (p.selling_rate * i.stock_quantity) as subtotal
            FROM products p
            JOIN inventory i ON p.product_id = i.product_id
            ORDER BY p.name
        """)
        
        inventory_data = cur.fetchall()
        
        # Inventory section header
        if ws is not None:
            ws.append([''])
            ws.append(['INVENTORY DETAILS'])
            ws['A4'].font = Font(bold=True, size=14)
            ws.merge_cells('A4:H4')
        
        # Add column headers for inventory
        headers = ['Product ID', 'Product Name', 'SKU', 'Pack Size', 'Selling Rate (₹)', 
                  'GST Rate (%)', 'Stock Quantity', 'Subtotal (₹)']
        if ws is not None:
            ws.append(headers)
        
        # Style the headers
        if ws is not None:
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
        
        # Add inventory data
        row_num = 6
        for row_data in inventory_data:
            if ws is not None:
                ws.append([
                    row_data['product_id'],
                    row_data['name'] or '',
                    row_data['sku'] or '',
                    row_data['pack_size'] or '',
                    float(row_data['selling_rate']) if row_data['selling_rate'] else 0.0,
                    float(row_data['gst_rate']) if row_data['gst_rate'] else 0.0,
                    int(row_data['stock_quantity']) if row_data['stock_quantity'] else 0,
                    float(row_data['subtotal']) if row_data['subtotal'] else 0.0
                ])
                row_num += 1
        
        # Add summary row
        total_products = len(inventory_data)
        total_stock_value = sum(row['subtotal'] if row['subtotal'] else 0 for row in inventory_data)
        total_stock_quantity = sum(row['stock_quantity'] if row['stock_quantity'] else 0 for row in inventory_data)
        
        if ws is not None:
            ws.append([''])
            ws.append(['SUMMARY'])
            ws.cell(row=row_num+2, column=1).font = Font(bold=True, size=14)
            ws.merge_cells(start_row=row_num+2, start_column=1, end_row=row_num+2, end_column=8)
            
            ws.append(['Total Products', total_products, '', '', '', '', 'Total Stock Quantity', total_stock_quantity])
            ws.append(['', '', '', '', '', '', 'Total Stock Value (₹)', total_stock_value])
            
            # Style summary
            summary_row = row_num + 4
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
            ws.cell(row=summary_row, column=2).font = Font(bold=True)
            ws.cell(row=summary_row, column=7).font = Font(bold=True)
            ws.cell(row=summary_row, column=8).font = Font(bold=True)
            ws.cell(row=summary_row+1, column=7).font = Font(bold=True)
            ws.cell(row=summary_row+1, column=8).font = Font(bold=True)
        
        # Auto-adjust column widths
        if ws is not None:
            for col_num in range(1, len(headers) + 1):
                max_length = 0
                column_letter = chr(64 + col_num)  # Convert 1-26 to A-Z
                for row_num in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook to a bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create response
        filename = f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
        return response

    except Exception as e:
        print(f"Error exporting inventory report to Excel: {str(e)}")
        return jsonify({'message': 'Failed to export inventory report', 'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)


# ── Profit & Loss Report ──────────────────────────────────────────────────────
