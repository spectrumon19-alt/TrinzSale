from flask import Blueprint, request, jsonify, Response
from db import get_db_connection, release_db_connection
from auth import token_required
from psycopg2.extras import RealDictCursor
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

reports_purchase_bp = Blueprint('reports_purchase', __name__)

@reports_purchase_bp.route('/reports/purchase', methods=['GET'])
@token_required
def get_purchase_report(payload):
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Get detailed purchase data
        detailed_query = """
            SELECT 
                po.purchase_order_number,
                po.purchase_date,
                s.supplier_name,
                s.supplier_gst_number,
                po.total_amount,
                u.username as created_by
            FROM purchase_orders po
            JOIN suppliers s ON po.supplier_id = s.supplier_id
            JOIN users u ON po.user_id = u.user_id
            WHERE po.status = 'Completed'
        """
        
        detailed_params = []
        if start_date:
            detailed_query += " AND DATE(po.purchase_date) >= %s"
            detailed_params.append(start_date)
            
        if end_date:
            detailed_query += " AND DATE(po.purchase_date) <= %s"
            detailed_params.append(end_date)
            
        detailed_query += " ORDER BY po.purchase_date DESC"
        
        cur.execute(detailed_query, detailed_params)
        detailed_purchases = cur.fetchall()
        
        # Get itemized purchase data
        itemized_query = """
            SELECT 
                po.purchase_order_number,
                po.purchase_date,
                s.supplier_name,
                p.name as product_name,
                p.pack_size,
                poi.quantity,
                poi.purchase_rate,
                poi.gst_rate,
                poi.taxable_value,
                poi.sgst,
                poi.cgst,
                poi.total_amount as item_total_amount
            FROM purchase_order_items poi
            JOIN purchase_orders po ON poi.purchase_order_id = po.purchase_order_id
            JOIN suppliers s ON po.supplier_id = s.supplier_id
            JOIN products p ON poi.product_id = p.product_id
            WHERE po.status = 'Completed'
        """
        
        itemized_params = []
        if start_date:
            itemized_query += " AND DATE(po.purchase_date) >= %s"
            itemized_params.append(start_date)
            
        if end_date:
            itemized_query += " AND DATE(po.purchase_date) <= %s"
            itemized_params.append(end_date)
            
        itemized_query += " ORDER BY po.purchase_date DESC, poi.item_id"
        
        cur.execute(itemized_query, itemized_params)
        itemized_purchases = cur.fetchall()
        
        # Calculate overall totals
        total_purchase_amount = sum(row['total_amount'] if row['total_amount'] else 0 for row in detailed_purchases)
        total_purchase_orders = len(detailed_purchases)
        
        # Format the response
        return jsonify({
            'total_purchase_amount': float(total_purchase_amount) if total_purchase_amount else 0.0,
            'total_purchase_orders': int(total_purchase_orders) if total_purchase_orders else 0,
            'purchase_orders': detailed_purchases,
            'purchase_items': itemized_purchases
        }), 200
    except Exception as e:
        print(f"Error generating purchase report: {str(e)}")
        return jsonify({'message': 'Failed to generate purchase report', 'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)

@reports_purchase_bp.route('/reports/purchase/export', methods=['GET'])
@token_required
def export_purchase_report_excel(payload):
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Purchase Report"
        
        # Add header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # Report title
        if ws is not None:
            ws['A1'] = 'PURCHASE REPORT'
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:H1')
            ws['A1'].alignment = center_alignment
        
        # Date range
        date_range = f"From {start_date or 'Beginning'} to {end_date or 'Today'}"
        if ws is not None:
            ws['A2'] = date_range
            ws['A2'].font = Font(bold=True)
            ws.merge_cells('A2:H2')
            ws['A2'].alignment = center_alignment
        
        # Add some spacing
        if ws is not None:
            ws.append([''])
        
        # Get detailed purchase data
        detailed_query = """
            SELECT 
                po.purchase_order_number,
                po.purchase_date,
                s.supplier_name,
                s.supplier_gst_number,
                po.total_amount,
                u.username as created_by
            FROM purchase_orders po
            JOIN suppliers s ON po.supplier_id = s.supplier_id
            JOIN users u ON po.user_id = u.user_id
            WHERE po.status = 'Completed'
        """
        
        detailed_params = []
        if start_date:
            detailed_query += " AND DATE(po.purchase_date) >= %s"
            detailed_params.append(start_date)
            
        if end_date:
            detailed_query += " AND DATE(po.purchase_date) <= %s"
            detailed_params.append(end_date)
            
        detailed_query += " ORDER BY po.purchase_date DESC"
        
        cur.execute(detailed_query, detailed_params)
        detailed_purchases = cur.fetchall()
        
        # Detailed purchase section header
        if ws is not None:
            ws.append([''])
            ws.append(['PURCHASE ORDER DETAILS'])
            ws['A4'].font = Font(bold=True, size=14)
            ws.merge_cells('A4:G4')
        
        # Add column headers for detailed purchases
        headers = ['Purchase Order Number', 'Date', 'Supplier Name', 'Supplier GST Number', 
                  'Total Amount (₹)', 'Created By']
        if ws is not None:
            ws.append(headers)
        
        # Style the headers
        if ws is not None:
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
        
        # Add detailed purchase data
        row_num = 6
        for row_data in detailed_purchases:
            if ws is not None:
                ws.append([
                    row_data['purchase_order_number'],
                    row_data['purchase_date'].strftime('%Y-%m-%d %H:%M:%S') if row_data['purchase_date'] else '',
                    row_data['supplier_name'] or '',
                    row_data['supplier_gst_number'] or '',
                    float(row_data['total_amount']) if row_data['total_amount'] else 0.0,
                    row_data['created_by'] or ''
                ])
                row_num += 1
        
        # Add some spacing
        if ws is not None:
            ws.append([''])
            row_num += 1
        
        # Get itemized purchase data
        itemized_query = """
            SELECT 
                po.purchase_order_number,
                po.purchase_date,
                s.supplier_name,
                p.name as product_name,
                p.pack_size,
                poi.quantity,
                poi.purchase_rate,
                poi.gst_rate,
                poi.taxable_value,
                poi.sgst,
                poi.cgst,
                poi.total_amount as item_total_amount
            FROM purchase_order_items poi
            JOIN purchase_orders po ON poi.purchase_order_id = po.purchase_order_id
            JOIN suppliers s ON po.supplier_id = s.supplier_id
            JOIN products p ON poi.product_id = p.product_id
            WHERE po.status = 'Completed'
        """
        
        itemized_params = []
        if start_date:
            itemized_query += " AND DATE(po.purchase_date) >= %s"
            itemized_params.append(start_date)
            
        if end_date:
            itemized_query += " AND DATE(po.purchase_date) <= %s"
            itemized_params.append(end_date)
            
        itemized_query += " ORDER BY po.purchase_date DESC, poi.item_id"
        
        cur.execute(itemized_query, itemized_params)
        itemized_purchases = cur.fetchall()
        
        # Itemized purchase section header
        if ws is not None:
            ws.append([''])
            ws.append(['ITEMIZED PURCHASE DETAILS'])
            ws.cell(row=row_num+2, column=1).font = Font(bold=True, size=14)
            ws.merge_cells(start_row=row_num+2, start_column=1, end_row=row_num+2, end_column=12)
            row_num += 2
        
        # Add column headers for itemized purchases
        item_headers = ['Purchase Order Number', 'Date', 'Supplier Name', 'Product Name', 'Pack Size', 'Quantity', 
                       'Purchase Rate (₹)', 'GST Rate (%)', 'Taxable Value (₹)', 'SGST (₹)', 'CGST (₹)', 'Total Amount (₹)']
        if ws is not None:
            ws.append(item_headers)
            row_num += 1
        
        # Style the item headers
        if ws is not None:
            for col_num, header in enumerate(item_headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
        
        # Add itemized purchase data
        for row_data in itemized_purchases:
            if ws is not None:
                ws.append([
                    row_data['purchase_order_number'],
                    row_data['purchase_date'].strftime('%Y-%m-%d') if row_data['purchase_date'] else '',
                    row_data['supplier_name'] or '',
                    row_data['product_name'] or '',
                    row_data['pack_size'] or '',
                    int(row_data['quantity']) if row_data['quantity'] else 0,
                    float(row_data['purchase_rate']) if row_data['purchase_rate'] else 0.0,
                    float(row_data['gst_rate']) if row_data['gst_rate'] else 0.0,
                    float(row_data['taxable_value']) if row_data['taxable_value'] else 0.0,
                    float(row_data['sgst']) if row_data['sgst'] else 0.0,
                    float(row_data['cgst']) if row_data['cgst'] else 0.0,
                    float(row_data['item_total_amount']) if row_data['item_total_amount'] else 0.0
                ])
                row_num += 1
        
        # Auto-adjust column widths
        if ws is not None:
            for col_num in range(1, len(item_headers) + 1):
                max_length = 0
                column_letter = chr(64 + col_num)  # Convert 1-26 to A-Z
                for row_num in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook to a bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create response
        filename = f"purchase_report_{start_date or 'beginning'}_to_{end_date or 'today'}.xlsx"
        response = Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
        return response
        
    except Exception as e:
        print(f"Error exporting purchase report to Excel: {str(e)}")
        return jsonify({'message': 'Failed to export purchase report', 'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)
