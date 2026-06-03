from flask import Blueprint, request, jsonify, Response
from db import get_db_connection, release_db_connection
from auth import token_required, admin_required
from psycopg2.extras import RealDictCursor
from openpyxl import Workbook, load_workbook
from io import BytesIO
import traceback
import re

supplier_data_bp = Blueprint('supplier_data', __name__)

def validate_gst_number(gst_number):
    """Validate GST number format"""
    if not gst_number:
        return False
    # Convert to string and strip whitespace
    gst_str = str(gst_number).strip().upper()
    # Handle special case for "NA" or empty-like values
    if gst_str in ["NA", "N/A", "", "NULL", "NONE"]:
        return False
    # GST format: 2 digits + 5 uppercase letters + 4 digits + 1 uppercase letter + 1 digit/letter + Z + 1 digit/letter
    gst_pattern = r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}[Z]{1}[0-9A-Z]{1}$'
    return re.match(gst_pattern, gst_str) is not None

def validate_mobile_number(mobile):
    """Validate mobile number format"""
    if not mobile:
        return False
    # Convert to string and remove any whitespace
    mobile_str = str(mobile).strip()
    # Simple mobile validation: 10 digits
    mobile_pattern = r'^[0-9]{10}$'
    return re.match(mobile_pattern, mobile_str) is not None

def safe_str_convert(cell, default=""):
    """Safely convert cell value to string"""
    if cell is None or cell == "" or (isinstance(cell, str) and cell.strip().upper() in ["NA", "N/A", "NULL", "NONE"]):
        return default
    return str(cell).strip()

def safe_float_convert(cell, default=0.0):
    """Safely convert cell value to float"""
    if cell is None:
        return default
    try:
        return float(cell)
    except (ValueError, TypeError):
        return default

@supplier_data_bp.route('/suppliers-export', methods=['GET'])
@token_required
def export_suppliers_excel(payload):
    """Export all suppliers to Excel file"""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Get all suppliers
        cur.execute("""
            SELECT 
                supplier_name, supplier_gst_number, contact_person, email, 
                mobile, address, bank_name, bank_account_number, ifsc_code
            FROM suppliers 
            ORDER BY supplier_name
        """)
        
        suppliers = cur.fetchall()
        
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Suppliers"
        
        # Add headers
        headers = ["Supplier Name", "GST Number", "Contact Person", "Email", "Mobile", "Address", "Bank Name", "Account Number", "IFSC Code"]
        if ws is not None:
            ws.append(headers)
        
        # Add data
        for supplier in suppliers:
            if ws is not None:
                ws.append([
                    supplier['supplier_name'] or '',
                    supplier['supplier_gst_number'] or '',
                    supplier['contact_person'] or '',
                    supplier['email'] or '',
                    supplier['mobile'] or '',
                    supplier['address'] or '',
                    supplier['bank_name'] or '',
                    supplier['bank_account_number'] or '',
                    supplier['ifsc_code'] or ''
                ])
        
        # Save the workbook to a bytes buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create response
        filename = "suppliers_export.xlsx"
        response = Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
        return response
        
    except Exception as e:
        print(f"Error exporting suppliers to Excel: {str(e)}")
        return jsonify({'message': 'Failed to export suppliers', 'error': str(e)}), 500
    finally:
        cur.close()
        release_db_connection(conn)

@supplier_data_bp.route('/suppliers-import', methods=['POST'])
@admin_required
def import_suppliers_excel(payload):
    """Import suppliers from Excel file"""
    if 'file' not in request.files:
        return jsonify({'message': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': 'No file selected'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'message': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'}), 400
    
    conn = None
    cur = None
    workbook = None
    
    try:
        # Save file temporarily
        temp_file_path = f"/tmp/suppliers_import_{payload.get('user_id')}.xlsx"
        file.save(temp_file_path)
        
        # Load workbook
        workbook = load_workbook(temp_file_path, read_only=True)
        worksheet = workbook.active
        
        if worksheet is None:
            raise ValueError("Could not load worksheet from Excel file")
        
        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            headers.append(safe_str_convert(cell, ""))
        
        # Required columns
        required_columns = ['Supplier Name', 'GST Number', 'Mobile']
        
        # Check if all required columns are present
        for col in required_columns:
            if col not in headers:
                raise ValueError(f"Missing required column: {col}")
        
        # Get column indices
        name_col = headers.index('Supplier Name')
        gst_col = headers.index('GST Number')
        contact_col = headers.index('Contact Person') if 'Contact Person' in headers else None
        email_col = headers.index('Email') if 'Email' in headers else None
        mobile_col = headers.index('Mobile')
        address_col = headers.index('Address') if 'Address' in headers else None
        bank_name_col = headers.index('Bank Name') if 'Bank Name' in headers else None
        account_col = headers.index('Account Number') if 'Account Number' in headers else None
        ifsc_col = headers.index('IFSC Code') if 'IFSC Code' in headers else None
        
        # Connect to database
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        record_count = 0
        error_count = 0
        errors = []
        
        # Keep track of GST numbers in the current file to detect duplicates
        gst_numbers_in_file = set()
        
        # Process each row
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None for cell in row):  # Skip empty rows
                continue
            
            try:
                # Extract data
                supplier_name = safe_str_convert(row[name_col])
                gst_number = safe_str_convert(row[gst_col])
                contact_person = safe_str_convert(row[contact_col]) if contact_col is not None and len(row) > contact_col else None
                email = safe_str_convert(row[email_col]) if email_col is not None and len(row) > email_col else None
                mobile = safe_str_convert(row[mobile_col])
                address = safe_str_convert(row[address_col]) if address_col is not None and len(row) > address_col else None
                bank_name = safe_str_convert(row[bank_name_col]) if bank_name_col is not None and len(row) > bank_name_col else None
                account_number = safe_str_convert(row[account_col]) if account_col is not None and len(row) > account_col else None
                ifsc_code = safe_str_convert(row[ifsc_col]) if ifsc_col is not None and len(row) > ifsc_col else None
                
                # Skip rows with missing required data
                if not all([supplier_name, gst_number, mobile]):
                    errors.append(f"Row {row_num}: Missing required data - Name: '{supplier_name}', GST: '{gst_number}', Mobile: '{mobile}'")
                    error_count += 1
                    continue
                
                # Additional check for "NA" or invalid GST numbers
                if not gst_number or gst_number.upper() in ["NA", "N/A", "NULL", "NONE", ""]:
                    errors.append(f"Row {row_num}: Invalid GST number '{gst_number}' - GST number is mandatory for all suppliers")
                    error_count += 1
                    continue
                
                # Check for duplicate GST numbers within the file
                if gst_number in gst_numbers_in_file:
                    errors.append(f"Row {row_num}: Duplicate GST number '{gst_number}' found in this file")
                    error_count += 1
                    continue
                
                # Add GST number to the set
                gst_numbers_in_file.add(gst_number)
                
                # Validate data
                gst_valid = validate_gst_number(gst_number)
                if not gst_valid:
                    errors.append(f"Row {row_num}: Invalid GST number format '{gst_number}' - GST number is mandatory and must follow the correct format (e.g., 22AAAAA0000A1Z5)")
                    error_count += 1
                    continue
                
                mobile_valid = validate_mobile_number(mobile)
                if not mobile_valid:
                    errors.append(f"Row {row_num}: Invalid mobile number format '{mobile}' - must be exactly 10 digits")
                    error_count += 1
                    continue
                
                # Check if supplier with same GST number already exists
                cur.execute("""
                    SELECT supplier_id FROM suppliers 
                    WHERE supplier_gst_number = %s
                """, (gst_number,))
                
                existing_supplier = cur.fetchone()
                if existing_supplier:
                    # Update existing supplier
                    cur.execute("""
                        UPDATE suppliers 
                        SET supplier_name = %s, contact_person = %s, email = %s, mobile = %s, 
                            address = %s, bank_name = %s, bank_account_number = %s, ifsc_code = %s
                        WHERE supplier_gst_number = %s
                    """, (
                        supplier_name, contact_person, email, mobile, address,
                        bank_name, account_number, ifsc_code, gst_number
                    ))
                else:
                    # Insert new supplier
                    cur.execute("""
                        INSERT INTO suppliers (
                            supplier_name, supplier_gst_number, contact_person, email, 
                            mobile, address, bank_name, bank_account_number, ifsc_code
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        supplier_name, gst_number, contact_person, email, mobile,
                        address, bank_name, account_number, ifsc_code
                    ))
                
                record_count += 1
                
            except Exception as row_error:
                errors.append(f"Row {row_num}: {str(row_error)}")
                error_count += 1
                continue
        
        conn.commit()
        
        # Clean up temp file
        import os
        temp_file_path = f"/tmp/suppliers_import_{payload.get('user_id')}.xlsx"
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        
        return jsonify({
            'message': f'Successfully processed {record_count} suppliers. {error_count} errors occurred.',
            'processed_count': record_count,
            'error_count': error_count,
            'errors': errors
        }), 200
        
    except Exception as e:
        if conn:
            conn.rollback()
        
        # Clean up temp file
        import os
        temp_file_path = f"/tmp/suppliers_import_{payload.get('user_id')}.xlsx"
        try:
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
        except:
            pass
            
        print(f"Error importing suppliers from Excel: {str(e)}")
        traceback.print_exc()
        return jsonify({'message': 'Failed to import suppliers', 'error': str(e)}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            release_db_connection(conn)
        # Close workbook if it was opened
        if workbook:
            workbook.close()