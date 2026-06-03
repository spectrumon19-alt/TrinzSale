"""
Generates tests/reports/bug_report.xlsx â€” a complete bug & fix registry
for the TrintzPOS POS system.

Usage:
    python create_bug_report.py

Requires: pip install openpyxl
"""

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

TODAY = date.today().strftime("%Y-%m-%d")
OUT_DIR = os.path.join(os.path.dirname(__file__), "tests", "reports")
OUT_FILE = os.path.join(OUT_DIR, "bug_report.xlsx")

# â”€â”€ Colour palette â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
RED    = "FFC7CE"   # critical
ORANGE = "FFE0B2"   # high
YELLOW = "FFF9C4"   # medium
GREEN  = "C8E6C9"   # low / fixed
BLUE   = "BBDEFB"   # info / header
GRAY   = "ECEFF1"   # zebra row
WHITE  = "FFFFFF"
DARK   = "263238"   # header text
FIXED  = "1B5E20"   # "Fixed" text
OPEN_  = "B71C1C"   # "Open" text

SEVERITY_COLOR = {
    "Critical": RED,
    "High":     ORANGE,
    "Medium":   YELLOW,
    "Low":      GREEN,
}

STATUS_COLOR = {
    "Fixed":    "C8E6C9",
    "Open":     "FFCDD2",
    "Verified": "B3E5FC",
    "Known":    "FFF9C4",
}

# â”€â”€ Helper utilities â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _thin():
    s = Side(style="thin", color="B0BEC5")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _cell_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def style_header_row(ws, row: int, col_count: int, fill_hex: str = BLUE):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=DARK, size=10)
        cell.fill = _hdr_fill(fill_hex)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin()

def write_row(ws, row: int, values: list, fill_hex: str = WHITE, bold=False):
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = Font(bold=bold, size=9)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = _thin()
        cell.fill = _cell_fill(fill_hex)

def set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# DATA
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

BUGS = [
    # â”€â”€ SCHEMA BUGS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    {
        "id": "BUG-001",
        "category": "Schema",
        "severity": "Critical",
        "status": "Fixed",
        "file": "schema.sql",
        "title": "users table missing full_name, email, mobile columns",
        "description": (
            "schema.sql CREATE TABLE users does not define full_name, email, "
            "or mobile columns. Every call to GET/POST/PUT /api/admin/users "
            "references these columns and throws a PostgreSQL error on a "
            "fresh database install, returning HTTP 500."
        ),
        "root_cause": "Schema was never updated after admin.py added the columns.",
        "fix": (
            "Added full_name VARCHAR, email VARCHAR, mobile VARCHAR to the "
            "CREATE TABLE users statement in schema.sql, plus "
            "ALTER TABLE users ADD COLUMN IF NOT EXISTS guards for "
            "existing databases."
        ),
        "test_ref": "tests/api/test_admin.py::TestUserCRUD",
        "fixed_date": TODAY,
        "found_by": "Static code analysis + conftest EXTRA_DDL workaround",
    },
    {
        "id": "BUG-002",
        "category": "Schema",
        "severity": "Critical",
        "status": "Fixed",
        "file": "schema.sql",
        "title": "login_activity table missing from schema",
        "description": (
            "routes/auth.py inserts a row into login_activity on every "
            "login attempt (success and failure). The table was never "
            "defined in schema.sql. A fresh DB install causes every login "
            "to 500 after authentication succeeds because the INSERT fails."
        ),
        "root_cause": "Table was added to routes but never added to schema.sql.",
        "fix": (
            "Added CREATE TABLE IF NOT EXISTS login_activity with all "
            "required columns (id, user_id FK, username, login_timestamp, "
            "ip_address, user_agent, browser, os, device_type, "
            "location_city, location_country, login_status CHECK, "
            "failure_reason) and supporting indexes."
        ),
        "test_ref": "tests/api/test_auth.py::TestLoginActivity",
        "fixed_date": TODAY,
        "found_by": "Integration test failure analysis",
    },
    {
        "id": "BUG-003",
        "category": "Schema",
        "severity": "Critical",
        "status": "Fixed",
        "file": "schema.sql",
        "title": "user_permissions table missing from schema",
        "description": (
            "admin.py has full CRUD for user_permissions (GET, PUT, check), "
            "including an ON CONFLICT upsert that requires a unique "
            "constraint on (user_id, page_id). The table was never in "
            "schema.sql. All permissions endpoints return 500 on fresh install."
        ),
        "root_cause": "Permissions feature was added to routes without updating schema.sql.",
        "fix": (
            "Added CREATE TABLE IF NOT EXISTS user_permissions with "
            "user_id FK (ON DELETE CASCADE), page_id VARCHAR, has_access BOOL, "
            "updated_at TIMESTAMP, and UNIQUE(user_id, page_id) constraint."
        ),
        "test_ref": "tests/api/test_admin.py::TestUserPermissions",
        "fixed_date": TODAY,
        "found_by": "Integration test failure analysis",
    },
    {
        "id": "BUG-004",
        "category": "Schema / Route Mismatch",
        "severity": "High",
        "status": "Fixed",
        "file": "schema.sql, routes/admin.py",
        "title": "Role CHECK constraint conflicts with allowed roles in admin routes",
        "description": (
            "schema.sql defined: CHECK (role IN ('Admin', 'Cashier')). "
            "routes/admin.py allows 'Admin', 'Cashier', 'Super Admin', 'Manager'. "
            "auth.py admin_required also allows 'Super Admin' and 'Manager'. "
            "Creating a user with role 'Super Admin' via API succeeds the "
            "route validation but is then rejected by PostgreSQL with a "
            "CHECK constraint violation â†’ HTTP 500."
        ),
        "root_cause": "Schema was never updated when new roles were introduced in routes.",
        "fix": (
            "Expanded schema.sql CHECK to: "
            "CHECK (role IN ('Admin','Cashier','Super Admin','Manager')). "
            "Updated admin.py validation message to use VALID_ROLES list."
        ),
        "test_ref": "tests/api/test_admin.py::TestCreateUser::test_super_admin_role_accepted",
        "fixed_date": TODAY,
        "found_by": "Cross-file code review",
    },
    {
        "id": "BUG-005",
        "category": "Schema",
        "severity": "High",
        "status": "Fixed",
        "file": "schema.sql",
        "title": "login_activity missing user_agent column",
        "description": (
            "routes/auth.py INSERT into login_activity includes user_agent "
            "as a column but the schema did not define it. Every login "
            "attempt would fail with 'column user_agent does not exist' "
            "causing the login activity to never be recorded."
        ),
        "root_cause": "Column omitted when schema was written.",
        "fix": "Added user_agent TEXT column to login_activity table in schema.sql.",
        "test_ref": "tests/api/test_auth.py::TestLoginActivity::test_login_records_user_agent",
        "fixed_date": TODAY,
        "found_by": "Route vs schema column diff",
    },
    # â”€â”€ LOGIC BUGS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    {
        "id": "BUG-006",
        "category": "Business Logic",
        "severity": "High",
        "status": "Fixed",
        "file": "routes/sales.py",
        "title": "grand_total double-counts GST in sales response",
        "description": (
            "calculate_invoice_item() treats selling_rate as GST-inclusive "
            "(total_line_amount = qty * selling_rate, where selling_rate "
            "already includes GST). total_amount stored in DB was the "
            "GST-inclusive total. However, the API response computed: "
            "grand_total = total_amount + total_gst, double-counting GST. "
            "Example: 1 unit at â‚¹100 (18% GST) â†’ total_amount=100, "
            "total_gst=15.25, grand_total=115.25 instead of 100.00."
        ),
        "root_cause": (
            "Naming ambiguity: total_amount was intended as the exclusive "
            "(ex-GST) subtotal but the code accumulated total_line_amount "
            "(GST-inclusive) into it."
        ),
        "fix": (
            "Changed create_sale() to accumulate exclusive_gst_amount "
            "(ex-GST base) into total_invoice_amount instead of "
            "total_line_amount. DB now stores ex-GST subtotal in "
            "total_amount so grand_total = total_amount + total_gst "
            "is mathematically correct."
        ),
        "test_ref": "tests/api/test_sales.py::TestGSTCalculation::test_grand_total_not_double_gst",
        "fixed_date": TODAY,
        "found_by": "Business logic review of calculate_invoice_item()",
    },
    {
        "id": "BUG-007",
        "category": "Business Logic",
        "severity": "Medium",
        "status": "Fixed",
        "file": "routes/products.py",
        "title": "Product search returns inactive (deactivated) products",
        "description": (
            "GET /api/products?q=<term> searched with name ILIKE and "
            "sku ILIKE but omitted the status='Active' filter. "
            "Searching returned deactivated products, which appeared "
            "in the sales product-search dropdown and could be sold "
            "after deactivation."
        ),
        "root_cause": "status filter existed only in the non-search (no q) branch.",
        "fix": (
            "Added AND p.status = 'Active' to the search branch WHERE clause "
            "in get_products(), matching the behaviour of the no-query branch."
        ),
        "test_ref": "tests/api/test_products.py::TestProductSearch::test_search_excludes_inactive",
        "fixed_date": TODAY,
        "found_by": "API integration test review",
    },
    {
        "id": "BUG-008",
        "category": "Business Logic",
        "severity": "Medium",
        "status": "Fixed",
        "file": "routes/products.py",
        "title": "Admin role check uses inline string compare instead of @admin_required decorator",
        "description": (
            "create_product, update_product, delete_product, deactivate_product "
            "each did: if payload.get('role') != 'Admin': return 403. "
            "This blocked 'Super Admin' and 'Manager' roles (which pass "
            "admin_required in other routes) from managing products. "
            "The inconsistency silently broke product management for new roles."
        ),
        "root_cause": "Route was written before admin_required decorator was created.",
        "fix": (
            "Replaced inline role check with @admin_required decorator on "
            "create_product, update_product, delete_product, and "
            "deactivate_product. Added admin_required to imports."
        ),
        "test_ref": "tests/api/test_products.py::TestProductAdminAccess",
        "fixed_date": TODAY,
        "found_by": "Role-access matrix regression test",
    },
    {
        "id": "BUG-009",
        "category": "Security",
        "severity": "High",
        "status": "Fixed",
        "file": "routes/auth.py, limiter.py, app.py, requirements.txt",
        "title": "No rate limiting on /api/login â€” brute force possible",
        "description": (
            "POST /api/login had no rate limiting. An attacker could send "
            "unlimited login attempts per second. With weak passwords, "
            "enumeration and brute-force attacks were trivially possible."
        ),
        "root_cause": "No Flask-Limiter or middleware rate limiting configured.",
        "fix": (
            "Created limiter.py with Flask-Limiter using get_remote_address key "
            "and memory:// storage. Initialized via limiter.init_app(app) in "
            "create_app(). Added @limiter.limit('10 per minute; 50 per hour') "
            "to /api/login. Added Flask-Limiter==3.5.0 to requirements.txt."
        ),
        "test_ref": "tests/security/test_security.py::TestBruteForce",
        "fixed_date": TODAY,
        "found_by": "Security test review",
    },
    {
        "id": "BUG-010",
        "category": "Security",
        "severity": "High",
        "status": "Fixed",
        "file": "routes/sales.py, routes/products.py, routes/auth.py",
        "title": "Full Python traceback exposed in error responses to clients",
        "description": (
            "On sales creation failure the API returned: "
            "{'message': ..., 'error': str(e), 'details': traceback_string}. "
            "This leaked stack frame paths, variable names, and internal "
            "implementation details to any API client, aiding attackers."
        ),
        "root_cause": "Debug-level error detail (print(), traceback, str(e)) left in production paths.",
        "fix": (
            "Added logging module to all three route files. Replaced all "
            "print() debug statements with logger.exception() / logger.warning(). "
            "Removed import traceback and all 'error'/'details' keys from "
            "client-facing error JSON. All exception handlers now return only "
            "{'message': '<safe description>'} with no internal details."
        ),
        "test_ref": "tests/security/test_security.py::TestInfoDisclosure",
        "fixed_date": TODAY,
        "found_by": "API response inspection",
    },
    {
        "id": "BUG-011",
        "category": "Business Logic",
        "severity": "Medium",
        "status": "Fixed",
        "file": "routes/sales.py, schema.sql",
        "title": "generate_invoice_number has a race condition under concurrent sales",
        "description": (
            "generate_invoice_number() fetched all invoice numbers, found "
            "the max, then incremented in Python. Under concurrent requests, "
            "two transactions could both read the same 'last' invoice number "
            "and generate duplicate invoice numbers, causing a UNIQUE "
            "constraint violation on sales_invoices.invoice_number â†’ 500."
        ),
        "root_cause": "No DB-level sequence or SELECT FOR UPDATE locking used.",
        "fix": (
            "Added CREATE SEQUENCE IF NOT EXISTS invoice_seq to schema.sql. "
            "Rewrote generate_invoice_number() to use nextval('invoice_seq') "
            "atomically: seq = nextval(); d = ((seq-1)//100)+1; p = ((seq-1)%100)+1; "
            "return f'D{d:02d}P{p:03d}_{date_str}'. Atomic at DB level, "
            "no Python-side race possible."
        ),
        "test_ref": "tests/api/test_sales.py::TestConcurrency::test_concurrent_sales_no_duplicate_invoice",
        "fixed_date": TODAY,
        "found_by": "Concurrency / load test analysis",
    },
    # â”€â”€ UI BUGS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    {
        "id": "BUG-012",
        "category": "UI / HTML",
        "severity": "Low",
        "status": "Fixed",
        "file": "dashboard.html, index.html, login.html",
        "title": "Missing data-testid attributes â€” fragile UI test selectors",
        "description": (
            "All original Playwright UI tests used class-based and id-based "
            "selectors (nth-child, .btn-login, #productSearch). These break "
            "on any CSS class rename. Tests had no stable hooks into the DOM."
        ),
        "root_cause": "data-testid attributes were never added to HTML elements.",
        "fix": (
            "Added data-testid to all testable elements in login.html "
            "(username-input, password-input, login-btn, login-error), "
            "dashboard.html (sidebar, nav-* links, logout-btn, nav-grid, "
            "card-* buttons), and index.html (all sales form fields, "
            "modal elements, filter buttons, totals)."
        ),
        "test_ref": "tests/ui/test_ui_playwright.py",
        "fixed_date": TODAY,
        "found_by": "UI test stability review",
    },
]

TEST_COVERAGE = [
    # category, file, count, areas_covered
    ("Auth / Login",     "tests/api/test_auth.py",           28, "Login success/fail, empty creds, SQL injection, JWT validation, /users/me, login activity recording"),
    ("Products CRUD",    "tests/api/test_products.py",        32, "Create/read/update/delete, search (active-only bug), deactivate, duplicate SKU, type validation, stock init, admin guard"),
    ("Sales & GST",      "tests/api/test_sales.py",           30, "GST formula (E=T/(1+r), SGST=CGST), grand_total bug doc, stock decrement/restore, invoice/receipt format regex, cancel, concurrent race"),
    ("Inventory",        "tests/api/test_inventory.py",       14, "Stock add/subtract, role guards, zero/negative, product not found"),
    ("Suppliers",        "tests/api/test_suppliers.py",       26, "CRUD, GST format validation, mobile validation, balance math, transaction history"),
    ("Admin / Users",    "tests/api/test_admin.py",           32, "User CRUD, last-admin guard, password reset, role mismatch documented, permissions CRUD"),
    ("Credit",           "tests/api/test_credit.py",          20, "Customer CRUD, auto-generated codes, transaction types, balance math"),
    ("Purchase",         "tests/api/test_purchase.py",        16, "Order CRUD, stock increment on purchase, duplicate PO number"),
    ("Reports",          "tests/api/test_reports.py",         10, "Date filtering, export endpoint, empty date range"),
    ("Regression",       "tests/api/test_regression.py",      16, "Full stock cycle, GST parametrized, invoice sequencing, role access matrix"),
    ("DB Integrity",     "tests/database/test_db_integrity.py", 28, "FK constraints, UNIQUE violations, CHECK constraints, CASCADE DELETE, rollback, NULL handling"),
    ("Security",         "tests/security/test_security.py",   36, "SQL injection (6 payloads), JWT algorithm confusion/tampering/none-alg, BOLA, traceback leak, rate limit, XSS storage, concurrent login"),
    ("Performance",      "tests/performance/test_performance.py", 12, "SLA thresholds, 50 sequential reads, bulk 100-product dataset, gzip compression"),
    ("UI Playwright",    "tests/ui/test_ui_playwright.py",    60, "Login flows, dashboard role-based cards, sales form, invoice modal, sidebar nav, RBAC, security XSS/SQL, session expiry, responsive, accessibility"),
]

DEBUGGING_SUMMARY = [
    ("DB connection refused",
     "TEST_DB_HOST/PORT mismatch or PostgreSQL not running",
     "Run: pg_isready -h localhost -p 5432. Check run_tests.bat env vars."),
    ("fixture 'test_product' setup failed: 403",
     "admin_token fixture using wrong user_id or secret key mismatch",
     "Verify TEST_SECRET_KEY matches app SECRET_KEY. Check setup_test_database user_id."),
    ("column full_name does not exist",
     "Running against a DB that was created before schema.sql fix BUG-001",
     "Re-run schema.sql against the DB or run EXTRA_DDL manually."),
    ("unique constraint violation on invoice_number",
     "Race condition (BUG-011) or previous test run left data",
     "Ensure test isolation: each test creates unique products. Use function-scoped fixtures."),
    ("gzip decode error in parse_json()",
     "Response not actually gzip-encoded",
     "Check Content-Encoding header. Use parse_json() helper which checks before decompressing."),
    ("Playwright TimeoutError",
     "App not running or element has no data-testid",
     "Start Flask: python app.py. Verify data-testid exists in HTML. Use --headed for visual debug."),
    ("JWT decode error in conftest",
     "PyJWT version API changed (v1 vs v2)",
     "v2: jwt.encode() returns str. v1: returns bytes. requirements-test.txt pins PyJWT>=2.0."),
    ("psycopg2.errors.ForeignKeyViolation in teardown",
     "Test created dependent records (sales) that weren't cleaned up",
     "Use cascade deletes or explicit teardown ordering in fixtures."),
]

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# BUILD WORKBOOK
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def build_workbook() -> Workbook:
    wb = Workbook()

    # â”€â”€ Sheet 1: Summary â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "POS System â€” Bug & Fix Registry"
    title.font = Font(bold=True, size=16, color=WHITE)
    title.fill = _hdr_fill(DARK)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value = f"TrintzPOS  Â·  Generated: {TODAY}  Â·  Total Bugs: {len(BUGS)}"
    sub.font = Font(italic=True, size=10, color=DARK)
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Stats row
    fixed  = sum(1 for b in BUGS if b["status"] == "Fixed")
    open_  = sum(1 for b in BUGS if b["status"] == "Open")
    crit   = sum(1 for b in BUGS if b["severity"] == "Critical")
    high   = sum(1 for b in BUGS if b["severity"] == "High")

    ws.merge_cells("A4:B4"); ws["A4"].value = "Total Bugs"
    ws.merge_cells("C4:D4"); ws["C4"].value = "Fixed"
    ws.merge_cells("E4:F4"); ws["E4"].value = "Open"
    ws.merge_cells("A5:B5"); ws["A5"].value = len(BUGS)
    ws.merge_cells("C5:D5"); ws["C5"].value = fixed
    ws.merge_cells("E5:F5"); ws["E5"].value = open_

    for cell in [ws["A4"], ws["C4"], ws["E4"]]:
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center")
    for cell in [ws["A5"], ws["C5"], ws["E5"]]:
        cell.font = Font(bold=True, size=18)
        cell.alignment = Alignment(horizontal="center")

    ws["C5"].fill = _cell_fill(GREEN)
    ws["E5"].fill = _cell_fill(RED)

    # Severity breakdown
    ws["A7"] = "Severity"; ws["B7"] = "Count"; ws["C7"] = "% of Total"
    style_header_row(ws, 7, 3)
    svr_data = [
        ("Critical", crit, "Critical"),
        ("High", high, "High"),
        ("Medium", sum(1 for b in BUGS if b["severity"]=="Medium"), "Medium"),
        ("Low", sum(1 for b in BUGS if b["severity"]=="Low"), "Low"),
    ]
    for i, (sev, cnt, key) in enumerate(svr_data, 8):
        pct = f"{cnt/len(BUGS)*100:.0f}%"
        write_row(ws, i, [sev, cnt, pct], fill_hex=SEVERITY_COLOR.get(key, WHITE))

    # Test file summary
    ws["A14"] = "Test File"; ws["B14"] = "Tests"; ws["C14"] = "Areas Covered"
    style_header_row(ws, 14, 3)
    total_tests = 0
    for i, (cat, fname, cnt, areas) in enumerate(TEST_COVERAGE, 15):
        fill = GREEN if i % 2 == 0 else WHITE
        write_row(ws, i, [fname, cnt, areas], fill_hex=fill)
        total_tests += cnt
    # Total row
    write_row(ws, 15 + len(TEST_COVERAGE), ["TOTAL", total_tests, ""], fill_hex=BLUE, bold=True)

    set_col_widths(ws, [30, 12, 80])
    freeze(ws, "A3")

    # â”€â”€ Sheet 2: Bugs & Fixes â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws2 = wb.create_sheet("Bugs & Fixes")
    ws2.sheet_view.showGridLines = False

    headers = [
        "Bug ID", "Category", "Severity", "Status", "File(s)",
        "Title", "Description", "Root Cause", "Fix Applied",
        "Test Reference", "Fixed Date", "Found By",
    ]
    write_row(ws2, 1, headers, fill_hex=DARK, bold=True)
    for c in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.row_dimensions[1].height = 22

    for i, bug in enumerate(BUGS, 2):
        sev_fill = SEVERITY_COLOR.get(bug["severity"], WHITE)
        stat_fill = STATUS_COLOR.get(bug["status"], WHITE)
        row_fill = GRAY if i % 2 == 0 else WHITE
        vals = [
            bug["id"], bug["category"], bug["severity"], bug["status"],
            bug["file"], bug["title"], bug["description"],
            bug["root_cause"], bug["fix"], bug["test_ref"],
            bug["fixed_date"], bug["found_by"],
        ]
        write_row(ws2, i, vals, fill_hex=row_fill)
        # Override severity and status cells
        ws2.cell(i, 3).fill = _cell_fill(sev_fill)
        ws2.cell(i, 3).font = Font(bold=True, size=9)
        ws2.cell(i, 4).fill = _cell_fill(stat_fill)
        ws2.cell(i, 4).font = Font(
            bold=True, size=9,
            color=FIXED if bug["status"] == "Fixed" else OPEN_
        )
        ws2.row_dimensions[i].height = 90

    set_col_widths(ws2, [10, 18, 10, 10, 28, 38, 58, 40, 58, 38, 12, 22])
    freeze(ws2, "A2")
    ws2.auto_filter.ref = f"A1:L{len(BUGS)+1}"

    # â”€â”€ Sheet 3: Test Coverage â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws3 = wb.create_sheet("Test Coverage")
    ws3.sheet_view.showGridLines = False

    cov_headers = ["Module", "Test File", "Test Count", "Key Areas"]
    write_row(ws3, 1, cov_headers, fill_hex=DARK, bold=True)
    for c in range(1, 5):
        cell = ws3.cell(row=1, column=c)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    grand_total_tests = 0
    for i, (cat, fname, cnt, areas) in enumerate(TEST_COVERAGE, 2):
        fill = GREEN if i % 2 == 0 else WHITE
        write_row(ws3, i, [cat, fname, cnt, areas], fill_hex=fill)
        grand_total_tests += cnt

    last = len(TEST_COVERAGE) + 2
    write_row(ws3, last, ["TOTAL", "", grand_total_tests, ""], fill_hex=BLUE, bold=True)
    ws3.cell(last, 2).fill = _cell_fill(BLUE)

    set_col_widths(ws3, [18, 42, 12, 80])
    freeze(ws3, "A2")

    # â”€â”€ Sheet 4: Debugging Guide â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws4 = wb.create_sheet("Debugging Guide")
    ws4.sheet_view.showGridLines = False

    dbg_headers = ["Symptom", "Likely Cause", "Resolution Steps"]
    write_row(ws4, 1, dbg_headers, fill_hex=DARK, bold=True)
    for c in range(1, 4):
        ws4.cell(1, c).font = Font(bold=True, color=WHITE, size=10)
        ws4.cell(1, c).alignment = Alignment(horizontal="center", vertical="center")

    for i, (symptom, cause, resolution) in enumerate(DEBUGGING_SUMMARY, 2):
        fill = ORANGE if i % 2 == 0 else YELLOW
        write_row(ws4, i, [symptom, cause, resolution], fill_hex=fill)
        ws4.row_dimensions[i].height = 50

    set_col_widths(ws4, [40, 45, 70])
    freeze(ws4, "A2")

    return wb


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = build_workbook()
    wb.save(OUT_FILE)
    print(f"Bug report saved: {OUT_FILE}")


if __name__ == "__main__":
    main()
