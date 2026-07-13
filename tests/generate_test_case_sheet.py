"""
Generate an Excel workbook documenting every test case in the three
session test files. Reads the test files via AST (no execution), derives a
human-readable scenario + expected result per test, and writes:

  - "Summary"  sheet: counts per module + per category
  - one sheet per module with: ID, Category, Scenario, Preconditions,
    Steps, Expected Result, Type, Priority, Auth Role

Run:  python tests/generate_test_case_sheet.py
Output: tests/reports/TrintzERP_Test_Cases.xlsx
"""

import ast
import os
import re

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

FILES = [
    ("Credit Management", "tests/api/test_credit_management.py"),
    ("Invoice Cancellation", "tests/api/test_sales_cancel_professional.py"),
    ("Returns & GST Reporting", "tests/api/test_returns_gst_reporting.py"),
]

# Module → short ID prefix
PREFIX = {
    "Credit Management": "CM",
    "Invoice Cancellation": "INV-CAN",
    "Returns & GST Reporting": "RET-GST",
}

# Map class names to a readable category label
def category_for(cls: str) -> str:
    spaced = re.sub(r"(?<!^)(?=[A-Z])", " ", cls.replace("Test", "", 1))
    return spaced.strip() or cls


def humanize(name: str) -> str:
    """test_cancel_blocked_when_return_exists -> 'Cancel blocked when return exists'."""
    s = name[len("test_"):] if name.startswith("test_") else name
    s = s.replace("_", " ").strip()
    return s[:1].upper() + s[1:]


def expected_from_test(node: ast.FunctionDef) -> str:
    """
    Build an 'expected result' string from the test's asserts and any
    status_code checks, so the sheet reflects what the test actually verifies.
    """
    expectations = []
    status_codes = set()
    for n in ast.walk(node):
        # status_code == NNN  or  .status_code NNN
        if isinstance(n, ast.Compare):
            src = ast.dump(n)
            for code in re.findall(r"status_code", src):
                pass
        if isinstance(n, ast.Assert):
            try:
                txt = ast.unparse(n.test)
            except Exception:
                txt = ""
            # capture explicit status code assertions
            for m in re.findall(r"status_code\s*==\s*(\d{3})", txt):
                status_codes.add(m)
            for m in re.findall(r"\.status_code\b.*?(\d{3})", txt):
                status_codes.add(m)
            if "reconciled" in txt and "True" in txt:
                expectations.append("Account reconciled (balance matches ledger)")
            if "reconciled" in txt and "False" in txt:
                expectations.append("Account flagged as drifted (balance != ledger)")
            if "aging" in txt:
                expectations.append("Aging bucket value as expected")
            if "warning" in txt and "not in" in txt:
                expectations.append("No credit-linkage warning returned")
            elif "warning" in txt:
                expectations.append("Credit-linkage warning returned")
    if status_codes:
        codes = ", ".join(sorted(status_codes))
        expectations.insert(0, f"HTTP {codes}")
    if not expectations:
        # fall back to docstring first line
        doc = ast.get_docstring(node)
        if doc:
            expectations.append(doc.strip().splitlines()[0])
    # de-dup, keep order
    seen = []
    for e in expectations:
        if e not in seen:
            seen.append(e)
    return "; ".join(seen) or "Assertion(s) pass"


def auth_role(name: str, doc: str) -> str:
    blob = (name + " " + (doc or "")).lower()
    if "unauthenticated" in blob or "requires_auth" in blob or "require_auth" in blob:
        return "None (401 expected)"
    if "cashier_cannot" in blob or "requires_admin" in blob or "cannot_cancel" in blob:
        return "Cashier (403 expected)"
    if "cashier" in blob:
        return "Cashier"
    return "Admin"


def test_type(cls: str, name: str) -> str:
    blob = (cls + " " + name).lower()
    if "auth" in blob:
        return "Negative / Security"
    if any(w in blob for w in ["not_found", "blocked", "cannot", "over_return",
                                "already", "double", "drift", "mismatch", "400", "403", "401", "404", "409"]):
        return "Negative"
    return "Positive"


def priority(cls: str, name: str) -> str:
    blob = (cls + " " + name).lower()
    if any(w in blob for w in ["double_restore", "blocked", "reconcil", "net", "gst",
                                "inventory", "audit", "fifo"]):
        return "High"
    if "auth" in blob or "shape" in blob or "sort" in blob:
        return "Medium"
    return "Medium"


def collect():
    rows = []  # (module, id, category, scenario, precond, steps, expected, type, prio, auth)
    for module, relpath in FILES:
        path = os.path.join(ROOT, relpath)
        src = open(path, encoding="utf-8").read()
        tree = ast.parse(src)
        counter = 0
        for cls in [n for n in tree.body if isinstance(n, ast.ClassDef)]:
            cat = category_for(cls.name)
            for fn in [m for m in cls.body if isinstance(m, ast.FunctionDef)
                       and m.name.startswith("test_")]:
                counter += 1
                tid = f"{PREFIX[module]}-{counter:03d}"
                doc = ast.get_docstring(fn) or ""
                scenario = humanize(fn.name)
                # Preconditions inferred from fixtures used in the signature
                args = [a.arg for a in fn.args.args if a.arg != "self"]
                precond = []
                if "seeded_customer" in args:
                    precond.append("Credit customer seeded with controlled ledger")
                if "seeded_supplier" in args:
                    precond.append("Supplier seeded with controlled ledger")
                if "test_product" in args:
                    precond.append("Product with stock exists")
                if "cashier_headers" in args:
                    precond.append("Cashier token available")
                if not precond:
                    precond.append("Authenticated admin session")
                steps = doc.strip().splitlines()[0] if doc else scenario
                rows.append((
                    module, tid, cat, scenario,
                    "; ".join(precond), steps,
                    expected_from_test(fn),
                    test_type(cls.name, fn.name),
                    priority(cls.name, fn.name),
                    auth_role(fn.name, doc),
                ))
    return rows


def build_workbook(rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    hdr_fill = PatternFill("solid", fgColor="4F46E5")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    title_font = Font(bold=True, size=14, color="4F46E5")
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    type_fill = {
        "Positive": PatternFill("solid", fgColor="DCFCE7"),
        "Negative": PatternFill("solid", fgColor="FEF3C7"),
        "Negative / Security": PatternFill("solid", fgColor="FEE2E2"),
    }
    prio_fill = {
        "High": PatternFill("solid", fgColor="FECACA"),
        "Medium": PatternFill("solid", fgColor="FEF9C3"),
        "Low": PatternFill("solid", fgColor="E0E7FF"),
    }

    COLUMNS = ["Test ID", "Category", "Scenario", "Preconditions", "Test Steps / Intent",
               "Expected Result", "Type", "Priority", "Auth Role", "Status"]
    WIDTHS = [12, 22, 34, 30, 46, 40, 16, 9, 20, 10]

    # ── Per-module sheets ────────────────────────────────────────────────────
    modules = []
    for module, _ in FILES:
        modules.append(module)
        ws = wb.create_sheet(title=module[:31])
        ws["A1"] = f"TrintzERP — Test Cases: {module}"
        ws["A1"].font = title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

        # header row at row 3
        for c, (head, w) in enumerate(zip(COLUMNS, WIDTHS), start=1):
            cell = ws.cell(row=3, column=c, value=head)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(c)].width = w

        r = 4
        for row in [x for x in rows if x[0] == module]:
            (_, tid, cat, scenario, precond, steps, expected, ttype, prio, auth) = row
            values = [tid, cat, scenario, precond, steps, expected, ttype, prio, auth, "Not Run"]
            for c, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.alignment = wrap
                cell.border = border
                cell.font = Font(size=9)
            ws.cell(row=r, column=7).fill = type_fill.get(ttype, PatternFill())
            ws.cell(row=r, column=8).fill = prio_fill.get(prio, PatternFill())
            ws.cell(row=r, column=8).alignment = center
            r += 1
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(COLUMNS))}{r-1}"

    # ── Summary sheet (first) ────────────────────────────────────────────────
    summary = wb["Sheet"]
    summary.title = "Summary"
    summary["A1"] = "TrintzERP — Test Case Summary"
    summary["A1"].font = title_font
    summary.merge_cells("A1:D1")

    summary["A3"] = "Generated from the session's pytest suites (read-only export)."
    summary["A3"].font = Font(italic=True, size=9, color="6B7280")

    # counts per module
    summary["A5"] = "Module"; summary["B5"] = "Total"; summary["C5"] = "Positive"; summary["D5"] = "Negative"
    for c in range(1, 5):
        cell = summary.cell(row=5, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center; cell.border = border

    rr = 6
    grand = pos = neg = 0
    for module in modules:
        mod_rows = [x for x in rows if x[0] == module]
        p = sum(1 for x in mod_rows if x[7] == "Positive")
        n = len(mod_rows) - p
        summary.cell(row=rr, column=1, value=module).border = border
        summary.cell(row=rr, column=2, value=len(mod_rows)).border = border
        summary.cell(row=rr, column=3, value=p).border = border
        summary.cell(row=rr, column=4, value=n).border = border
        grand += len(mod_rows); pos += p; neg += n
        rr += 1
    # total row
    summary.cell(row=rr, column=1, value="TOTAL").font = Font(bold=True)
    summary.cell(row=rr, column=2, value=grand).font = Font(bold=True)
    summary.cell(row=rr, column=3, value=pos).font = Font(bold=True)
    summary.cell(row=rr, column=4, value=neg).font = Font(bold=True)
    for c in range(1, 5):
        summary.cell(row=rr, column=c).fill = PatternFill("solid", fgColor="EEF2FF")
        summary.cell(row=rr, column=c).border = border

    for col, w in zip("ABCD", [34, 12, 12, 12]):
        summary.column_dimensions[col].width = w

    # legend
    lr = rr + 3
    summary.cell(row=lr, column=1, value="Legend").font = Font(bold=True)
    summary.cell(row=lr + 1, column=1, value="Positive = happy-path / expected-success cases")
    summary.cell(row=lr + 2, column=1, value="Negative = error/validation/security cases (401/403/404/409/400)")
    summary.cell(row=lr + 3, column=1, value="Status column is for manual run tracking (Not Run / Pass / Fail)")

    # reorder: Summary first
    wb.move_sheet("Summary", -(len(wb.sheetnames) - 1))

    out_dir = os.path.join(ROOT, "tests", "reports")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "TrintzERP_Test_Cases.xlsx")
    wb.save(out_path)
    return out_path, grand


if __name__ == "__main__":
    rows = collect()
    path, total = build_workbook(rows)
    print(f"Wrote {total} test cases to: {path}")
